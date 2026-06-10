from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

import backtest_nxt31_utc7_latest as base
import backtest_nxt32_native_1d_latest as native


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "outputs" / "nxt33_ssl14_test"
OUT_JSON = OUT_DIR / "nxt33_ssl14_test_results.json"
OUT_XLSX = OUT_DIR / "NXT33_SSL10_vs_SSL14_Test.xlsx"


def enrich_with_ssl_period(candles: list[dict], ssl_period: int) -> list[dict]:
    enriched = base.enrich([dict(c) for c in candles])
    highs = [c["high"] for c in enriched]
    lows = [c["low"] for c in enriched]
    high_sma = base.sma(highs, ssl_period)
    low_sma = base.sma(lows, ssl_period)
    state = 0
    for i, c in enumerate(enriched):
        if high_sma[i] is None or low_sma[i] is None:
            c["ssl"] = None
            continue
        if c["close"] > high_sma[i]:
            state = 1
        elif c["close"] < low_sma[i]:
            state = -1
        c["ssl"] = state
    return enriched


def backtest_symbol(symbol: str, candles: list[dict], ssl_period: int) -> list[dict]:
    trades, pos, n = [], None, 1
    last_profitable_runner_exit = None
    for i in range(55, len(candles) - 1):
        c, prev, nxt = candles[i], candles[i - 1], candles[i + 1]
        next_date = base.date.fromisoformat(nxt["localDate"])
        if next_date < native.START_DATE or next_date >= native.END_DATE:
            continue
        if pos:
            side = pos["side"]
            ssl_flip = (side == "LONG" and prev["ssl"] == 1 and c["ssl"] == -1) or (side == "SHORT" and prev["ssl"] == -1 and c["ssl"] == 1)
            exit_price = reason = None
            if side == "LONG":
                if c["low"] <= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["high"] >= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["tp"] - pos["entry"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bearish flip"
            else:
                if c["high"] >= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["low"] <= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["entry"] - pos["tp"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bullish flip"
            if exit_price is not None:
                rem = 0.5 if pos["triggered"] else 1.0
                rem_r = (exit_price - pos["entry"]) / pos["risk"] if side == "LONG" else (pos["entry"] - exit_price) / pos["risk"]
                gross = pos["realizedR"] + rem * rem_r
                net = gross - base.cost_r(pos["entry"], pos["risk"])
                trades.append({
                    "variant": f"SSL {ssl_period}",
                    "symbol": symbol,
                    "tradeNo": n,
                    "side": side,
                    "signalTime": pos["signalDate"],
                    "entryTime": pos["entryDate"],
                    "entryPrice": pos["entry"],
                    "initialStop": pos["initialStop"],
                    "finalStop": pos["stop"],
                    "riskPerUnit": pos["risk"],
                    "tp1": pos["tp"],
                    "tp1Time": pos["tp1Time"],
                    "exitTime": c["localDate"],
                    "exitPrice": exit_price,
                    "exitReason": reason,
                    "grossRMultiple": gross,
                    "costR": base.cost_r(pos["entry"], pos["risk"]),
                    "rMultiple": net,
                    "atr14": pos["atr14"],
                    "rsi14": pos["rsi14"],
                    "distanceToEma50Atr": pos["distance"],
                    "notes": f"NXT v3.3 Runner A; SSL period {ssl_period}",
                })
                if net > 0 and reason.startswith("Runner exit"):
                    last_profitable_runner_exit = {"index": i, "side": side}
                n += 1
                pos = None
            if pos:
                continue

        if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]) or prev["ssl"] is None:
            continue
        dist = abs(c["close"] - c["ema50"]) / c["atr14"]
        long_ok = prev["ssl"] == -1 and c["ssl"] == 1 and base.recent_cross(candles, i, "LONG") and dist <= 2 and c["rsi14"] > 50
        short_ok = prev["ssl"] == 1 and c["ssl"] == -1 and base.recent_cross(candles, i, "SHORT") and dist <= 2 and c["rsi14"] < 50
        if last_profitable_runner_exit and i - last_profitable_runner_exit["index"] <= 1:
            if long_ok and last_profitable_runner_exit["side"] == "SHORT":
                long_ok = False
            if short_ok and last_profitable_runner_exit["side"] == "LONG":
                short_ok = False
        if not (long_ok or short_ok):
            continue
        side = "LONG" if long_ok else "SHORT"
        risk = c["atr14"] * 1.5
        entry = nxt["open"]
        pos = {
            "side": side,
            "signalDate": c["localDate"],
            "entryDate": nxt["localDate"],
            "entry": entry,
            "initialStop": entry - risk if side == "LONG" else entry + risk,
            "stop": entry - risk if side == "LONG" else entry + risk,
            "risk": risk,
            "tp": entry + c["atr14"] * 2.5 if side == "LONG" else entry - c["atr14"] * 2.5,
            "triggered": False,
            "tp1Time": "",
            "realizedR": 0.0,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": dist,
        }
    return trades


def profit_factor(rows: list[dict]) -> float | None:
    gross_profit = sum(t["rMultiple"] for t in rows if t["rMultiple"] > 0)
    gross_loss = -sum(t["rMultiple"] for t in rows if t["rMultiple"] < 0)
    return gross_profit / gross_loss if gross_loss else None


def side_stats(rows: list[dict]) -> dict:
    return {side: base.stats([t for t in rows if t["side"] == side]) for side in ["LONG", "SHORT"]}


def trade_key(t: dict) -> tuple:
    return (t["symbol"], t["side"], t["signalTime"], t["entryTime"])


def build_workbook(result: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "NXT v3.3 SSL 10 vs SSL 14 Test"
    ws["A2"] = "Only SSL Channel period changes; all other rules stay unchanged."
    headers = ["Variant", "Trades", "Win Rate", "Total R", "Delta R", "Avg R", "Max DD R", "Delta DD R", "Best R", "Worst R", "Profit Factor", "20K Ending"]
    ws.append([])
    ws.append(headers)
    for row in result["summary"]:
        ws.append([
            row["variant"], row["trades"], row["winRate"], row["totalR"], row["deltaTotalR"],
            row["avgR"], row["maxDrawdownR"], row["deltaMaxDrawdownR"], row["bestR"], row["worstR"],
            row["profitFactor"], row["ending20k"],
        ])

    trades_ws = wb.create_sheet("Trades SSL14")
    trade_headers = ["Symbol", "No", "Side", "Signal Date", "Entry Date", "Entry Price", "Initial Stop", "Final Stop", "Risk / Unit", "TP1", "TP1 Date", "Exit Date", "Exit Price", "Exit Reason", "R", "ATR14", "RSI14", "Distance EMA50 ATR", "Notes"]
    trades_ws.append(trade_headers)
    for t in result["variants"]["SSL 14"]["trades"]:
        trades_ws.append([t["symbol"].replace("USDT", ""), t["tradeNo"], t["side"], t["signalTime"], t["entryTime"], t["entryPrice"], t["initialStop"], t["finalStop"], t["riskPerUnit"], t["tp1"], t["tp1Time"], t["exitTime"], t["exitPrice"], t["exitReason"], t["rMultiple"], t["atr14"], t["rsi14"], t["distanceToEma50Atr"], t["notes"]])

    diff_ws = wb.create_sheet("Changed Trades")
    diff_ws.append(["Change", *trade_headers])
    for label, rows in [("Removed from SSL10", result["removedFromBaseline"]), ("Added by SSL14", result["addedBySSL14"])]:
        for t in rows:
            diff_ws.append([label, t["symbol"].replace("USDT", ""), t["tradeNo"], t["side"], t["signalTime"], t["entryTime"], t["entryPrice"], t["initialStop"], t["finalStop"], t["riskPerUnit"], t["tp1"], t["tp1Time"], t["exitTime"], t["exitPrice"], t["exitReason"], t["rMultiple"], t["atr14"], t["rsi14"], t["distanceToEma50Atr"], t["notes"]])

    side_ws = wb.create_sheet("By Side")
    side_ws.append(["Variant", "Side", "Trades", "Win Rate", "Total R", "Avg R", "Max DD R", "Best R", "Worst R"])
    for variant, payload in result["variants"].items():
        for side, stats in payload["sideStats"].items():
            side_ws.append([variant, side, stats["trades"], stats["winRate"], stats["totalR"], stats["avgR"], stats["maxDrawdownR"], stats["bestR"], stats["worstR"]])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A4" if sheet.title == "Summary" else "A2"
        for cell in sheet[1 if sheet.title != "Summary" else 4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 15
    ws.column_dimensions["A"].width = 22
    trades_ws.column_dimensions["S"].width = 34
    diff_ws.column_dimensions["A"].width = 20
    diff_ws.column_dimensions["T"].width = 34
    wb.save(OUT_XLSX)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    raw_by_symbol = {symbol: native.fetch_native_1d(symbol) for symbol in native.SYMBOLS}
    variants = {}
    baseline_stats = None
    summary = []
    for ssl_period in [10, 14]:
        rows = []
        for symbol, raw in raw_by_symbol.items():
            rows.extend(backtest_symbol(symbol, enrich_with_ssl_period(raw, ssl_period), ssl_period))
        rows.sort(key=lambda x: x["exitTime"])
        stats = base.stats(rows)
        stats["variant"] = f"SSL {ssl_period}"
        stats["profitFactor"] = profit_factor(rows)
        stats["ending20k"] = 20000 + stats["totalR"] * 1000
        if baseline_stats is None:
            baseline_stats = stats
        stats["deltaTotalR"] = stats["totalR"] - baseline_stats["totalR"]
        stats["deltaMaxDrawdownR"] = stats["maxDrawdownR"] - baseline_stats["maxDrawdownR"]
        summary.append(stats)
        variants[f"SSL {ssl_period}"] = {"stats": stats, "sideStats": side_stats(rows), "trades": rows}

    baseline_keys = {trade_key(t): t for t in variants["SSL 10"]["trades"]}
    ssl14_keys = {trade_key(t): t for t in variants["SSL 14"]["trades"]}
    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "systemVersion": "NXT v3.3 SSL 10 vs SSL 14 test",
        "summary": summary,
        "variants": variants,
        "removedFromBaseline": [baseline_keys[k] for k in sorted(set(baseline_keys) - set(ssl14_keys))],
        "addedBySSL14": [ssl14_keys[k] for k in sorted(set(ssl14_keys) - set(baseline_keys))],
        "assumptions": [
            "Baseline uses SSL Channel SMA high/low period 10.",
            "Test variant uses SSL Channel SMA high/low period 14.",
            "All other NXT v3.3 rules are unchanged: native 1D, Runner A, anti-immediate-reversal, no continuation, no risk-off.",
        ],
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    build_workbook(result)
    print(json.dumps({"outJson": str(OUT_JSON), "outXlsx": str(OUT_XLSX), "summary": summary}, indent=2))


if __name__ == "__main__":
    main()
