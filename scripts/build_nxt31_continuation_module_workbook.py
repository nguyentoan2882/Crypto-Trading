from __future__ import annotations

import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\Workspace\Codex\Crypto trading")
SRC = ROOT / "outputs" / "nxt31_continuation_module_6y" / "nxt31_continuation_module_6y_results.json"
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
OUT = ROOT / "outputs" / "nxt31_continuation_module_6y" / "NXT31_Continuation_Module_6Y_BTC_SOL_SUI_20K.xlsx"


def clear(ws, rows=320, cols=28):
    for row in ws.iter_rows(min_row=1, max_row=max(rows, ws.max_row), min_col=1, max_col=max(cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def rowset(ws, r, values):
    for c, v in enumerate(values, 1):
        ws.cell(r, c).value = v


def copy_layout(src, dst):
    dst.sheet_format.defaultColWidth = src.sheet_format.defaultColWidth
    dst.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    for c in range(1, max(src.max_column, dst.max_column) + 1):
        letter = get_column_letter(c)
        if src.column_dimensions.get(letter) and src.column_dimensions[letter].width:
            dst.column_dimensions[letter].width = src.column_dimensions[letter].width
    for r in range(1, max(src.max_row, dst.max_row) + 1):
        sr = r if r <= src.max_row else 5
        if src.row_dimensions.get(sr) and src.row_dimensions[sr].height:
            dst.row_dimensions[r].height = src.row_dimensions[sr].height
        for c in range(1, max(src.max_column, dst.max_column) + 1):
            sc = min(c, src.max_column)
            a = src.cell(sr, sc)
            b = dst.cell(r, c)
            if a.has_style:
                b._style = copy(a._style)
            b.number_format = a.number_format
            b.alignment = copy(a.alignment)
            b.font = copy(a.font)
            b.fill = copy(a.fill)
            b.border = copy(a.border)


def main():
    data = json.loads(SRC.read_text(encoding="utf-8"))
    tpl = load_workbook(TEMPLATE)
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        clear(ws)

    ws = wb["Summary"]
    ws["A1"] = "NXT v3.1 Continuation Module Test"
    ws["A2"] = "Baseline Runner A unchanged; continuation module adds non-overlapping continuation entries."
    rowset(ws, 4, ["Variant", "Trades", "Win Rate", "Total R After Risk-Off", "Avg R", "Max DD R", "Continuation Trades", "Continuation R", "Captures 13-Dec-2020"])
    for i, r in enumerate(data["rankingAfterRiskOff"], 5):
        rowset(ws, i, [r["name"], r["trades"], r["winRate"], r["totalR"], r["avgR"], r["maxDrawdownR"], r["continuationTrades"], r["continuationR"], "Yes" if r["capturesDec2020"] else "No"])

    headers = ["Variant", "Symbol", "No", "Side", "Signal Time", "Entry Time", "Entry Price", "TP1", "TP1 Time", "Exit Time", "Exit Price", "Exit Reason", "Net R", "ATR14", "RSI14", "Distance EMA50 ATR", "Notes"]
    all_cont = []
    for key, variant in data["variants"].items():
        for t in variant["continuationTrades"]:
            all_cont.append((variant["name"], t))
    all_cont.sort(key=lambda x: x[1]["exitTime"])
    ws = wb["Trades"]
    ws["A1"] = "Continuation Trades"
    ws["A2"] = "All continuation candidates by variant."
    rowset(ws, 4, headers)
    for i, (name, t) in enumerate(all_cont, 5):
        rowset(ws, i, [name, t["symbol"].replace("USDT", ""), t["tradeNo"], t["side"], t["signalTime"], t["entryTime"], t["entryPrice"], t["tp1"], t.get("tp1Time", ""), t["exitTime"], t["exitPrice"], t["exitReason"], t["rMultiple"], t["atr14"], t["rsi14"], t["distanceToEma50Atr"], t["notes"]])

    for sheet_name in ["Summary", "Trades", "BTC", "SOL", "SUI", "Assumptions", "Data Quality", "20K Account", "Equity Curve"]:
        if sheet_name in wb.sheetnames and sheet_name in tpl.sheetnames:
            copy_layout(tpl[sheet_name], wb[sheet_name])
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
