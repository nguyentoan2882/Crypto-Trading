from __future__ import annotations

import json
import os
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
LATEST_JSON = Path(os.environ.get(
    "NXT_LATEST_JSON",
    ROOT / "latest" / "NXT_Latest_NXT35_USDM_BlockShortAfterLosingLong_FundingAdjusted_20K.json",
))
LATEST_XLSX = Path(os.environ.get(
    "NXT_LATEST_FUNDING_XLSX",
    ROOT / "latest" / "NXT_Latest_NXT35_USDM_BlockShortAfterLosingLong_FundingAdjusted_20K.xlsx",
))
ONE_R_DOLLARS = 1_000
STARTING_EQUITY = 20_000.0
EQUAL_CAP_ALLOCATION = {"BTCUSDT": 0.02, "BNBUSDT": 0.02, "SOLUSDT": 0.02}
SHEET_NAME = "Order Plan"


def side_values(side: str) -> dict:
    if side == "LONG":
        return {
            "entrySide": "BUY",
            "closeSide": "SELL",
            "stopOrderSide": "SELL",
            "takeProfitSide": "SELL",
            "positionSide": "LONG",
            "stopDirection": "below entry",
            "tpDirection": "above entry",
        }
    return {
        "entrySide": "SELL",
        "closeSide": "BUY",
        "stopOrderSide": "BUY",
        "takeProfitSide": "BUY",
        "positionSide": "SHORT",
        "stopDirection": "above entry",
        "tpDirection": "below entry",
    }


def qty_for_trade(trade: dict) -> float:
    return ONE_R_DOLLARS / trade["riskPerUnit"]


def simulate_sizing(trades: list[dict]) -> dict[str, dict]:
    events = defaultdict(lambda: {"entries": [], "exits": []})
    for index, trade in enumerate(trades):
        events[trade["entryTime"]]["entries"].append((index, trade))
        events[trade["exitTime"]]["exits"].append((index, trade))

    equity = STARTING_EQUITY
    fixed_equity = STARTING_EQUITY
    open_risk: dict[int, dict] = {}
    trade_plan: dict[str, dict] = {}

    def trade_key(trade: dict) -> str:
        return f"{trade['symbol']}-{trade['tradeNo']:03d}-{trade['signalTime']}"

    def current_open_risk(symbol: str | None = None) -> float:
        return sum(
            row["riskAmount"]
            for row in open_risk.values()
            if symbol is None or row["symbol"] == symbol
        )

    def close_trade(index: int, trade: dict) -> None:
        nonlocal equity, fixed_equity
        risk_record = open_risk.pop(index, None)
        risk_amount = risk_record["riskAmount"] if risk_record else 0.0
        pnl = risk_amount * float(trade["netRAfterFunding"])
        equity += pnl
        fixed_pnl = ONE_R_DOLLARS * float(trade["netRAfterFunding"])
        fixed_equity += fixed_pnl
        rec = trade_plan.setdefault(trade_key(trade), {})
        rec.update(
            {
                "fixedRisk": ONE_R_DOLLARS,
                "fixedPnl": fixed_pnl,
                "fixedEquityAfterExit": fixed_equity,
                "capEqualRisk": risk_amount,
                "capEqualPnl": pnl,
                "capEqualEquityAfterExit": equity,
            }
        )

    for event_date in sorted(events):
        same_day_indexes = {
            index
            for index, trade in events[event_date]["entries"]
            if trade["exitTime"] == event_date
        }
        regular_exits = [
            item for item in events[event_date]["exits"]
            if item[0] not in same_day_indexes
        ]
        same_day_exits = [
            item for item in events[event_date]["exits"]
            if item[0] in same_day_indexes
        ]

        for index, trade in sorted(regular_exits, key=lambda item: (item[1]["symbol"], item[1]["tradeNo"])):
            close_trade(index, trade)
        for index, trade in sorted(events[event_date]["entries"], key=lambda item: (item[1]["symbol"], item[1]["tradeNo"])):
            symbol = trade["symbol"]
            symbol_capacity = max(0.0, equity * EQUAL_CAP_ALLOCATION[symbol] - current_open_risk(symbol))
            portfolio_capacity = max(0.0, equity * sum(EQUAL_CAP_ALLOCATION.values()) - current_open_risk())
            risk_amount = min(symbol_capacity, portfolio_capacity)
            trade_plan.setdefault(trade_key(trade), {})["capEqualRisk"] = risk_amount
            if risk_amount > 0:
                open_risk[index] = {"symbol": symbol, "riskAmount": risk_amount}
        for index, trade in sorted(same_day_exits, key=lambda item: (item[1]["symbol"], item[1]["tradeNo"])):
            close_trade(index, trade)

    return trade_plan


def scenario_values(trade: dict, qty_fraction: float, ref_price: float | str, plan: dict, show_trade_result: bool = False) -> list:
    price = ref_price if isinstance(ref_price, (int, float)) else None
    fixed_risk = plan.get("fixedRisk", ONE_R_DOLLARS)
    cap_risk = plan.get("capEqualRisk", 0.0)
    fixed_qty = fixed_risk / trade["riskPerUnit"] * qty_fraction
    cap_qty = cap_risk / trade["riskPerUnit"] * qty_fraction if cap_risk else 0.0
    return [
        fixed_risk,
        fixed_qty,
        fixed_qty * price if price else "",
        plan.get("fixedPnl", "") if show_trade_result else "",
        plan.get("fixedEquityAfterExit", "") if show_trade_result else "",
        cap_risk,
        cap_qty,
        cap_qty * price if price else "",
        plan.get("capEqualPnl", "") if show_trade_result else "",
        plan.get("capEqualEquityAfterExit", "") if show_trade_result else "",
    ]


def order_rows_for_trade(global_no: int, trade: dict, plan_by_key: dict[str, dict], runner_rule: dict | None = None) -> list[list]:
    vals = side_values(trade["side"])
    qty = qty_for_trade(trade)
    runner_rule = runner_rule or {}
    tp1_fraction = float(runner_rule.get("tp1Fraction", 0.5))
    partial_fraction = float(runner_rule.get("partialFraction", 0.0))
    partial_at_r = float(runner_rule.get("partialAtR", 0.0))
    tail_fraction = float(runner_rule.get("tailFraction", 1.0 - tp1_fraction - partial_fraction))
    tp1_qty = qty * tp1_fraction
    partial_qty = qty * partial_fraction
    tail_qty = qty * tail_fraction
    partial_price = trade["entryPrice"] + trade["riskPerUnit"] * partial_at_r if trade["side"] == "LONG" else trade["entryPrice"] - trade["riskPerUnit"] * partial_at_r
    symbol = trade["symbol"]
    key = f"{symbol}-{trade['tradeNo']:03d}-{trade['signalTime']}"
    plan = plan_by_key.get(key, {})
    common = [
        global_no,
        key,
        symbol,
        trade["tradeNo"],
        trade["signalType"],
        trade["side"],
        vals["positionSide"],
        trade["signalTime"],
    ]
    rows = []
    rows.append(common + [
        1,
        trade["entryTime"],
        "OPEN_POSITION",
        vals["entrySide"],
        "MARKET_OR_LIMIT_AT_OPEN",
        "OPEN",
        qty,
        1.0,
        trade["entryPrice"],
        "",
        "",
        "Open full position at next daily open after signal close.",
        "Submit only if signal remains valid at daily close; use isolated/cross setting per account policy.",
    ] + scenario_values(trade, 1.0, trade["entryPrice"], plan))
    rows.append(common + [
        2,
        trade["entryTime"],
        "PLACE_INITIAL_STOP",
        vals["stopOrderSide"],
        "STOP_MARKET reduceOnly",
        "PROTECT",
        qty,
        1.0,
        "",
        trade["initialStop"],
        "",
        f"Protect full position; stop is 1.5 ATR {vals['stopDirection']}.",
        "Must be reduceOnly. Cancel/replace after TP1 if TP1 fills.",
    ] + scenario_values(trade, 1.0, trade["initialStop"], plan))
    rows.append(common + [
        3,
        trade["entryTime"],
        "PLACE_TP1",
        vals["takeProfitSide"],
        "TAKE_PROFIT_MARKET reduceOnly",
        "PARTIAL_CLOSE",
        tp1_qty,
        tp1_fraction,
        "",
        trade["tp1"],
        "",
        f"Close {tp1_fraction:.0%} at TP1; target is 2.5 ATR {vals['tpDirection']}.",
        "If TP1 is not hit, this order remains pending until final exit/stop.",
    ] + scenario_values(trade, tp1_fraction, trade["tp1"], plan))
    next_order_no = 4
    if trade.get("earlyBeTriggered"):
        rows.append(common + [
            next_order_no,
            f"{trade.get('earlyBeTime', '')} + next D1",
            "EARLY_BE_7PCT_MOVE_STOP_TO_ENTRY",
            vals["stopOrderSide"],
            "STOP_MARKET reduceOnly",
            "PROTECT_FULL_POSITION",
            qty,
            1.0,
            "",
            trade["entryPrice"],
            "",
            "A 7% favorable High/Low move occurred before TP1 on a daily candle after entry; cancel initial stop and move full-position stop to entry from the next daily candle.",
            "LONG High trigger = Entry x 1.07; SHORT Low trigger = Entry x 0.93.",
        ] + scenario_values(trade, 1.0, trade["entryPrice"], plan))
        next_order_no += 1
    if trade.get("tp1Time"):
        rows.append(common + [
            next_order_no,
            trade["tp1Time"],
            f"TP1_FILLED_CLOSE_{tp1_fraction:.0%}",
            vals["closeSide"],
            "FILLED_TP1 reduceOnly",
            "PARTIAL_CLOSE",
            tp1_qty,
            tp1_fraction,
            trade["tp1"],
            "",
            "",
            f"TP1 filled; {tp1_fraction:.0%} position closed.",
            f"Realized gross +{tp1_fraction * 2.5:.4f}R before trading cost/funding.",
        ] + scenario_values(trade, tp1_fraction, trade["tp1"], plan))
        next_order_no += 1
        rows.append(common + [
            next_order_no,
            trade["tp1Time"],
            "MOVE_STOP_TO_BREAKEVEN",
            vals["stopOrderSide"],
            "STOP_MARKET reduceOnly",
            "PROTECT_RUNNER",
            tail_qty + partial_qty,
            tail_fraction + partial_fraction,
            "",
            trade["entryPrice"],
            "",
            f"Cancel initial stop and replace with breakeven stop for remaining {tail_fraction + partial_fraction:.0%}.",
            "This is the runner risk-control step after TP1.",
        ] + scenario_values(trade, tail_fraction + partial_fraction, trade["entryPrice"], plan))
        next_order_no += 1
        if partial_fraction:
            rows.append(common + [
                next_order_no, trade.get("partialTime") or trade["entryTime"], f"PARTIAL_CLOSE_{partial_fraction:.0%}_AT_{partial_at_r:.1f}R", vals["takeProfitSide"], "TAKE_PROFIT_MARKET reduceOnly", "PARTIAL_CLOSE", partial_qty, partial_fraction, partial_price, "", "", f"Close {partial_fraction:.0%} at {partial_at_r:.1f}R; then leave {tail_fraction:.0%} tail.", "Order remains pending if final exit occurs first.",
            ] + scenario_values(trade, partial_fraction, partial_price, plan))
        final_order_no = next_order_no + 1
        final_qty = tail_qty if trade.get("partialTime") else tail_qty + partial_qty
        final_fraction = tail_fraction if trade.get("partialTime") else tail_fraction + partial_fraction
    else:
        final_order_no = next_order_no
        final_qty = qty
        final_fraction = 1.0

    final_action = "STOP_LOSS_OR_RUNNER_EXIT"
    if trade["exitReason"] == "Stop loss":
        final_action = "STOP_LOSS_CLOSE"
    elif trade["exitReason"] == "Breakeven stop":
        final_action = "BREAKEVEN_STOP_CLOSE_RUNNER" if trade.get("tp1Time") else "BREAKEVEN_STOP_CLOSE_FULL"
    elif trade["exitReason"].startswith("Runner exit"):
        final_action = "RUNNER_EXIT_ON_EMA50" if "EMA50" in trade["exitReason"] else "RUNNER_EXIT_ON_SSL_FLIP"

    rows.append(common + [
        final_order_no,
        trade["exitTime"],
        final_action,
        vals["closeSide"],
        "MARKET_OR_STOP reduceOnly",
        "FINAL_CLOSE",
        final_qty,
        final_fraction,
        trade["exitPrice"],
        trade["finalStop"] if "stop" in trade["exitReason"].lower() else "",
        "",
        trade["exitReason"],
        f"Trade net after trading cost and funding: {trade['netRAfterFunding']:.4f}R.",
    ] + scenario_values(trade, final_fraction, trade["exitPrice"], plan, show_trade_result=True))
    return rows


def style_sheet(ws) -> None:
    ws.freeze_panes = "A5"
    for cell in ws[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    widths = {
        "A": 8, "B": 28, "C": 12, "D": 8, "E": 14, "F": 10, "G": 12, "H": 12,
        "I": 8, "J": 12, "K": 26, "L": 10, "M": 24, "N": 16, "O": 14,
        "P": 12, "Q": 14, "R": 14, "S": 12, "T": 44, "U": 52,
        "V": 14, "W": 14, "X": 16, "Y": 16, "Z": 18, "AA": 14, "AB": 14,
        "AC": 16, "AD": 16, "AE": 18,
    }
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(letter, 16)
    for row in ws.iter_rows(min_row=5, min_col=15, max_col=18):
        for cell in row:
            cell.number_format = "0.000000"
    for row in ws.iter_rows(min_row=5, min_col=22, max_col=31):
        for cell in row:
            if cell.column in {23, 28}:
                cell.number_format = "0.000000"
            else:
                cell.number_format = '"$"#,##0.00'
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    latest = json.loads(LATEST_JSON.read_text(encoding="utf-8"))
    plan_by_key = simulate_sizing(latest["trades"])
    wb = load_workbook(LATEST_XLSX)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME, 1)
    ws["A1"] = "Binance Futures Order-Level Plan"
    ws["A2"] = "Assumptions: scenario 1 fixes risk at $1,000/trade; scenario 2 caps total open risk at 6% with BTC/BNB/SOL each capped at 2% of current realized equity. Order value = scenario quantity x reference price."
    headers = [
        "Trade #", "Trade Key", "Symbol", "Symbol Trade #", "Signal Type", "Trade Side", "Position Side",
        "Signal Date", "Order #", "Action Date", "Action", "Order Side", "Order Type", "Purpose",
        "Qty Base", "Position Fraction", "Price", "Trigger/Stop", "Leverage Note", "Rule / Reason", "Control Note",
        "Fixed Risk $", "Fixed Qty", "Fixed Order Value $", "Fixed Trade P&L $", "Fixed Equity After Exit $",
        "Cap 6% Eq Risk $", "Cap 6% Eq Qty", "Cap 6% Eq Order Value $", "Cap 6% Eq Trade P&L $",
        "Cap 6% Eq Equity After Exit $",
    ]
    for col, value in enumerate(headers, 1):
        ws.cell(4, col).value = value
    row_no = 5
    for i, trade in enumerate(latest["trades"], 1):
        for row in order_rows_for_trade(i, trade, plan_by_key, latest.get("runnerRule")):
            for col, value in enumerate(row, 1):
                ws.cell(row_no, col).value = value
            row_no += 1
    style_sheet(ws)
    for sheet in wb.worksheets:
        if sheet.freeze_panes is None:
            for selection in sheet.sheet_view.selection:
                selection.pane = None
    wb.save(LATEST_XLSX)
    print(json.dumps({
        "workbook": str(LATEST_XLSX),
        "sheet": SHEET_NAME,
        "trades": len(latest["trades"]),
        "orderRows": row_no - 5,
    }, indent=2))


if __name__ == "__main__":
    main()
