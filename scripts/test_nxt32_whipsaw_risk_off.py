from __future__ import annotations

import json
import sys
from copy import deepcopy
from datetime import date, datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

import backtest_nxt31_utc7_latest as base
import backtest_nxt32_native_1d_latest as native
import rebuild_nxt32_native_1d_tv_atr_latest as tv_atr
from test_nxt32_whipsaw_regime_score import (
    WHIPSAW_WINDOWS,
    whipsaw_features,
    whipsaw_score,
    window_breakdown,
)


ROOT = Path.cwd()
OUT_DIR = ROOT / "outputs" / "nxt32_whipsaw_risk_off"
OUT_JSON = OUT_DIR / "nxt32_whipsaw_risk_off_results.json"
OUT_XLSX = OUT_DIR / "NXT32_Whipsaw_Risk_Off_Best_6Y_BTC_SOL_SUI_20K.xlsx"
BASELINE_JSON = ROOT / "latest" / "NXT_Latest_NXT32_Native1D_RunnerA_NoContinuation_NoRiskOff_6Y_BTC_SOL_SUI_20K.json"
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"


def index_candles(candles_by_symbol: dict[str, list[dict]]) -> dict[str, dict[str, int]]:
    return {
        symbol: {c["localDate"]: i for i, c in enumerate(candles)}
        for symbol, candles in candles_by_symbol.items()
    }


def apply_risk_off(trades: list[dict], candles_by_symbol: dict[str, list[dict]], date_index: dict[str, dict[str, int]], cfg: dict) -> dict:
    adjusted = []
    risk_off_events = []
    for trade in trades:
        t = deepcopy(trade)
        symbol = t["symbol"]
        signal_date = t["signalTime"]
        i = date_index[symbol][signal_date]
        features = whipsaw_features(candles_by_symbol[symbol], i, cfg)
        score, reasons = whipsaw_score(features, cfg)
        multiplier = cfg["riskOffMultiplier"] if score >= cfg["riskOffScore"] else 1.0
        t["originalRMultiple"] = t["rMultiple"]
        t["riskMultiplier"] = multiplier
        t["rMultiple"] = t["rMultiple"] * multiplier
        t["whipsawScore"] = score
        t["whipsawReasons"] = ", ".join(reasons)
        t["efficiencyRatio"] = features["er"]
        t["medianDistanceToEma50Atr"] = features["medianDistanceToEma50Atr"]
        t["ema50SlopeAtr"] = features["ema50SlopeAtr"]
        t["sslFlipCount"] = features["sslFlipCount"]
        t["notes"] = (
            f"Whipsaw risk-off test; raw R={t['originalRMultiple']:.4f}; "
            f"risk multiplier={multiplier:.2f}; score={score}; reasons={', '.join(reasons)}"
        )
        adjusted.append(t)
        if multiplier != 1.0:
            risk_off_events.append(
                {
                    "symbol": symbol,
                    "signalTime": t["signalTime"],
                    "entryTime": t["entryTime"],
                    "side": t["side"],
                    "rawR": t["originalRMultiple"],
                    "adjustedR": t["rMultiple"],
                    "riskMultiplier": multiplier,
                    "score": score,
                    "reasons": reasons,
                    "features": features,
                }
            )
    return {
        "trades": adjusted,
        "stats": base.stats(adjusted),
        "whipsawWindows": window_breakdown(adjusted),
        "riskOffEvents": risk_off_events,
    }


def fix_workbook_labels(path: Path, result: dict) -> None:
    wb = load_workbook(path)
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        ws["A1"] = "NXT v3.2 Latest Test - Whipsaw Risk-Off"
        ws["A2"] = f"BTC/SOL/SUI 6Y | {result['name']} | risk-off scales R in whipsaw regime; no entry skip."
    if "Trades" in wb.sheetnames:
        wb["Trades"]["A1"] = "Detailed Trades - NXT Whipsaw Risk-Off Test"
    for sheet in ["BTC", "SOL", "SUI"]:
        if sheet in wb.sheetnames:
            wb[sheet]["A1"] = f"{sheet} - NXT Whipsaw Risk-Off Test"
    wb.save(path)


def write_compact_workbook(path: Path, best: dict, baseline: dict, datasets: dict) -> None:
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=max(ws.max_row, 350), min_col=1, max_col=max(ws.max_column, 24)):
            for cell in row:
                if cell.__class__.__name__ != "MergedCell":
                    cell.value = None

    def rowset(ws, r, values):
        for c, v in enumerate(values, 1):
            ws.cell(r, c).value = v

    ws = wb["Summary"]
    ws["A1"] = "NXT v3.2 Latest Test - Whipsaw Risk-Off"
    ws["A2"] = best["name"]
    rows = [
        ["Metric", "Latest", "Risk-Off"],
        ["Trades", baseline["stats"]["trades"], best["stats"]["trades"]],
        ["Win Rate", baseline["stats"]["winRate"], best["stats"]["winRate"]],
        ["Total R", baseline["stats"]["totalR"], best["stats"]["totalR"]],
        ["Average R", baseline["stats"]["avgR"], best["stats"]["avgR"]],
        ["Max DD R", baseline["stats"]["maxDrawdownR"], best["stats"]["maxDrawdownR"]],
        ["Whipsaw Combined R", window_breakdown(baseline["trades"])[-1]["totalR"], best["whipsawWindows"][-1]["totalR"]],
        ["Risk-Off Trades", 0, len(best["riskOffEvents"])],
    ]
    for i, row in enumerate(rows, 4):
        rowset(ws, i, row)

    headers = [
        "Symbol",
        "No",
        "Side",
        "Signal Date",
        "Entry Date",
        "Exit Date",
        "Raw R",
        "Risk Mult",
        "Adjusted R",
        "Whipsaw Score",
        "Whipsaw Reasons",
        "ER20",
        "Median EMA50 Dist",
        "EMA50 Slope",
        "SSL Flips",
        "Exit Reason",
    ]
    subsets = {
        "Trades": best["trades"],
        "BTC": [t for t in best["trades"] if t["symbol"] == "BTCUSDT"],
        "SOL": [t for t in best["trades"] if t["symbol"] == "SOLUSDT"],
        "SUI": [t for t in best["trades"] if t["symbol"] == "SUIUSDT"],
    }
    for sheet, trades in subsets.items():
        ws = wb[sheet]
        ws["A1"] = f"{sheet} - NXT Whipsaw Risk-Off Test"
        ws["A2"] = "Adjusted R reflects risk multiplier in whipsaw regime."
        rowset(ws, 4, headers)
        for i, t in enumerate(trades, 5):
            rowset(
                ws,
                i,
                [
                    t["symbol"].replace("USDT", ""),
                    t["tradeNo"],
                    t["side"],
                    t["signalTime"],
                    t["entryTime"],
                    t["exitTime"],
                    t["originalRMultiple"],
                    t["riskMultiplier"],
                    t["rMultiple"],
                    t["whipsawScore"],
                    t["whipsawReasons"],
                    t["efficiencyRatio"],
                    t["medianDistanceToEma50Atr"],
                    t["ema50SlopeAtr"],
                    t["sslFlipCount"],
                    t["exitReason"],
                ],
            )

    ws = wb["20K Account"]
    ws["A1"] = "20K Account"
    rowset(ws, 4, ["Starting Equity", "Latest Total R", "Risk-Off Total R", "Risk-Off Ending Equity"])
    rowset(ws, 5, [20000, baseline["stats"]["totalR"], best["stats"]["totalR"], 20000 + best["stats"]["totalR"] * 1000])

    ws = wb["Assumptions"]
    ws["A1"] = "Assumptions"
    rowset(ws, 4, ["#", "Assumption"])
    assumptions = [
        "Baseline is current latest NXT v3.2 with Binance native 1D candles, TV ATR RMA, Runner A, 20% reversal-skip, no continuation, and no risk-off.",
        "This test does not skip entries; it scales R by the risk multiplier when pre-entry whipsaw score reaches the threshold.",
        f"Best rule: {best['name']}.",
        "This workbook is a test artifact and is not promoted to latest.",
    ]
    for i, text in enumerate(assumptions, 5):
        rowset(ws, i, [i - 4, text])

    ws = wb["Data Quality"]
    ws["A1"] = "Data Quality"
    rowset(ws, 4, ["Symbol", "Daily Rows", "First Day", "Last Day", "Source"])
    for i, (symbol, q) in enumerate(datasets.items(), 5):
        rowset(ws, i, [symbol.replace("USDT", ""), q["dailyRows"], q["firstDay"], q["lastDay"], q["source"]])

    wb.save(path)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    baseline = json.loads(BASELINE_JSON.read_text(encoding="utf-8"))
    candles_by_symbol = {symbol: tv_atr.enrich_tv_atr(native.fetch_native_1d(symbol)) for symbol in base.SYMBOLS}
    date_index = index_candles(candles_by_symbol)
    datasets = {
        symbol: {
            "dailyRows": len(candles),
            "firstDay": candles[0]["localDate"],
            "lastDay": candles[-1]["localDate"],
            "source": "Binance spot native 1D klines",
        }
        for symbol, candles in candles_by_symbol.items()
    }

    base_cfg = {
        "flipLookback": 20,
        "erLookback": 20,
        "compressionLookback": 10,
        "slopeLookback": 20,
        "minEntryDistance": 0.0,
    }
    variants = []
    for max_flips in [3, 4, 5]:
        for max_er in [0.15, 0.20, 0.25, 0.30]:
            for max_median_dist in [0.45, 0.60, 0.75]:
                for max_slope in [0.40, 0.60, 0.80]:
                    for risk_off_score in [2, 3, 4]:
                        for multiplier in [0.25, 0.50, 0.75]:
                            cfg = deepcopy(base_cfg)
                            cfg.update(
                                {
                                    "name": f"risk{multiplier:g}_score{risk_off_score}_flip{max_flips}_er{max_er:g}_dist{max_median_dist:g}_slope{max_slope:g}",
                                    "maxFlips": max_flips,
                                    "maxEr": max_er,
                                    "maxMedianDistance": max_median_dist,
                                    "maxEma50Slope": max_slope,
                                    "riskOffScore": risk_off_score,
                                    "riskOffMultiplier": multiplier,
                                }
                            )
                            variants.append(cfg)

    baseline_whipsaw = window_breakdown(baseline["trades"])
    baseline_combined = baseline_whipsaw[-1]["totalR"]
    results = []
    for cfg in variants:
        result = apply_risk_off(baseline["trades"], candles_by_symbol, date_index, cfg)
        stats = result["stats"]
        combined = result["whipsawWindows"][-1]["totalR"]
        results.append(
            {
                "name": cfg["name"],
                "config": cfg,
                "stats": stats,
                "whipsawWindows": result["whipsawWindows"],
                "riskOffEvents": result["riskOffEvents"],
                "trades": result["trades"],
                "summary": {
                    "name": cfg["name"],
                    "trades": stats["trades"],
                    "totalR": stats["totalR"],
                    "maxDrawdownR": stats["maxDrawdownR"],
                    "winRate": stats["winRate"],
                    "whipsawCombinedR": combined,
                    "whipsawImprovementR": combined - baseline_combined,
                    "riskOffTrades": len(result["riskOffEvents"]),
                },
            }
        )

    summaries = [r["summary"] for r in results]
    candidates = [
        s
        for s in summaries
        if s["totalR"] >= baseline["stats"]["totalR"] and s["whipsawImprovementR"] > 0.5
    ]
    if not candidates:
        candidates = [
            s
            for s in summaries
            if s["totalR"] >= baseline["stats"]["totalR"] - 2.0 and s["whipsawImprovementR"] > 1.0
        ]
    best_summary = max(candidates or summaries, key=lambda s: (s["whipsawImprovementR"], s["totalR"], -abs(s["maxDrawdownR"])))
    best = next(r for r in results if r["name"] == best_summary["name"])

    payload = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "baseline": {
            "name": "latest_nxt32",
            "stats": baseline["stats"],
            "whipsawWindows": baseline_whipsaw,
        },
        "best": best,
        "fullSummary": sorted(summaries, key=lambda s: (-s["totalR"], -s["whipsawImprovementR"])),
        "topByWhipsawImprovement": sorted(summaries, key=lambda s: (-s["whipsawImprovementR"], -s["totalR"]))[:40],
    }
    OUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    workbook_payload = {
        "generatedAt": payload["generatedAt"],
        "systemVersion": f"NXT v3.2 Latest Test + Whipsaw Risk-Off - {best['name']}",
        "period": {
            "start": base.START_DATE.isoformat(),
            "end": (base.END_DATE - base.timedelta(days=1)).isoformat(),
            "timezone": "Binance native daily candles",
        },
        "symbols": base.SYMBOLS,
        "stats": best["stats"],
        "trades": best["trades"],
        "datasets": datasets,
        "assumptions": [
            "Baseline is current latest NXT v3.2 with Binance native 1D candles, TradingView ATR RMA, Runner A, 20% reversal-skip, no continuation, and no risk-off.",
            "Risk-off test does not skip entries. It scales a trade's R by the risk multiplier when the pre-entry whipsaw score reaches the configured threshold.",
            f"Best tested rule: {best['name']}.",
            "This is a test output only and is not promoted to latest.",
        ],
    }
    write_compact_workbook(OUT_XLSX, best, baseline, datasets)

    print(
        json.dumps(
            {
                "outJson": str(OUT_JSON),
                "outXlsx": str(OUT_XLSX),
                "baseline": {
                    "totalR": baseline["stats"]["totalR"],
                    "maxDrawdownR": baseline["stats"]["maxDrawdownR"],
                    "whipsawCombinedR": baseline_combined,
                },
                "best": best_summary,
                "top": payload["topByWhipsawImprovement"][:10],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
