from __future__ import annotations

import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\Workspace\Codex\Crypto trading")
BASE = ROOT / "outputs" / "nxt_v31_runner_ab_6y" / "nxt_v31_runner_ab_6y_results.json"
CONT = ROOT / "outputs" / "nxt31_continuation_module_6y" / "nxt31_continuation_module_6y_results.json"
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
OUT = ROOT / "outputs" / "nxt31_continuation_module_6y" / "NXT31_Full_Trades_With_SSL_Reentry_Continuation_6Y_BTC_SOL_SUI_20K.xlsx"
OUT_JSON = ROOT / "outputs" / "nxt31_continuation_module_6y" / "nxt31_full_trades_with_ssl_reentry_continuation.json"
VARIANT_KEY = "cont_ssl_reentry_dist25"


def clear(ws, rows=380, cols=32):
    for row in ws.iter_rows(min_row=1, max_row=max(rows, ws.max_row), min_col=1, max_col=max(cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def rowset(ws, r, values):
    for c, v in enumerate(values, 1):
        ws.cell(r, c).value = v


def copy_layout(src, dst):
    dst.sheet_format.defaultColWidth = src.sheet_format.defaultColWidth
    dst.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    for c in range(1, max(src.max_column, dst.max_column) + 1):
        letter = get_column_letter(c)
        if src.column_dimensions.get(letter) and src.column_dimensions[letter].width:
            dst.column_dimensions[letter].width = src.column_dimensions[letter].width
    for r in range(1, max(src.max_row, dst.max_row) + 1):
        sr = r if r <= src.max_row else 5
        if src.row_dimensions.get(sr) and src.row_dimensions[sr].height:
            dst.row_dimensions[r].height = src.row_dimensions[sr].height
        for c in range(1, max(src.max_column, dst.max_column) + 1):
            sc = min(c, src.max_column)
            a = src.cell(sr, sc)
            b = dst.cell(r, c)
            if a.has_style:
                b._style = copy(a._style)
            b.number_format = a.number_format
            b.alignment = copy(a.alignment)
            b.font = copy(a.font)
            b.fill = copy(a.fill)
            b.border = copy(a.border)


def max_dd(trades, key="rMultiple"):
    cum = peak = dd = 0.0
    for t in sorted(trades, key=lambda x: x["exitTime"]):
        cum += t[key]
        peak = max(peak, cum)
        dd = min(dd, cum - peak)
    return dd


def stats(trades, key="rMultiple"):
    rows = sorted(trades, key=lambda x: x["exitTime"])
    total = sum(t[key] for t in rows)
    wins = sum(1 for t in rows if t[key] > 0)
    cont = [t for t in rows if t.get("signalType") == "Continuation"]
    return {
        "trades": len(rows),
        "wins": wins,
        "losses": len(rows) - wins,
        "winRate": wins / len(rows),
        "totalR": total,
        "avgR": total / len(rows),
        "maxDrawdownR": max_dd(rows, key),
        "continuationTrades": len(cont),
        "continuationR": sum(t[key] for t in cont),
    }


def riskoff(trades, threshold=-4.0, scale=0.4):
    rows = []
    cum = peak = 0.0
    for t in sorted(trades, key=lambda x: x["exitTime"]):
        row = dict(t)
        pre = cum - peak
        mult = scale if pre <= threshold else 1.0
        row["preTradeDrawdownR"] = pre
        row["sizeMultiplier"] = mult
        row["baseR"] = row["rMultiple"]
        row["riskOffR"] = row["rMultiple"] * mult
        cum += row["riskOffR"]
        peak = max(peak, cum)
        row["postTradeDrawdownR"] = cum - peak
        rows.append(row)
    return rows


def normalize_primary(t):
    row = dict(t)
    row["signalType"] = "Primary"
    row["tp1Time"] = row.get("tp1Time", "")
    row["tp1"] = row.get("tp1", row.get("tp2", ""))
    row["rsi14"] = row.get("rsi14", "")
    return row


def normalize_continuation(t):
    row = dict(t)
    row["finalStop"] = row.get("finalStop", row.get("stop", ""))
    row["tp1"] = row.get("tp1", row.get("tp", ""))
    row["rsi14"] = row.get("rsi14", "")
    return row


def main():
    base = json.loads(BASE.read_text(encoding="utf-8"))
    cont = json.loads(CONT.read_text(encoding="utf-8"))
    primary = [normalize_primary(t) for t in next(v for v in base["variants"] if v["key"] == "runner_a_50_50_ssl")["trades"]]
    continuation = [normalize_continuation(t) for t in cont["variants"][VARIANT_KEY]["continuationTrades"]]
    combined = sorted([*primary, *continuation], key=lambda x: x["exitTime"])
    risk_rows = riskoff(combined)

    result = {
        "variant": cont["variants"][VARIANT_KEY]["name"],
        "statsBeforeRiskOff": stats(combined),
        "statsAfterRiskOff": stats(risk_rows, "riskOffR"),
        "trades": risk_rows,
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")

    tpl = load_workbook(TEMPLATE)
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        clear(ws)

    ws = wb["Summary"]
    ws["A1"] = "NXT v3.1 Full Trades + SSL Re-entry Continuation"
    ws["A2"] = "Primary Runner A trades plus continuation trades that capture cases like 13-Dec-2020."
    rows = [
        ["Metric", "Before Risk-Off", "After Risk-Off"],
        ["Trades", result["statsBeforeRiskOff"]["trades"], result["statsAfterRiskOff"]["trades"]],
        ["Win Rate", result["statsBeforeRiskOff"]["winRate"], result["statsAfterRiskOff"]["winRate"]],
        ["Total R", result["statsBeforeRiskOff"]["totalR"], result["statsAfterRiskOff"]["totalR"]],
        ["Average R", result["statsBeforeRiskOff"]["avgR"], result["statsAfterRiskOff"]["avgR"]],
        ["Max DD R", result["statsBeforeRiskOff"]["maxDrawdownR"], result["statsAfterRiskOff"]["maxDrawdownR"]],
        ["Continuation Trades", result["statsBeforeRiskOff"]["continuationTrades"], result["statsAfterRiskOff"]["continuationTrades"]],
        ["Continuation R", result["statsBeforeRiskOff"]["continuationR"], result["statsAfterRiskOff"]["continuationR"]],
    ]
    for i, row in enumerate(rows, 4):
        rowset(ws, i, row)

    headers = ["Symbol", "No", "Signal Type", "Side", "Signal Time", "Entry Time", "Entry Price", "Initial Stop", "Final Stop", "Risk / Unit", "TP1 2.5 ATR", "TP1 Time", "Exit Time", "Exit Price", "Exit Reason", "Base R", "Risk-Off R", "Size Mult", "Pre DD R", "Post DD R", "ATR14", "RSI14", "Distance EMA50 ATR", "Notes"]
    for sheet, subset in [
        ("Trades", risk_rows),
        ("BTC", [t for t in risk_rows if t["symbol"] == "BTCUSDT"]),
        ("SOL", [t for t in risk_rows if t["symbol"] == "SOLUSDT"]),
        ("SUI", [t for t in risk_rows if t["symbol"] == "SUIUSDT"]),
    ]:
        ws = wb[sheet]
        ws["A1"] = f"{sheet} - Full Trades With Continuation" if sheet != "Trades" else "Detailed Trades - Full With Continuation"
        ws["A2"] = "Primary and continuation entries in one list."
        rowset(ws, 4, headers)
        for i, t in enumerate(subset, 5):
            rowset(ws, i, [
                t["symbol"].replace("USDT", ""), t.get("tradeNo", ""), t.get("signalType", ""), t["side"],
                t["signalTime"].replace("T", " ").replace("Z", ""), t["entryTime"].replace("T", " ").replace("Z", ""),
                t["entryPrice"], t["initialStop"], t.get("finalStop", ""), t.get("riskPerUnit", ""), t.get("tp1", ""),
                t.get("tp1Time", "").replace("T", " ").replace("Z", "") if t.get("tp1Time") else "",
                t["exitTime"].replace("T", " ").replace("Z", ""), t["exitPrice"], t["exitReason"],
                t["baseR"], t["riskOffR"], t["sizeMultiplier"], t["preTradeDrawdownR"], t["postTradeDrawdownR"],
                t.get("atr14", ""), t.get("rsi14", ""), t.get("distanceToEma50Atr", ""), t.get("notes", ""),
            ])

    ass = wb["Assumptions"]
    ass["A1"] = "Assumptions"
    ass["A2"] = "NXT v3.1 Runner A unchanged; SSL re-entry continuation module added."
    lines = [
        "Primary NXT v3.1 trades are unchanged.",
        "Continuation rule used: SSL re-entry after failed distance, distance <= 2.5 ATR.",
        "Continuation uses Runner A exit: 50% at 2.5 ATR, 50% runner exits on SSL flip or stop.",
        "Risk-off overlay remains the latest rule: DD <= -4R uses 40% size.",
        "This variant explicitly captures BTCUSDT continuation around 13-Dec-2020.",
    ]
    rowset(ass, 4, ["#", "Assumption"])
    for i, line in enumerate(lines, 5):
        rowset(ass, i, [i - 4, line])

    for sheet in ["Summary", "Trades", "BTC", "SOL", "SUI", "Assumptions", "Data Quality", "20K Account", "Equity Curve"]:
        if sheet in wb.sheetnames and sheet in tpl.sheetnames:
            copy_layout(tpl[sheet], wb[sheet])

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
