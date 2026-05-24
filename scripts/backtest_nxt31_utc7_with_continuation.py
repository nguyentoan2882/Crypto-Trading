from __future__ import annotations

import importlib.util
import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\Workspace\Codex\Crypto trading")
UTC7_SCRIPT = ROOT / "scripts" / "backtest_nxt31_utc7_latest.py"
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
OUT_DIR = ROOT / "outputs" / "nxt31_utc7_continuation_6y"
OUT_JSON = OUT_DIR / "nxt31_utc7_continuation_6y_results.json"
OUT_XLSX = OUT_DIR / "NXT31_UTC7_With_Continuation_6Y_BTC_SOL_SUI_20K.xlsx"


spec = importlib.util.spec_from_file_location("utc7", UTC7_SCRIPT)
utc7 = importlib.util.module_from_spec(spec)
spec.loader.exec_module(utc7)


CONT_CFG = {
    "key": "ssl_reentry_dist25",
    "name": "SSL re-entry continuation dist<=2.5",
    "maxDist": 2.5,
    "longRsi": 55,
    "shortRsi": 45,
    "barsAfterFlip": 3,
}


def continuation_signal(candles, i):
    c = candles[i]
    if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]):
        return None
    dist = abs(c["close"] - c["ema50"]) / c["atr14"]
    if dist > CONT_CFG["maxDist"]:
        return None
    recent_bull = False
    recent_bear = False
    for j in range(max(1, i - CONT_CFG["barsAfterFlip"]), i + 1):
        recent_bull = recent_bull or (candles[j - 1]["ssl"] == -1 and candles[j]["ssl"] == 1)
        recent_bear = recent_bear or (candles[j - 1]["ssl"] == 1 and candles[j]["ssl"] == -1)
    if recent_bull and c["ssl"] == 1 and c["close"] > c["ema20"] and c["rsi14"] > CONT_CFG["longRsi"] and c["close"] > c["open"]:
        return "LONG"
    if recent_bear and c["ssl"] == -1 and c["close"] < c["ema20"] and c["rsi14"] < CONT_CFG["shortRsi"] and c["close"] < c["open"]:
        return "SHORT"
    return None


def ranges(trades):
    out = []
    for t in trades:
        out.append((t["symbol"], t["entryTime"], t["exitTime"]))
    return out


def is_overlapping_primary(symbol, entry_date, primary_ranges):
    return any(sym == symbol and start <= entry_date <= end for sym, start, end in primary_ranges)


def backtest_continuation_symbol(symbol, candles, primary_ranges):
    trades = []
    pos = None
    n = 1
    for i in range(55, len(candles) - 1):
        c, prev, nxt = candles[i], candles[i - 1], candles[i + 1]
        next_date = __import__("datetime").date.fromisoformat(nxt["localDate"])
        if next_date < utc7.START_DATE or next_date >= utc7.END_DATE:
            continue
        if pos:
            side = pos["side"]
            ssl_flip = (side == "LONG" and prev["ssl"] == 1 and c["ssl"] == -1) or (side == "SHORT" and prev["ssl"] == -1 and c["ssl"] == 1)
            exit_price = reason = None
            if side == "LONG":
                if c["low"] <= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["high"] >= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["tp"] - pos["entry"]) / pos["risk"])
                    if pos["triggered"] and ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bearish flip"
            else:
                if c["high"] >= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["low"] <= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["entry"] - pos["tp"]) / pos["risk"])
                    if pos["triggered"] and ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bullish flip"
            if exit_price is not None:
                rem = 0.5 if pos["triggered"] else 1.0
                rem_r = (exit_price - pos["entry"]) / pos["risk"] if side == "LONG" else (pos["entry"] - exit_price) / pos["risk"]
                gross = pos["realizedR"] + rem * rem_r
                net = gross - utc7.cost_r(pos["entry"], pos["risk"])
                trades.append({
                    "symbol": symbol,
                    "tradeNo": n,
                    "signalType": "Continuation",
                    "side": side,
                    "signalTime": pos["signalDate"],
                    "entryTime": pos["entryDate"],
                    "entryPrice": pos["entry"],
                    "initialStop": pos["initialStop"],
                    "finalStop": pos["stop"],
                    "riskPerUnit": pos["risk"],
                    "tp1": pos["tp"],
                    "tp1Time": pos["tp1Time"],
                    "exitTime": c["localDate"],
                    "exitPrice": exit_price,
                    "exitReason": reason,
                    "grossRMultiple": gross,
                    "costR": utc7.cost_r(pos["entry"], pos["risk"]),
                    "rMultiple": net,
                    "atr14": pos["atr14"],
                    "rsi14": pos["rsi14"],
                    "distanceToEma50Atr": pos["distance"],
                    "notes": CONT_CFG["name"],
                })
                n += 1
                pos = None
            if pos:
                continue
        side = continuation_signal(candles, i)
        if not side:
            continue
        if is_overlapping_primary(symbol, nxt["localDate"], primary_ranges):
            continue
        risk = c["atr14"] * 1.5
        entry = nxt["open"]
        pos = {
            "side": side,
            "signalDate": c["localDate"],
            "entryDate": nxt["localDate"],
            "entry": entry,
            "initialStop": entry - risk if side == "LONG" else entry + risk,
            "stop": entry - risk if side == "LONG" else entry + risk,
            "risk": risk,
            "tp": entry + c["atr14"] * 2.5 if side == "LONG" else entry - c["atr14"] * 2.5,
            "triggered": False,
            "tp1Time": "",
            "realizedR": 0.0,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": abs(c["close"] - c["ema50"]) / c["atr14"],
        }
    return trades


def clear(ws, rows=420, cols=32):
    for row in ws.iter_rows(min_row=1, max_row=max(rows, ws.max_row), min_col=1, max_col=max(cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def rowset(ws, r, values):
    for c, v in enumerate(values, 1):
        ws.cell(r, c).value = v


def copy_layout(src, dst):
    dst.sheet_format.defaultColWidth = src.sheet_format.defaultColWidth
    dst.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    for c in range(1, max(src.max_column, dst.max_column) + 1):
        letter = get_column_letter(c)
        if src.column_dimensions.get(letter) and src.column_dimensions[letter].width:
            dst.column_dimensions[letter].width = src.column_dimensions[letter].width
    for r in range(1, max(src.max_row, dst.max_row) + 1):
        sr = r if r <= src.max_row else 5
        if src.row_dimensions.get(sr) and src.row_dimensions[sr].height:
            dst.row_dimensions[r].height = src.row_dimensions[sr].height
        for c in range(1, max(src.max_column, dst.max_column) + 1):
            a = src.cell(sr, min(c, src.max_column))
            b = dst.cell(r, c)
            if a.has_style:
                b._style = copy(a._style)
            b.number_format = a.number_format
            b.alignment = copy(a.alignment)
            b.font = copy(a.font)
            b.fill = copy(a.fill)
            b.border = copy(a.border)


def build_workbook(result):
    tpl = load_workbook(TEMPLATE)
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        clear(ws)
    ws = wb["Summary"]
    ws["A1"] = "NXT v3.1 UTC+7 With Continuation"
    ws["A2"] = "Primary Runner A plus SSL re-entry continuation module; risk-off applied."
    rows = [
        ["Metric", "Before Risk-Off", "After Risk-Off"],
        ["Trades", result["stats"]["trades"], result["riskOffStats"]["trades"]],
        ["Win Rate", result["stats"]["winRate"], result["riskOffStats"]["winRate"]],
        ["Total R", result["stats"]["totalR"], result["riskOffStats"]["totalR"]],
        ["Average R", result["stats"]["avgR"], result["riskOffStats"]["avgR"]],
        ["Max DD R", result["stats"]["maxDrawdownR"], result["riskOffStats"]["maxDrawdownR"]],
        ["Continuation Trades", result["continuationStats"]["trades"], result["continuationRiskOffStats"]["trades"]],
        ["Continuation R", result["continuationStats"]["totalR"], result["continuationRiskOffStats"]["totalR"]],
    ]
    for i, row in enumerate(rows, 4):
        rowset(ws, i, row)
    headers = ["Symbol", "No", "Signal Type", "Side", "Signal Date UTC+7", "Entry Date UTC+7", "Entry Price", "Initial Stop", "Final Stop", "Risk / Unit", "TP1", "TP1 Date", "Exit Date UTC+7", "Exit Price", "Exit Reason", "Base R", "Risk-Off R", "Size Mult", "Pre DD R", "Post DD R", "ATR14", "RSI14", "Distance EMA50 ATR", "Notes"]
    for sheet, subset in [
        ("Trades", result["riskOffTrades"]),
        ("BTC", [t for t in result["riskOffTrades"] if t["symbol"] == "BTCUSDT"]),
        ("SOL", [t for t in result["riskOffTrades"] if t["symbol"] == "SOLUSDT"]),
        ("SUI", [t for t in result["riskOffTrades"] if t["symbol"] == "SUIUSDT"]),
    ]:
        ws = wb[sheet]
        ws["A1"] = "Detailed Trades - UTC+7 With Continuation" if sheet == "Trades" else f"{sheet} - UTC+7 With Continuation"
        ws["A2"] = "Primary and continuation trades in one list."
        rowset(ws, 4, headers)
        for i, t in enumerate(subset, 5):
            rowset(ws, i, [t["symbol"].replace("USDT", ""), t["tradeNo"], t["signalType"], t["side"], t["signalTime"], t["entryTime"], t["entryPrice"], t["initialStop"], t["finalStop"], t["riskPerUnit"], t["tp1"], t["tp1Time"], t["exitTime"], t["exitPrice"], t["exitReason"], t["rMultiple"], t["riskOffR"], t["sizeMultiplier"], t["preTradeDrawdownR"], t["postTradeDrawdownR"], t["atr14"], t["rsi14"], t["distanceToEma50Atr"], t["notes"]])
    ass = wb["Assumptions"]
    ass["A1"] = "Assumptions"
    rowset(ass, 4, ["#", "Assumption"])
    for i, line in enumerate(result["assumptions"], 5):
        rowset(ass, i, [i - 4, line])
    for sheet in ["Summary", "Trades", "BTC", "SOL", "SUI", "Assumptions", "Data Quality", "20K Account", "Equity Curve"]:
        if sheet in wb.sheetnames and sheet in tpl.sheetnames:
            copy_layout(tpl[sheet], wb[sheet])
    wb.save(OUT_XLSX)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    datasets = {}
    primary = []
    candles_by_symbol = {}
    for symbol in utc7.SYMBOLS:
        h = utc7.fetch_1h(symbol)
        d = utc7.enrich(utc7.resample_utc7(h))
        candles_by_symbol[symbol] = d
        datasets[symbol] = {"hourlyRows": len(h), "dailyRows": len(d), "firstDay": d[0]["localDate"], "lastDay": d[-1]["localDate"], "source": "Binance spot 1H klines resampled to UTC+7 day"}
        for t in utc7.backtest_symbol(symbol, d):
            t["signalType"] = "Primary"
            primary.append(t)
    primary.sort(key=lambda x: x["exitTime"])
    pr = ranges(primary)
    cont = []
    for symbol in utc7.SYMBOLS:
        cont.extend(backtest_continuation_symbol(symbol, candles_by_symbol[symbol], pr))
    combined = sorted([*primary, *cont], key=lambda x: x["exitTime"])
    ro = utc7.riskoff(combined)
    result = {
        "generatedAt": __import__("datetime").datetime.now(__import__("datetime").timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "systemVersion": "NXT v3.1 Runner A + Risk-Off + SSL Re-entry Continuation, UTC+7 daily candles",
        "continuationRule": CONT_CFG,
        "stats": utc7.stats(combined),
        "riskOffStats": utc7.stats(ro, "riskOffR"),
        "continuationStats": utc7.stats(cont),
        "continuationRiskOffStats": utc7.stats([t for t in ro if t["signalType"] == "Continuation"], "riskOffR") if cont else utc7.stats([]),
        "trades": combined,
        "riskOffTrades": ro,
        "datasets": datasets,
        "assumptions": [
            "Daily candles are resampled from Binance spot 1H klines using Asia/Saigon UTC+7 calendar days.",
            "Primary NXT v3.1 rules are unchanged.",
            "Continuation module: SSL re-entry within 3 candles, distance <= 2.5 ATR, LONG RSI >55 and close above EMA20, SHORT RSI <45 and close below EMA20.",
            "Continuation entries are skipped if a primary trade on the same symbol is already open.",
            "Runner A and risk-off overlay are unchanged.",
        ],
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    build_workbook(result)
    print(json.dumps({"outJson": str(OUT_JSON), "outXlsx": str(OUT_XLSX), "stats": result["stats"], "riskOffStats": result["riskOffStats"], "continuationStats": result["continuationStats"], "continuationRiskOffStats": result["continuationRiskOffStats"]}, indent=2))


if __name__ == "__main__":
    main()
