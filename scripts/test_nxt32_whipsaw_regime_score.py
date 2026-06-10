from __future__ import annotations

import json
import statistics
import sys
from copy import deepcopy
from datetime import date, datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

import backtest_nxt31_utc7_latest as base
import backtest_nxt32_native_1d_latest as native
import rebuild_nxt32_native_1d_tv_atr_latest as tv_atr


ROOT = Path.cwd()
OUT_DIR = ROOT / "outputs" / "nxt32_whipsaw_regime_score"
OUT_JSON = OUT_DIR / "nxt32_whipsaw_regime_score_results.json"
OUT_XLSX = OUT_DIR / "NXT32_Whipsaw_Regime_Score_Best_6Y_BTC_SOL_SUI_20K.xlsx"
BASELINE_JSON = ROOT / "latest" / "NXT_Latest_NXT32_Native1D_RunnerA_NoContinuation_NoRiskOff_6Y_BTC_SOL_SUI_20K.json"

WHIPSAW_WINDOWS = [
    ("2020-05-29", "2020-07-07"),
    ("2022-05-31", "2022-10-05"),
    ("2023-08-02", "2023-10-12"),
]


def trade_in_window(trade: dict, start: str, end: str) -> bool:
    entry = date.fromisoformat(trade["entryTime"])
    return date.fromisoformat(start) <= entry <= date.fromisoformat(end)


def window_breakdown(trades: list[dict]) -> list[dict]:
    rows = []
    for start, end in WHIPSAW_WINDOWS:
        subset = [t for t in trades if trade_in_window(t, start, end)]
        rows.append(
            {
                "start": start,
                "end": end,
                "trades": len(subset),
                "wins": sum(1 for t in subset if t["rMultiple"] > 0),
                "totalR": sum(t["rMultiple"] for t in subset),
                "bySymbol": {
                    symbol: {
                        "trades": len([t for t in subset if t["symbol"] == symbol]),
                        "totalR": sum(t["rMultiple"] for t in subset if t["symbol"] == symbol),
                    }
                    for symbol in sorted({t["symbol"] for t in subset})
                },
            }
        )
    rows.append(
        {
            "start": "combined",
            "end": "combined",
            "trades": sum(r["trades"] for r in rows),
            "wins": sum(r["wins"] for r in rows),
            "totalR": sum(r["totalR"] for r in rows),
            "bySymbol": {},
        }
    )
    return rows


def ssl_flip_count(candles: list[dict], index: int, lookback: int) -> int:
    start = max(1, index - lookback + 1)
    return sum(
        1
        for i in range(start, index + 1)
        if candles[i].get("ssl") is not None
        and candles[i - 1].get("ssl") is not None
        and candles[i]["ssl"] != candles[i - 1]["ssl"]
    )


def efficiency_ratio(candles: list[dict], index: int, lookback: int) -> float | None:
    if index < lookback:
        return None
    net = abs(candles[index]["close"] - candles[index - lookback]["close"])
    path = sum(abs(candles[i]["close"] - candles[i - 1]["close"]) for i in range(index - lookback + 1, index + 1))
    if path == 0:
        return 0.0
    return net / path


def median_ema50_distance(candles: list[dict], index: int, lookback: int) -> float | None:
    if index < lookback - 1:
        return None
    vals = []
    for i in range(index - lookback + 1, index + 1):
        c = candles[i]
        if c.get("ema50") is None or c.get("atr14") in (None, 0):
            return None
        vals.append(abs(c["close"] - c["ema50"]) / c["atr14"])
    return statistics.median(vals)


def ema50_slope_atr(candles: list[dict], index: int, lookback: int) -> float | None:
    if index < lookback:
        return None
    now = candles[index]
    old = candles[index - lookback]
    if now.get("ema50") is None or old.get("ema50") is None or now.get("atr14") in (None, 0):
        return None
    return abs(now["ema50"] - old["ema50"]) / now["atr14"]


def whipsaw_features(candles: list[dict], index: int, cfg: dict) -> dict:
    c = candles[index]
    dist = abs(c["close"] - c["ema50"]) / c["atr14"]
    return {
        "sslFlipCount": ssl_flip_count(candles, index, cfg["flipLookback"]),
        "er": efficiency_ratio(candles, index, cfg["erLookback"]),
        "medianDistanceToEma50Atr": median_ema50_distance(candles, index, cfg["compressionLookback"]),
        "ema50SlopeAtr": ema50_slope_atr(candles, index, cfg["slopeLookback"]),
        "distanceToEma50Atr": dist,
    }


def whipsaw_score(features: dict, cfg: dict) -> tuple[int, list[str]]:
    reasons = []
    if features["sslFlipCount"] > cfg["maxFlips"]:
        reasons.append("ssl_flip_density")
    if features["er"] is not None and features["er"] < cfg["maxEr"]:
        reasons.append("low_efficiency_ratio")
    if features["medianDistanceToEma50Atr"] is not None and features["medianDistanceToEma50Atr"] < cfg["maxMedianDistance"]:
        reasons.append("ema50_compression")
    if features["ema50SlopeAtr"] is not None and features["ema50SlopeAtr"] < cfg["maxEma50Slope"]:
        reasons.append("flat_ema50")
    if features["distanceToEma50Atr"] < cfg["minEntryDistance"]:
        reasons.append("entry_too_close_to_ema50")
    return len(reasons), reasons


def backtest_symbol(symbol: str, candles: list[dict], cfg: dict) -> tuple[list[dict], list[dict]]:
    trades, skipped, pos, n = [], [], None, 1
    skip_reversal = None
    for i in range(55, len(candles) - 1):
        c, prev, nxt = candles[i], candles[i - 1], candles[i + 1]
        next_date = base.date.fromisoformat(nxt["localDate"])
        if next_date < base.START_DATE or next_date >= base.END_DATE:
            continue

        skip_reversal = None
        if pos:
            side = pos["side"]
            if side == "LONG":
                pos["favorable20"] = pos["favorable20"] or c["high"] >= pos["entry"] * 1.20
            else:
                pos["favorable20"] = pos["favorable20"] or c["low"] <= pos["entry"] * 0.80

            ssl_flip = (side == "LONG" and prev["ssl"] == 1 and c["ssl"] == -1) or (
                side == "SHORT" and prev["ssl"] == -1 and c["ssl"] == 1
            )
            exit_price = reason = None
            if side == "LONG":
                if c["low"] <= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["high"] >= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["tp"] - pos["entry"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bearish flip"
                        if pos["favorable20"]:
                            skip_reversal = "SHORT"
            else:
                if c["high"] >= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["low"] <= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["entry"] - pos["tp"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bullish flip"
                        if pos["favorable20"]:
                            skip_reversal = "LONG"

            if exit_price is not None:
                rem = 0.5 if pos["triggered"] else 1.0
                rem_r = (exit_price - pos["entry"]) / pos["risk"] if side == "LONG" else (pos["entry"] - exit_price) / pos["risk"]
                gross = pos["realizedR"] + rem * rem_r
                cost = base.cost_r(pos["entry"], pos["risk"])
                net = gross - cost
                trades.append(
                    {
                        "symbol": symbol,
                        "tradeNo": n,
                        "side": side,
                        "signalTime": pos["signalDate"],
                        "entryTime": pos["entryDate"],
                        "entryPrice": pos["entry"],
                        "initialStop": pos["initialStop"],
                        "finalStop": pos["stop"],
                        "riskPerUnit": pos["risk"],
                        "tp1": pos["tp"],
                        "tp1Time": pos["tp1Time"],
                        "exitTime": c["localDate"],
                        "exitPrice": exit_price,
                        "exitReason": reason,
                        "grossRMultiple": gross,
                        "costR": cost,
                        "rMultiple": net,
                        "atr14": pos["atr14"],
                        "rsi14": pos["rsi14"],
                        "distanceToEma50Atr": pos["distance"],
                        "whipsawScore": pos["whipsawScore"],
                        "whipsawReasons": ", ".join(pos["whipsawReasons"]),
                        "efficiencyRatio": pos["features"]["er"],
                        "medianDistanceToEma50Atr": pos["features"]["medianDistanceToEma50Atr"],
                        "ema50SlopeAtr": pos["features"]["ema50SlopeAtr"],
                        "sslFlipCount": pos["features"]["sslFlipCount"],
                        "favorable20PctReached": pos["favorable20"],
                        "notes": cfg["notes"],
                    }
                )
                n += 1
                pos = None
            if pos:
                continue

        if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]) or prev["ssl"] is None:
            continue
        dist = abs(c["close"] - c["ema50"]) / c["atr14"]
        long_ok = prev["ssl"] == -1 and c["ssl"] == 1 and base.recent_cross(candles, i, "LONG") and dist <= 2 and c["rsi14"] > 50
        short_ok = prev["ssl"] == 1 and c["ssl"] == -1 and base.recent_cross(candles, i, "SHORT") and dist <= 2 and c["rsi14"] < 50

        if skip_reversal == "LONG" and long_ok:
            skipped.append({"symbol": symbol, "date": c["localDate"], "side": "LONG", "reason": "20pct_reversal_skip"})
            long_ok = False
        if skip_reversal == "SHORT" and short_ok:
            skipped.append({"symbol": symbol, "date": c["localDate"], "side": "SHORT", "reason": "20pct_reversal_skip"})
            short_ok = False

        features = whipsaw_features(candles, i, cfg)
        score, reasons = whipsaw_score(features, cfg)
        if (long_ok or short_ok) and score >= cfg["skipScore"]:
            skipped.append(
                {
                    "symbol": symbol,
                    "date": c["localDate"],
                    "side": "LONG" if long_ok else "SHORT",
                    "reason": "whipsaw_regime_score",
                    "score": score,
                    "reasons": reasons,
                    "features": features,
                }
            )
            long_ok = short_ok = False

        if not (long_ok or short_ok):
            continue
        side = "LONG" if long_ok else "SHORT"
        risk = c["atr14"] * 1.5
        entry = nxt["open"]
        pos = {
            "side": side,
            "signalDate": c["localDate"],
            "entryDate": nxt["localDate"],
            "entry": entry,
            "initialStop": entry - risk if side == "LONG" else entry + risk,
            "stop": entry - risk if side == "LONG" else entry + risk,
            "risk": risk,
            "tp": entry + c["atr14"] * 2.5 if side == "LONG" else entry - c["atr14"] * 2.5,
            "triggered": False,
            "tp1Time": "",
            "realizedR": 0.0,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": dist,
            "features": features,
            "whipsawScore": score,
            "whipsawReasons": reasons,
            "favorable20": False,
        }
    return trades, skipped


def run_variant(cfg: dict, candles_by_symbol: dict[str, list[dict]]) -> dict:
    all_trades, all_skipped = [], []
    for symbol, candles in candles_by_symbol.items():
        trades, skipped = backtest_symbol(symbol, candles, cfg)
        all_trades.extend(trades)
        all_skipped.extend(skipped)
    all_trades.sort(key=lambda x: x["exitTime"])
    return {
        "name": cfg["name"],
        "config": cfg,
        "stats": base.stats(all_trades),
        "whipsawWindows": window_breakdown(all_trades),
        "skippedSignals": all_skipped,
        "trades": all_trades,
    }


def fix_workbook_labels(path: Path, result: dict) -> None:
    wb = load_workbook(path)
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        ws["A1"] = "NXT v3.2 Latest Test - Whipsaw Regime Score"
        ws["A2"] = f"BTC/SOL/SUI 6Y | {result['name']} | 20% reversal-skip | no continuation | no risk-off."
    if "Trades" in wb.sheetnames:
        wb["Trades"]["A1"] = "Detailed Trades - NXT Whipsaw Regime Score Test"
    for sheet in ["BTC", "SOL", "SUI"]:
        if sheet in wb.sheetnames:
            wb[sheet]["A1"] = f"{sheet} - NXT Whipsaw Regime Score Test"
    wb.save(path)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    baseline = json.loads(BASELINE_JSON.read_text(encoding="utf-8"))
    candles_by_symbol = {symbol: tv_atr.enrich_tv_atr(native.fetch_native_1d(symbol)) for symbol in base.SYMBOLS}
    datasets = {
        symbol: {
            "dailyRows": len(candles),
            "firstDay": candles[0]["localDate"],
            "lastDay": candles[-1]["localDate"],
            "source": "Binance spot native 1D klines",
        }
        for symbol, candles in candles_by_symbol.items()
    }

    base_cfg = {
        "flipLookback": 20,
        "erLookback": 20,
        "compressionLookback": 10,
        "slopeLookback": 20,
        "minEntryDistance": 0.0,
        "notes": "NXT latest plus whipsaw regime score test",
    }
    variants = []
    for max_flips in [3, 4, 5]:
        for max_er in [0.20, 0.25, 0.30, 0.35]:
            for max_median_dist in [0.45, 0.60, 0.75, 0.90]:
                for max_slope in [0.25, 0.40, 0.60]:
                    for skip_score in [2, 3, 4]:
                        cfg = deepcopy(base_cfg)
                        cfg.update(
                            {
                                "name": f"score{skip_score}_flip{max_flips}_er{max_er:g}_dist{max_median_dist:g}_slope{max_slope:g}",
                                "maxFlips": max_flips,
                                "maxEr": max_er,
                                "maxMedianDistance": max_median_dist,
                                "maxEma50Slope": max_slope,
                                "skipScore": skip_score,
                            }
                        )
                        variants.append(cfg)

    results = [run_variant(cfg, candles_by_symbol) for cfg in variants]
    baseline_whipsaw = window_breakdown(baseline["trades"])
    baseline_combined = baseline_whipsaw[-1]["totalR"]

    summary = []
    for result in results:
        stats = result["stats"]
        combined = result["whipsawWindows"][-1]["totalR"]
        summary.append(
            {
                "name": result["name"],
                "trades": stats["trades"],
                "totalR": stats["totalR"],
                "maxDrawdownR": stats["maxDrawdownR"],
                "winRate": stats["winRate"],
                "whipsawCombinedR": combined,
                "whipsawImprovementR": combined - baseline_combined,
                "skipped": len(result["skippedSignals"]),
            }
        )

    practical_candidates = [
        s
        for s in summary
        if s["totalR"] >= baseline["stats"]["totalR"] and s["whipsawImprovementR"] > 0
    ]
    candidates = practical_candidates or [
        s
        for s in summary
        if s["totalR"] >= baseline["stats"]["totalR"] - 3.0 and s["whipsawImprovementR"] > 1.0
    ]
    best_summary = max(
        candidates or summary,
        key=lambda s: (s["whipsawImprovementR"], s["totalR"], -abs(s["maxDrawdownR"])),
    )
    best = next(r for r in results if r["name"] == best_summary["name"])

    payload = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "baseline": {
            "name": "latest_nxt32",
            "stats": baseline["stats"],
            "whipsawWindows": baseline_whipsaw,
        },
        "best": best,
        "fullSummary": sorted(summary, key=lambda s: (-s["totalR"], -s["whipsawImprovementR"])),
        "summary": sorted(summary, key=lambda s: (-s["whipsawImprovementR"], -s["totalR"]))[:40],
    }
    OUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    workbook_payload = {
        "generatedAt": payload["generatedAt"],
        "systemVersion": f"NXT v3.2 Latest Test + Whipsaw Regime Score - {best['name']}",
        "period": {
            "start": base.START_DATE.isoformat(),
            "end": (base.END_DATE - base.timedelta(days=1)).isoformat(),
            "timezone": "Binance native daily candles",
        },
        "symbols": base.SYMBOLS,
        "stats": best["stats"],
        "trades": best["trades"],
        "datasets": datasets,
        "assumptions": [
            "Baseline is current latest NXT v3.2 with Binance native 1D candles, TradingView ATR RMA, Runner A, 20% reversal-skip, no continuation, and no risk-off.",
            "Whipsaw score is evaluated before entry using SSL flip density, Kaufman-style efficiency ratio, median distance to EMA50, EMA50 slope, and optional entry distance to EMA50.",
            f"Best tested rule: {best['name']}. Skip entry when score >= {best['config']['skipScore']}.",
            "This is a test output only and is not promoted to latest.",
        ],
    }
    native.OUT_XLSX = OUT_XLSX
    if not OUT_XLSX.exists():
        native.build_workbook(workbook_payload)
    fix_workbook_labels(OUT_XLSX, best)

    print(
        json.dumps(
            {
                "outJson": str(OUT_JSON),
                "outXlsx": str(OUT_XLSX),
                "baseline": {
                    "totalR": baseline["stats"]["totalR"],
                    "maxDrawdownR": baseline["stats"]["maxDrawdownR"],
                    "whipsawCombinedR": baseline_combined,
                },
                "best": best_summary,
                "top": payload["summary"][:10],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
