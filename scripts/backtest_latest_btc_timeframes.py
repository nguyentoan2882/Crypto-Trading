from __future__ import annotations

import csv
import io
import json
import sys
import time
import urllib.error
import urllib.request
import zipfile
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

import backtest_nxt31_utc7_latest as base
import backtest_nxt32_native_1d_latest as native
import test_nxt33_long_only_pullback_continuation as cont
from test_nxt33_ssl14 import enrich_with_ssl_period


ROOT = Path(__file__).resolve().parents[1]
CACHE = ROOT / "data_cache" / "binance_spot_klines"
FUNDING_CACHE = ROOT / "data_cache" / "binance_futures_funding"
OUT_DIR = ROOT / "outputs" / "latest_btc_timeframes"
OUT_JSON = OUT_DIR / "latest_btc_h1_h4_w1_results.json"
OUT_XLSX = OUT_DIR / "NXT_Latest_BTC_H1_H4_W1_20K.xlsx"

SYMBOL = "BTCUSDT"
TIMEFRAMES = ["1h", "4h", "1w"]
START_DATE = native.START_DATE
END_DATE = native.END_DATE
WARMUP_DATE = native.WARMUP_DATE
START_MS = int(datetime(START_DATE.year, START_DATE.month, START_DATE.day, tzinfo=timezone.utc).timestamp() * 1000)
END_MS = int(datetime(END_DATE.year, END_DATE.month, END_DATE.day, tzinfo=timezone.utc).timestamp() * 1000)
WARMUP_MS = int(datetime(WARMUP_DATE.year, WARMUP_DATE.month, WARMUP_DATE.day, tzinfo=timezone.utc).timestamp() * 1000)
STARTING_EQUITY = 20_000
ONE_R_DOLLARS = 1_000

SYSTEM_VERSION = "NXT v3.4 Latest BTC timeframe test + SSL14 + Runner A + Anti-Immediate-Reversal + LONG-only Pullback Continuation + No Risk-Off"
RULES = [
    "Data: Binance spot klines for BTCUSDT on each tested timeframe.",
    "ATR14 uses the NXT ATR-SMA variant.",
    "SSL Channel: SMA(high,14) and SMA(low,14); state flips bullish when close is above high SMA and bearish when close is below low SMA.",
    "Primary LONG: SSL flips bullish, price crosses above EMA20 within the last 3 candles, distance from close to EMA50 <= 2 ATR14, and RSI14 > 50.",
    "Primary SHORT: SSL flips bearish, price crosses below EMA20 within the last 3 candles, distance from close to EMA50 <= 2 ATR14, and RSI14 < 50.",
    "Continuation LONG: SSL is bullish, close > EMA20 > EMA50, low touched EMA20 within the last 5 candles, close > EMA20, and close > previous close.",
    "Continuation is LONG-only; SHORT continuation is disabled.",
    "Anti-immediate-reversal: after a profitable runner exit by opposite SSL flip, block an opposite-direction entry for the next 1 candle.",
    "Initial stop: 1.5 ATR14 from entry.",
    "TP1: 2.5 ATR14 from entry; close 50% at TP1.",
    "Runner A: after TP1, move remaining 50% stop to breakeven and exit runner on opposite SSL flip or breakeven stop.",
    "Risk-off overlay is disabled.",
    "Trading cost is included in R results; funding is overlaid separately from Binance USD-M fundingRate data.",
    "20K account model: starting equity $20,000 and 1R = $1,000.",
]


def iso_label(ms: int, interval: str) -> str:
    dt = datetime.fromtimestamp(ms / 1000, timezone.utc)
    return dt.date().isoformat() if interval == "1w" else dt.strftime("%Y-%m-%d %H:%M")


def interval_ms(interval: str) -> int:
    if interval == "1h":
        return 3_600_000
    if interval == "4h":
        return 14_400_000
    if interval == "1w":
        return 604_800_000
    raise ValueError(interval)


def fetch_klines(symbol: str, interval: str) -> list[dict]:
    CACHE.mkdir(parents=True, exist_ok=True)
    path = CACHE / f"{symbol}_{interval}_{WARMUP_DATE.isoformat()}_{(END_DATE - timedelta(days=1)).isoformat()}.json"
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))

    rows = []
    start = WARMUP_MS
    end = END_MS
    step = interval_ms(interval)
    while start < end:
        url = f"https://data-api.binance.vision/api/v3/klines?symbol={symbol}&interval={interval}&startTime={start}&endTime={end}&limit=1000"
        with urllib.request.urlopen(url, timeout=30) as response:
            batch = json.loads(response.read().decode("utf-8"))
        if not batch:
            break
        for item in batch:
            rows.append({
                "time": int(item[0]),
                "open": float(item[1]),
                "high": float(item[2]),
                "low": float(item[3]),
                "close": float(item[4]),
                "volume": float(item[5]),
            })
        next_start = int(batch[-1][0]) + step
        if next_start <= start:
            break
        start = next_start
        time.sleep(0.03)

    rows = sorted({int(r["time"]): r for r in rows}.values(), key=lambda r: int(r["time"]))
    path.write_text(json.dumps(rows), encoding="utf-8")
    return rows


def candles_for_interval(interval: str) -> list[dict]:
    rows = fetch_klines(SYMBOL, interval)
    out = []
    for row in rows:
        item = dict(row)
        item["localDate"] = iso_label(int(row["time"]), interval)
        if WARMUP_MS <= int(row["time"]) <= END_MS:
            out.append(item)
    return out


def touch_reclaim_long(candles: list[dict], i: int, lookback: int) -> bool:
    start = max(1, i - lookback + 1)
    touched = any(candles[j]["low"] <= candles[j]["ema20"] for j in range(start, i + 1) if candles[j]["ema20"] is not None)
    return touched and candles[i]["close"] > candles[i]["ema20"] and candles[i]["close"] > candles[i - 1]["close"]


def backtest_symbol_timeframe(interval: str, candles: list[dict]) -> list[dict]:
    trades, pos, n = [], None, 1
    last_profitable_runner_exit = None
    for i in range(55, len(candles) - 1):
        c, prev, nxt = candles[i], candles[i - 1], candles[i + 1]
        if int(nxt["time"]) < START_MS or int(nxt["time"]) >= END_MS:
            continue
        if pos:
            side = pos["side"]
            ssl_flip = (side == "LONG" and prev["ssl"] == 1 and c["ssl"] == -1) or (side == "SHORT" and prev["ssl"] == -1 and c["ssl"] == 1)
            exit_price = reason = None
            if side == "LONG":
                if c["low"] <= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["high"] >= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["tp1Ms"] = int(c["time"])
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["tp"] - pos["entry"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bearish flip"
            else:
                if c["high"] >= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["low"] <= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["tp1Ms"] = int(c["time"])
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["entry"] - pos["tp"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bullish flip"
            if exit_price is not None:
                rem = 0.5 if pos["triggered"] else 1.0
                rem_r = (exit_price - pos["entry"]) / pos["risk"] if side == "LONG" else (pos["entry"] - exit_price) / pos["risk"]
                gross = pos["realizedR"] + rem * rem_r
                cost = base.cost_r(pos["entry"], pos["risk"])
                net = gross - cost
                trades.append({
                    "timeframe": interval.upper(),
                    "symbol": SYMBOL,
                    "tradeNo": n,
                    "signalType": pos["signalType"],
                    "side": side,
                    "signalTime": pos["signalDate"],
                    "signalMs": pos["signalMs"],
                    "entryTime": pos["entryDate"],
                    "entryMs": pos["entryMs"],
                    "entryPrice": pos["entry"],
                    "initialStop": pos["initialStop"],
                    "finalStop": pos["stop"],
                    "riskPerUnit": pos["risk"],
                    "tp1": pos["tp"],
                    "tp1Time": pos["tp1Time"],
                    "tp1Ms": pos["tp1Ms"],
                    "exitTime": c["localDate"],
                    "exitMs": int(c["time"]),
                    "exitPrice": exit_price,
                    "exitReason": reason,
                    "grossRMultiple": gross,
                    "costR": cost,
                    "rMultiple": net,
                    "atr14": pos["atr14"],
                    "rsi14": pos["rsi14"],
                    "distanceToEma50Atr": pos["distance"],
                    "ema20": pos["ema20"],
                    "ema50": pos["ema50"],
                    "notes": pos["notes"],
                })
                if net > 0 and reason.startswith("Runner exit"):
                    last_profitable_runner_exit = {"index": i, "side": side}
                n += 1
                pos = None
            if pos:
                continue

        if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]) or prev["ssl"] is None:
            continue
        dist = abs(c["close"] - c["ema50"]) / c["atr14"]
        long_primary = prev["ssl"] == -1 and c["ssl"] == 1 and base.recent_cross(candles, i, "LONG") and dist <= 2 and c["rsi14"] > 50
        short_primary = prev["ssl"] == 1 and c["ssl"] == -1 and base.recent_cross(candles, i, "SHORT") and dist <= 2 and c["rsi14"] < 50
        long_cont = c["ssl"] == 1 and c["close"] > c["ema20"] > c["ema50"] and touch_reclaim_long(candles, i, 5)
        if last_profitable_runner_exit and i - last_profitable_runner_exit["index"] <= 1:
            if (long_primary or long_cont) and last_profitable_runner_exit["side"] == "SHORT":
                long_primary = long_cont = False
            if short_primary and last_profitable_runner_exit["side"] == "LONG":
                short_primary = False
        if not (long_primary or short_primary or long_cont):
            continue
        side = "LONG" if (long_primary or long_cont) else "SHORT"
        signal_type = "Continuation" if long_cont and not long_primary else "Primary"
        risk = c["atr14"] * 1.5
        entry = nxt["open"]
        pos = {
            "side": side,
            "signalType": signal_type,
            "signalDate": c["localDate"],
            "signalMs": int(c["time"]),
            "entryDate": nxt["localDate"],
            "entryMs": int(nxt["time"]),
            "entry": entry,
            "initialStop": entry - risk if side == "LONG" else entry + risk,
            "stop": entry - risk if side == "LONG" else entry + risk,
            "risk": risk,
            "tp": entry + c["atr14"] * 2.5 if side == "LONG" else entry - c["atr14"] * 2.5,
            "triggered": False,
            "tp1Time": "",
            "tp1Ms": None,
            "realizedR": 0.0,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": dist,
            "ema20": c["ema20"],
            "ema50": c["ema50"],
            "notes": "Primary NXT latest" if signal_type == "Primary" else "LONG-only pullback/touch EMA20 continuation",
        }
    return trades


def month_iter(start: date, end: date):
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        yield y, m
        m += 1
        if m == 13:
            y += 1
            m = 1


def fetch_monthly_funding(symbol: str, start: date, end: date) -> list[dict]:
    FUNDING_CACHE.mkdir(parents=True, exist_ok=True)
    start_ms = int(datetime(start.year, start.month, start.day, tzinfo=timezone.utc).timestamp() * 1000)
    end_ms = int(datetime(end.year, end.month, end.day, tzinfo=timezone.utc).timestamp() * 1000)
    path = FUNDING_CACHE / f"{symbol}_{start_ms}_{end_ms}.json"
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))

    rows = []
    for y, m in month_iter(start, end):
        url = f"https://data.binance.vision/data/futures/um/monthly/fundingRate/{symbol}/{symbol}-fundingRate-{y}-{m:02d}.zip"
        try:
            with urllib.request.urlopen(url, timeout=30) as response:
                payload = response.read()
        except urllib.error.HTTPError as exc:
            if exc.code == 404:
                continue
            raise
        with zipfile.ZipFile(io.BytesIO(payload)) as zf:
            names = [name for name in zf.namelist() if name.endswith(".csv")]
            if not names:
                continue
            with zf.open(names[0]) as raw:
                reader = csv.DictReader(io.TextIOWrapper(raw, encoding="utf-8"))
                for row in reader:
                    ts = int(float(row["calc_time"]))
                    if start_ms <= ts <= end_ms:
                        rows.append({"fundingTime": ts, "fundingRate": float(row["last_funding_rate"])})
        time.sleep(0.02)
    rows = sorted({r["fundingTime"]: r for r in rows}.values(), key=lambda r: r["fundingTime"])
    path.write_text(json.dumps(rows), encoding="utf-8")
    return rows


def funding_for_trade(trade: dict, funding_rows: list[dict]) -> dict:
    entry_ms = int(trade["entryMs"])
    exit_ms = int(trade["exitMs"])
    tp1_ms = int(trade["tp1Ms"]) if trade.get("tp1Ms") else None
    side_mult = -1 if trade["side"] == "LONG" else 1
    entry_to_risk = trade["entryPrice"] / trade["riskPerUnit"]
    total = 0.0
    count = 0
    paid = 0.0
    received = 0.0
    for row in funding_rows:
        ts = int(row["fundingTime"])
        if ts < entry_ms or ts > exit_ms:
            continue
        fraction = 0.5 if tp1_ms is not None and ts >= tp1_ms else 1.0
        funding_r = side_mult * float(row["fundingRate"]) * entry_to_risk * fraction
        total += funding_r
        count += 1
        if funding_r < 0:
            paid += funding_r
        else:
            received += funding_r
    return {"fundingR": total, "fundingEvents": count, "fundingPaidR": paid, "fundingReceivedR": received}


def profit_factor(rows: list[dict], key: str) -> float | None:
    gross_profit = sum(t[key] for t in rows if t[key] > 0)
    gross_loss = -sum(t[key] for t in rows if t[key] < 0)
    return gross_profit / gross_loss if gross_loss else None


def stats(rows: list[dict], key: str = "rMultiple") -> dict:
    result = base.stats(rows, key)
    result["profitFactor"] = profit_factor(rows, key)
    result["ending20k"] = STARTING_EQUITY + result["totalR"] * ONE_R_DOLLARS
    return result


def equity_curve(trades: list[dict], key: str) -> list[dict]:
    equity = STARTING_EQUITY
    peak = equity
    rows = []
    for i, trade in enumerate(trades, 1):
        pnl = trade[key] * ONE_R_DOLLARS
        equity += pnl
        peak = max(peak, equity)
        rows.append({
            "no": i,
            "timeframe": trade["timeframe"],
            "exitTime": trade["exitTime"],
            "side": trade["side"],
            "signalType": trade["signalType"],
            "rMultiple": trade[key],
            "pnl": pnl,
            "equity": equity,
            "drawdown": equity - peak,
        })
    return rows


def group_stats(rows: list[dict], key_fn, stat_key: str = "rMultiple") -> list[dict]:
    groups = {}
    for trade in rows:
        groups.setdefault(key_fn(trade), []).append(trade)
    out = []
    for key, subset in sorted(groups.items()):
        row = stats(subset, stat_key)
        row["group"] = key
        out.append(row)
    return out


def write_row(ws, row: int, values: list) -> None:
    for col, value in enumerate(values, 1):
        ws.cell(row, col).value = value


def style_sheet(ws, header_row: int = 4) -> None:
    ws.freeze_panes = f"A{header_row + 1}"
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(max(ws.column_dimensions[get_column_letter(col)].width or 12, 12), 26)


def build_workbook(result: dict) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "NXT Latest BTC - H1 H4 W1"
    summary["A2"] = "BTCUSDT only | same latest NXT v3.4 rule set | trading cost included | funding shown separately."
    write_row(summary, 4, ["Timeframe", "Trades", "Win Rate", "Total R", "Funding R", "Funding Adj R", "Max DD R", "Funding Adj Max DD R", "Profit Factor", "Funding Adj PF", "Ending 20K Adj"])
    for r, row in enumerate(result["byTimeframe"], 5):
        write_row(summary, r, [
            row["group"], row["trades"], row["winRate"], row["totalR"], row["fundingR"],
            row["fundingAdjustedTotalR"], row["maxDrawdownR"], row["fundingAdjustedMaxDrawdownR"],
            row["profitFactor"], row["fundingAdjustedProfitFactor"], row["fundingAdjustedEnding20k"],
        ])
    style_sheet(summary)

    trades_ws = wb.create_sheet("Trades")
    headers = ["TF", "No", "Signal Type", "Side", "Signal", "Entry", "Entry Price", "Initial Stop", "Risk/Unit", "TP1", "TP1 Time", "Exit", "Exit Price", "Exit Reason", "Gross R", "Cost R", "Net R", "Funding R", "Funding Events", "Funding Adj R", "ATR14", "RSI14", "Dist EMA50 ATR", "Notes"]
    write_row(trades_ws, 4, headers)
    for r, trade in enumerate(result["trades"], 5):
        write_row(trades_ws, r, [
            trade["timeframe"], trade["tradeNo"], trade["signalType"], trade["side"], trade["signalTime"],
            trade["entryTime"], trade["entryPrice"], trade["initialStop"], trade["riskPerUnit"],
            trade["tp1"], trade["tp1Time"], trade["exitTime"], trade["exitPrice"], trade["exitReason"],
            trade["grossRMultiple"], trade["costR"], trade["rMultiple"], trade["fundingR"],
            trade["fundingEvents"], trade["netRAfterFunding"], trade["atr14"], trade["rsi14"],
            trade["distanceToEma50Atr"], trade["notes"],
        ])
    style_sheet(trades_ws)
    trades_ws.column_dimensions["X"].width = 42

    for tf in ["H1", "H4", "W1"]:
        ws = wb.create_sheet(tf)
        ws["A1"] = f"{tf} Trades"
        write_row(ws, 4, headers)
        for r, trade in enumerate([t for t in result["trades"] if t["timeframe"] == tf], 5):
            write_row(ws, r, [
                trade["timeframe"], trade["tradeNo"], trade["signalType"], trade["side"], trade["signalTime"],
                trade["entryTime"], trade["entryPrice"], trade["initialStop"], trade["riskPerUnit"],
                trade["tp1"], trade["tp1Time"], trade["exitTime"], trade["exitPrice"], trade["exitReason"],
                trade["grossRMultiple"], trade["costR"], trade["rMultiple"], trade["fundingR"],
                trade["fundingEvents"], trade["netRAfterFunding"], trade["atr14"], trade["rsi14"],
                trade["distanceToEma50Atr"], trade["notes"],
            ])
        style_sheet(ws)
        ws.column_dimensions["X"].width = 42

    equity = wb.create_sheet("20K Account")
    equity["A1"] = "20K Account - Funding Adjusted"
    write_row(equity, 4, ["No", "TF", "Exit", "Side", "Signal Type", "R", "P&L $", "Equity $", "Drawdown $"])
    for r, row in enumerate(result["equityCurveFundingAdjusted"], 5):
        write_row(equity, r, [row["no"], row["timeframe"], row["exitTime"], row["side"], row["signalType"], row["rMultiple"], row["pnl"], row["equity"], row["drawdown"]])
    style_sheet(equity)

    assumptions = wb.create_sheet("Assumptions")
    write_row(assumptions, 4, ["#", "Assumption"])
    for r, line in enumerate(result["assumptions"], 5):
        write_row(assumptions, r, [r - 4, line])
    style_sheet(assumptions)
    assumptions.column_dimensions["B"].width = 82

    quality = wb.create_sheet("Data Quality")
    write_row(quality, 4, ["Timeframe", "Rows", "First Bar", "Last Bar", "Source"])
    for r, row in enumerate(result["datasets"], 5):
        write_row(quality, r, [row["timeframe"], row["rows"], row["firstBar"], row["lastBar"], row["source"]])
    style_sheet(quality)

    wb.save(OUT_XLSX)


def build_result() -> dict:
    all_trades = []
    datasets = []
    for interval in TIMEFRAMES:
        raw = candles_for_interval(interval)
        candles = enrich_with_ssl_period(raw, 14)
        datasets.append({
            "timeframe": interval.upper(),
            "rows": len(candles),
            "firstBar": candles[0]["localDate"],
            "lastBar": candles[-1]["localDate"],
            "source": f"Binance spot {interval} klines",
        })
        all_trades.extend(backtest_symbol_timeframe(interval, candles))
    all_trades.sort(key=lambda trade: (trade["exitMs"], trade["timeframe"]))

    funding_rows = fetch_monthly_funding(SYMBOL, START_DATE, END_DATE)
    for trade in all_trades:
        funding = funding_for_trade(trade, funding_rows)
        trade.update(funding)
        trade["netRAfterFunding"] = trade["rMultiple"] + trade["fundingR"]

    curve = equity_curve(all_trades, "netRAfterFunding")
    by_tf = []
    for row in group_stats(all_trades, lambda trade: trade["timeframe"]):
        subset = [t for t in all_trades if t["timeframe"] == row["group"]]
        adjusted = stats(subset, "netRAfterFunding")
        row["fundingR"] = sum(t["fundingR"] for t in subset)
        row["fundingAdjustedTotalR"] = adjusted["totalR"]
        row["fundingAdjustedMaxDrawdownR"] = adjusted["maxDrawdownR"]
        row["fundingAdjustedProfitFactor"] = adjusted["profitFactor"]
        row["fundingAdjustedEnding20k"] = adjusted["ending20k"]
        by_tf.append(row)

    original_stats = stats(all_trades)
    adjusted_stats = stats(all_trades, "netRAfterFunding")
    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "systemVersion": SYSTEM_VERSION,
        "period": {"start": START_DATE.isoformat(), "end": (END_DATE - timedelta(days=1)).isoformat(), "timezone": "Binance native candles"},
        "symbols": [SYMBOL],
        "timeframes": [tf.upper() for tf in TIMEFRAMES],
        "stats": original_stats,
        "fundingAdjustedStats": adjusted_stats,
        "fundingSummary": {
            "totalFundingR": sum(t["fundingR"] for t in all_trades),
            "fundingEvents": sum(t["fundingEvents"] for t in all_trades),
            "fundingPaidR": sum(t["fundingPaidR"] for t in all_trades),
            "fundingReceivedR": sum(t["fundingReceivedR"] for t in all_trades),
        },
        "account": {
            "startingEquity": STARTING_EQUITY,
            "oneRDollars": ONE_R_DOLLARS,
            "endingEquity": adjusted_stats["ending20k"],
            "netProfit": adjusted_stats["totalR"] * ONE_R_DOLLARS,
            "maxDrawdownDollars": min((r["drawdown"] for r in curve), default=0),
        },
        "byTimeframe": by_tf,
        "trades": all_trades,
        "equityCurveFundingAdjusted": curve,
        "datasets": datasets,
        "assumptions": RULES,
    }
    return result


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    result = build_result()
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    build_workbook(result)
    print(json.dumps({
        "outJson": str(OUT_JSON),
        "outXlsx": str(OUT_XLSX),
        "period": result["period"],
        "byTimeframe": result["byTimeframe"],
        "fundingAdjustedStats": result["fundingAdjustedStats"],
        "account": result["account"],
    }, indent=2))


if __name__ == "__main__":
    main()
