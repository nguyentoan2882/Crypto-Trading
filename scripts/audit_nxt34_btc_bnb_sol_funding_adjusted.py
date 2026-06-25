from __future__ import annotations

import csv
import io
import json
import shutil
import sys
import time
import urllib.error
import urllib.request
import zipfile
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

import test_nxt33_long_only_pullback_continuation as cont


ROOT = Path(__file__).resolve().parents[1]
LATEST_JSON = ROOT / "latest" / "NXT_Latest_NXT35_BTC_BNB_SOL_LongOnlyPullbackContinuation_20K.json"
SOURCE_JSON_FALLBACK = ROOT / "outputs" / "nxt35_btc_bnb_sol_latest" / "nxt35_btc_bnb_sol_latest_results.json"
FUNDING_CACHE = ROOT / "data_cache" / "binance_futures_funding"
OUT_DIR = ROOT / "outputs" / "nxt35_btc_bnb_sol_funding_adjusted"
OUT_JSON = OUT_DIR / "nxt35_btc_bnb_sol_funding_adjusted_results.json"
OUT_XLSX = OUT_DIR / "NXT35_BTC_BNB_SOL_20K_Funding_Adjusted.xlsx"
LATEST_DIR = ROOT / "latest"
LATEST_FUNDING_JSON = LATEST_DIR / "NXT_Latest_NXT35_BTC_BNB_SOL_FundingAdjusted_20K.json"
LATEST_FUNDING_XLSX = LATEST_DIR / "NXT_Latest_NXT35_BTC_BNB_SOL_FundingAdjusted_20K.xlsx"
LATEST_SOURCE_DOCX = LATEST_DIR / "NXT_Latest_NXT35_BTC_BNB_SOL_System_And_Indicators.docx"
LATEST_FUNDING_DOCX = LATEST_DIR / "NXT_Latest_NXT35_BTC_BNB_SOL_FundingAdjusted_System_And_Indicators.docx"
LATEST_SUMMARY = LATEST_DIR / "NXT_Latest_Summary.md"
STARTING_EQUITY = 20_000
ONE_R_DOLLARS = 1_000


def month_iter(start: date, end: date):
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        yield y, m
        m += 1
        if m == 13:
            y += 1
            m = 1


def funding_cache_path(symbol: str, start_ms: int, end_ms: int) -> Path:
    return FUNDING_CACHE / f"{symbol}_{start_ms}_{end_ms}.json"


def fetch_monthly_funding(symbol: str, start: date, end: date) -> list[dict]:
    FUNDING_CACHE.mkdir(parents=True, exist_ok=True)
    start_ms = int(datetime(start.year, start.month, start.day, tzinfo=timezone.utc).timestamp() * 1000)
    end_ms = int(datetime(end.year, end.month, end.day, tzinfo=timezone.utc).timestamp() * 1000)
    path = funding_cache_path(symbol, start_ms, end_ms)
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))

    rows = []
    for y, m in month_iter(start, end):
        url = f"https://data.binance.vision/data/futures/um/monthly/fundingRate/{symbol}/{symbol}-fundingRate-{y}-{m:02d}.zip"
        try:
            with urllib.request.urlopen(url, timeout=30) as response:
                payload = response.read()
        except urllib.error.HTTPError as exc:
            if exc.code == 404:
                continue
            raise
        with zipfile.ZipFile(io.BytesIO(payload)) as zf:
            names = [name for name in zf.namelist() if name.endswith(".csv")]
            if not names:
                continue
            with zf.open(names[0]) as raw:
                reader = csv.DictReader(io.TextIOWrapper(raw, encoding="utf-8"))
                for row in reader:
                    ts = int(float(row["calc_time"]))
                    if start_ms <= ts <= end_ms:
                        rows.append({
                            "symbol": symbol,
                            "fundingTime": ts,
                            "fundingRate": float(row["last_funding_rate"]),
                            "fundingIntervalHours": float(row.get("funding_interval_hours") or 8),
                            "markPrice": None,
                        })
        time.sleep(0.02)
    rows = sorted({r["fundingTime"]: r for r in rows}.values(), key=lambda r: r["fundingTime"])
    path.write_text(json.dumps(rows), encoding="utf-8")
    return rows


def ms_for_day(day: str) -> int:
    d = date.fromisoformat(day)
    return int(datetime(d.year, d.month, d.day, tzinfo=timezone.utc).timestamp() * 1000)


def funding_for_trade(trade: dict, funding_rows: list[dict]) -> dict:
    entry_ms = ms_for_day(trade["entryTime"])
    exit_ms = ms_for_day(trade["exitTime"]) + 86_400_000 - 1
    tp1_ms = ms_for_day(trade["tp1Time"]) if trade.get("tp1Time") else None
    side_mult = -1 if trade["side"] == "LONG" else 1
    entry_to_risk = trade["entryPrice"] / trade["riskPerUnit"]
    total = 0.0
    count = 0
    paid = 0.0
    received = 0.0
    for row in funding_rows:
        ts = int(row["fundingTime"])
        if ts < entry_ms or ts > exit_ms:
            continue
        fraction = 0.5 if tp1_ms is not None and ts >= tp1_ms else 1.0
        funding_r = side_mult * float(row["fundingRate"]) * entry_to_risk * fraction
        total += funding_r
        count += 1
        if funding_r < 0:
            paid += funding_r
        else:
            received += funding_r
    return {
        "fundingR": total,
        "fundingEvents": count,
        "fundingPaidR": paid,
        "fundingReceivedR": received,
    }


def equity_curve(trades: list[dict], key: str) -> list[dict]:
    equity = STARTING_EQUITY
    peak = equity
    rows = []
    for i, trade in enumerate(trades, 1):
        pnl = trade[key] * ONE_R_DOLLARS
        equity += pnl
        peak = max(peak, equity)
        rows.append({
            "no": i,
            "exitTime": trade["exitTime"],
            "symbol": trade["symbol"],
            "side": trade["side"],
            "signalType": trade["signalType"],
            "rMultiple": trade[key],
            "pnl": pnl,
            "equity": equity,
            "drawdown": equity - peak,
        })
    return rows


def trade_key(trade: dict) -> str:
    return f"{trade['symbol']}:{trade['tradeNo']}:{trade['entryTime']}:{trade['exitTime']}"


def compounding_curve(trades: list[dict], risk_pct: float, key: str) -> dict:
    events = {}
    for trade in trades:
        events.setdefault(trade["entryTime"], {"entries": [], "exits": []})["entries"].append(trade)
        events.setdefault(trade["exitTime"], {"entries": [], "exits": []})["exits"].append(trade)

    equity = STARTING_EQUITY
    peak = equity
    max_drawdown = 0.0
    max_drawdown_pct = 0.0
    open_risk = {}
    rows = {}
    for event_date in sorted(events):
        same_day_keys = {
            trade_key(trade)
            for trade in events[event_date]["entries"]
            if trade["exitTime"] == event_date
        }
        regular_exits = [
            trade for trade in events[event_date]["exits"]
            if trade_key(trade) not in same_day_keys
        ]
        same_day_exits = [
            trade for trade in events[event_date]["exits"]
            if trade_key(trade) in same_day_keys
        ]

        def close_trade(trade: dict) -> None:
            nonlocal equity, peak, max_drawdown, max_drawdown_pct
            identity = trade_key(trade)
            risk_amount = open_risk.pop(identity)
            pnl = risk_amount * trade[key]
            equity += pnl
            peak = max(peak, equity)
            drawdown = equity - peak
            max_drawdown = min(max_drawdown, drawdown)
            max_drawdown_pct = min(max_drawdown_pct, drawdown / peak if peak else 0)
            rows[identity] = {
                "riskAmount": risk_amount,
                "pnl": pnl,
                "equity": equity,
                "drawdown": drawdown,
            }

        for trade in sorted(regular_exits, key=lambda t: (t["symbol"], t["tradeNo"])):
            close_trade(trade)
        for trade in sorted(events[event_date]["entries"], key=lambda t: (t["symbol"], t["tradeNo"])):
            open_risk[trade_key(trade)] = equity * risk_pct
        for trade in sorted(same_day_exits, key=lambda t: (t["symbol"], t["tradeNo"])):
            close_trade(trade)
    return {
        "riskPct": risk_pct,
        "endingEquity": equity,
        "netProfit": equity - STARTING_EQUITY,
        "maxDrawdownDollars": max_drawdown,
        "maxDrawdownPct": max_drawdown_pct,
        "trades": rows,
    }


def stats_for_key(trades: list[dict], key: str) -> dict:
    rows = [dict(t, rMultiple=t[key]) for t in trades]
    return cont.enriched_stats(rows)


def write_row(ws, row: int, values: list) -> None:
    for col, value in enumerate(values, 1):
        ws.cell(row, col).value = value


def style_sheet(ws, header_row: int = 4) -> None:
    ws.freeze_panes = f"A{header_row + 1}"
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(max(ws.column_dimensions[get_column_letter(col)].width or 12, 12), 26)


def build_workbook(result: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "NXT v3.5 BTC BNB SOL - Funding Adjusted Audit"
    ws["A2"] = "Trading cost is already included in original R via costR. This audit adds Binance USD-M perpetual funding per trade."
    write_row(ws, 4, ["Metric", "Original", "Funding Adjusted", "Delta"])
    pairs = [
        ("Trades", "trades"),
        ("Win Rate", "winRate"),
        ("Total R", "totalR"),
        ("Average R", "avgR"),
        ("Max DD R", "maxDrawdownR"),
        ("Profit Factor", "profitFactor"),
        ("Ending Equity", "ending20k"),
    ]
    for r, (label, key) in enumerate(pairs, 5):
        old = result["originalStats"][key]
        new = result["fundingAdjustedStats"][key]
        write_row(ws, r, [label, old, new, new - old if isinstance(old, (int, float)) and isinstance(new, (int, float)) else None])
    write_row(ws, 14, ["Funding R Total", "", result["fundingSummary"]["totalFundingR"], result["fundingSummary"]["totalFundingR"]])
    write_row(ws, 15, ["Funding Events", "", result["fundingSummary"]["fundingEvents"], ""])
    write_row(ws, 17, ["Compound 2% Ending Equity", "", result["compounding2Pct"]["endingEquity"], ""])
    write_row(ws, 18, ["Compound 2% Max DD %", "", result["compounding2Pct"]["maxDrawdownPct"], ""])
    write_row(ws, 19, ["Compound 5% Ending Equity", "", result["compounding5Pct"]["endingEquity"], ""])
    write_row(ws, 20, ["Compound 5% Max DD %", "", result["compounding5Pct"]["maxDrawdownPct"], ""])
    style_sheet(ws)

    trades_ws = wb.create_sheet("Trades Funding")
    headers = ["Symbol", "No", "Type", "Side", "Signal", "Entry", "Exit", "Entry Price", "Risk/Unit", "TP1 Date", "Gross R", "Trading Cost R", "Original Net R", "Funding R", "Funding Events", "Adjusted Net R", "Adjusted P&L $", "Funding Paid R", "Funding Received R", "Exit Reason"]
    write_row(trades_ws, 4, headers)
    for r, trade in enumerate(result["trades"], 5):
        write_row(trades_ws, r, [
            trade["symbol"].replace("USDT", ""), trade["tradeNo"], trade["signalType"], trade["side"],
            trade["signalTime"], trade["entryTime"], trade["exitTime"], trade["entryPrice"], trade["riskPerUnit"],
            trade.get("tp1Time", ""), trade["grossRMultiple"], trade["costR"], trade["rMultiple"],
            trade["fundingR"], trade["fundingEvents"], trade["netRAfterFunding"], trade["netRAfterFunding"] * ONE_R_DOLLARS,
            trade["fundingPaidR"], trade["fundingReceivedR"], trade["exitReason"],
        ])
    style_sheet(trades_ws)

    by_symbol = wb.create_sheet("By Symbol")
    write_row(by_symbol, 4, ["Symbol", "Original R", "Funding R", "Adjusted R", "Funding Events"])
    for r, row in enumerate(result["fundingBySymbol"], 5):
        write_row(by_symbol, r, [row["symbol"].replace("USDT", ""), row["originalR"], row["fundingR"], row["adjustedR"], row["fundingEvents"]])
    style_sheet(by_symbol)

    account = wb.create_sheet("20K Account")
    account["A1"] = "20K Account - Funding Adjusted Trade Detail"
    account["A2"] = "Starting equity $20,000, 1R = $1,000. Trading cost is already included in Original Net R; Funding R is added separately."
    account_headers = [
        "No", "Symbol", "Signal Type", "Side", "Signal Date", "Entry Date", "Exit Date",
        "Entry Price", "Exit Price", "Original Net R", "Funding R", "Adjusted Net R",
        "Fixed P&L $", "Fixed Equity $", "Fixed Drawdown $",
        "Risk 2% $", "P&L 2% $", "Equity 2% $", "Drawdown 2% $",
        "Risk 5% $", "P&L 5% $", "Equity 5% $", "Drawdown 5% $",
        "Funding Events", "Exit Reason",
    ]
    write_row(account, 4, account_headers)
    adjusted_curve = result["equityCurveFundingAdjusted"]
    for r, (trade, curve_row) in enumerate(zip(result["trades"], adjusted_curve), 5):
        identity = trade_key(trade)
        comp2 = result["compounding2Pct"]["trades"][identity]
        comp5 = result["compounding5Pct"]["trades"][identity]
        write_row(account, r, [
            curve_row["no"],
            trade["symbol"].replace("USDT", ""),
            trade["signalType"],
            trade["side"],
            trade["signalTime"],
            trade["entryTime"],
            trade["exitTime"],
            trade["entryPrice"],
            trade["exitPrice"],
            trade["rMultiple"],
            trade["fundingR"],
            trade["netRAfterFunding"],
            curve_row["pnl"],
            curve_row["equity"],
            curve_row["drawdown"],
            comp2["riskAmount"],
            comp2["pnl"],
            comp2["equity"],
            comp2["drawdown"],
            comp5["riskAmount"],
            comp5["pnl"],
            comp5["equity"],
            comp5["drawdown"],
            trade["fundingEvents"],
            trade["exitReason"],
        ])
    style_sheet(account)
    account.column_dimensions["Y"].width = 34

    wb.save(OUT_XLSX)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    source_json = LATEST_JSON if LATEST_JSON.exists() else SOURCE_JSON_FALLBACK
    latest = json.loads(source_json.read_text(encoding="utf-8"))
    trades = [dict(t) for t in latest["trades"]]
    start = date.fromisoformat(latest["period"]["start"])
    end = date.fromisoformat(latest["period"]["end"])
    funding_by_symbol = {
        symbol: fetch_monthly_funding(symbol, start, end)
        for symbol in latest["symbols"]
    }
    for trade in trades:
        funding = funding_for_trade(trade, funding_by_symbol[trade["symbol"]])
        trade.update(funding)
        trade["netRAfterFunding"] = trade["rMultiple"] + trade["fundingR"]

    original_stats = latest["stats"]
    adjusted_stats = stats_for_key(trades, "netRAfterFunding")
    original_stats = dict(original_stats)
    adjusted_stats = dict(adjusted_stats)
    original_stats["ending20k"] = STARTING_EQUITY + original_stats["totalR"] * ONE_R_DOLLARS
    adjusted_stats["ending20k"] = STARTING_EQUITY + adjusted_stats["totalR"] * ONE_R_DOLLARS
    adjusted_curve = equity_curve(trades, "netRAfterFunding")
    adjusted_stats["maxDrawdownDollars"] = min((r["drawdown"] for r in adjusted_curve), default=0)
    compounding_2pct = compounding_curve(trades, 0.02, "netRAfterFunding")
    compounding_5pct = compounding_curve(trades, 0.05, "netRAfterFunding")

    symbol_rows = []
    for symbol in latest["symbols"]:
        subset = [t for t in trades if t["symbol"] == symbol]
        symbol_rows.append({
            "symbol": symbol,
            "originalR": sum(t["rMultiple"] for t in subset),
            "fundingR": sum(t["fundingR"] for t in subset),
            "adjustedR": sum(t["netRAfterFunding"] for t in subset),
            "fundingEvents": sum(t["fundingEvents"] for t in subset),
        })

    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "sourceLatestJson": str(source_json),
        "method": {
            "tradingCost": "Existing rMultiple already subtracts costR = round-trip fee/slippage model; this audit does not subtract it again.",
            "funding": "Funding is calculated from Binance USD-M futures fundingRate monthly public data. Position notional is approximated from entry price and risk per unit; after TP1, remaining notional is treated as 50%.",
            "sign": "Positive funding rates are paid by LONG and received by SHORT; negative rates are received by LONG and paid by SHORT.",
        },
        "symbols": latest["symbols"],
        "originalStats": original_stats,
        "fundingAdjustedStats": adjusted_stats,
        "fundingSummary": {
            "totalFundingR": sum(t["fundingR"] for t in trades),
            "fundingEvents": sum(t["fundingEvents"] for t in trades),
            "fundingPaidR": sum(t["fundingPaidR"] for t in trades),
            "fundingReceivedR": sum(t["fundingReceivedR"] for t in trades),
        },
        "fundingBySymbol": symbol_rows,
        "trades": trades,
        "equityCurveFundingAdjusted": adjusted_curve,
        "compounding2Pct": compounding_2pct,
        "compounding5Pct": compounding_5pct,
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    build_workbook(result)
    LATEST_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(OUT_JSON, LATEST_FUNDING_JSON)
    shutil.copy2(OUT_XLSX, LATEST_FUNDING_XLSX)
    if LATEST_SOURCE_DOCX.exists():
        shutil.copy2(LATEST_SOURCE_DOCX, LATEST_FUNDING_DOCX)
    LATEST_SUMMARY.write_text(
        "\n".join([
            "# Latest NXT System",
            "",
            f"System: {latest['systemVersion']} + funding-adjusted audit",
            "",
            f"Symbols: {', '.join(latest['symbols'])}",
            f"Trades: {adjusted_stats['trades']}",
            f"Original Total R before funding: {original_stats['totalR']:.2f}R",
            f"Funding R: {result['fundingSummary']['totalFundingR']:.2f}R",
            f"Funding-adjusted Total R: {adjusted_stats['totalR']:.2f}R",
            f"Funding-adjusted Max DD R: {adjusted_stats['maxDrawdownR']:.2f}R",
            f"Funding-adjusted win rate: {adjusted_stats['winRate']:.2%}",
            f"Funding-adjusted profit factor: {adjusted_stats['profitFactor']:.2f}",
            f"Starting equity: ${STARTING_EQUITY:,.2f}",
            f"1R: ${ONE_R_DOLLARS:,.2f}",
            f"Funding-adjusted ending equity: ${adjusted_stats['ending20k']:,.2f}",
            f"Funding-adjusted max DD dollars: ${adjusted_stats['maxDrawdownDollars']:,.2f}",
            f"Compounding 2% ending equity: ${compounding_2pct['endingEquity']:,.2f}",
            f"Compounding 2% max DD: {compounding_2pct['maxDrawdownPct']:.2%}",
            f"Compounding 5% ending equity: ${compounding_5pct['endingEquity']:,.2f}",
            f"Compounding 5% max DD: {compounding_5pct['maxDrawdownPct']:.2%}",
            "",
            "Anti-Immediate-Reversal: after a profitable runner exits by SSL flip, block the opposite entry on the exit candle and the next candle.",
            "",
            f"Workbook: {LATEST_FUNDING_XLSX.name}",
            f"JSON: {LATEST_FUNDING_JSON.name}",
            f"System doc: {LATEST_FUNDING_DOCX.name}",
        ]),
        encoding="utf-8",
    )
    print(json.dumps({
        "outJson": str(OUT_JSON),
        "outXlsx": str(OUT_XLSX),
        "originalTotalR": original_stats["totalR"],
        "fundingAdjustedTotalR": adjusted_stats["totalR"],
        "fundingR": result["fundingSummary"]["totalFundingR"],
        "originalEnding": original_stats["ending20k"],
        "adjustedEnding": adjusted_stats["ending20k"],
        "adjustedMaxDrawdownDollars": adjusted_stats["maxDrawdownDollars"],
        "compounding2PctEnding": compounding_2pct["endingEquity"],
        "compounding5PctEnding": compounding_5pct["endingEquity"],
    }, indent=2))


if __name__ == "__main__":
    main()
