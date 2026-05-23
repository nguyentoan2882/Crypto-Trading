from __future__ import annotations

import json
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\Workspace\Codex\Crypto trading")
SOURCE_JSON = ROOT / "outputs" / "nxt_v31_runner_ab_6y" / "nxt_v31_runner_ab_6y_results.json"
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
LATEST = ROOT / "latest"
OUT_JSON = LATEST / "NXT_Latest_NXT31_RunnerA_RiskOff_6Y_BTC_SOL_SUI_20K.json"
OUT_XLSX = LATEST / "NXT_Latest_NXT31_RunnerA_RiskOff_6Y_BTC_SOL_SUI_20K.xlsx"
OUT_MD = LATEST / "NXT_Latest_Summary.md"

STARTING_EQUITY = 20_000
RISK_PCT = 0.02
RISK_OFF_DD_R = -4.0
RISK_OFF_SIZE = 0.4


def iso_now() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def sheet_clear(ws, max_rows=300, max_cols=30):
    for row in ws.iter_rows(min_row=1, max_row=max(max_rows, ws.max_row), min_col=1, max_col=max(max_cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def set_row(ws, row_idx, values):
    for col_idx, value in enumerate(values, 1):
        ws.cell(row=row_idx, column=col_idx).value = value


def apply_template_layout(template_ws, target_ws):
    target_ws.sheet_format.defaultColWidth = template_ws.sheet_format.defaultColWidth
    target_ws.sheet_format.defaultRowHeight = template_ws.sheet_format.defaultRowHeight
    target_ws.freeze_panes = template_ws.freeze_panes
    target_ws.sheet_view.showGridLines = template_ws.sheet_view.showGridLines

    max_col = max(template_ws.max_column, target_ws.max_column)
    max_row = max(template_ws.max_row, target_ws.max_row)
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        src_dim = template_ws.column_dimensions.get(letter)
        dst_dim = target_ws.column_dimensions[letter]
        if src_dim and src_dim.width:
            dst_dim.width = src_dim.width
        dst_dim.hidden = src_dim.hidden if src_dim else False

    template_data_row = 5 if template_ws.max_row >= 5 else template_ws.max_row
    for row_idx in range(1, max_row + 1):
        src_idx = row_idx if row_idx <= template_ws.max_row else template_data_row
        src_dim = template_ws.row_dimensions.get(src_idx)
        dst_dim = target_ws.row_dimensions[row_idx]
        if src_dim and src_dim.height:
            dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden if src_dim else False

    for row_idx in range(1, max_row + 1):
        src_row = row_idx if row_idx <= template_ws.max_row else template_data_row
        for col_idx in range(1, max_col + 1):
            src_col = col_idx if col_idx <= template_ws.max_column else template_ws.max_column
            src = template_ws.cell(row=src_row, column=src_col)
            dst = target_ws.cell(row=row_idx, column=col_idx)
            if src.has_style:
                dst._style = copy(src._style)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)


def max_drawdown(rows, key):
    cumulative = 0.0
    peak = 0.0
    max_dd = 0.0
    for row in rows:
        cumulative += row[key]
        peak = max(peak, cumulative)
        max_dd = min(max_dd, cumulative - peak)
    return max_dd


def summarize(rows):
    out = []
    for symbol in ["BTCUSDT", "SOLUSDT", "SUIUSDT"]:
        subset = [r for r in rows if r["symbol"] == symbol]
        total = sum(r["riskOffR"] for r in subset)
        wins = sum(1 for r in subset if r["riskOffR"] > 0)
        out.append(
            {
                "symbol": symbol,
                "trades": len(subset),
                "wins": wins,
                "losses": len(subset) - wins,
                "winRate": wins / len(subset) if subset else 0,
                "totalR": total,
                "avgR": total / len(subset) if subset else 0,
                "bestR": max((r["riskOffR"] for r in subset), default=0),
                "worstR": min((r["riskOffR"] for r in subset), default=0),
            }
        )
    return out


def account_rows(rows):
    equity = STARTING_EQUITY
    peak = STARTING_EQUITY
    out = []
    for idx, row in enumerate(rows, 1):
        before = equity
        risk_usd = before * RISK_PCT
        pnl = risk_usd * row["riskOffR"]
        equity += pnl
        peak = max(peak, equity)
        out.append([idx, row["exitTime"], row["symbol"].replace("USDT", ""), row["side"], row["riskOffR"], before, risk_usd, pnl, equity, equity / peak - 1])
    return out, equity


def build():
    LATEST.mkdir(parents=True, exist_ok=True)
    source = json.loads(SOURCE_JSON.read_text(encoding="utf-8"))
    runner_a = next(v for v in source["variants"] if v["key"] == "runner_a_50_50_ssl")
    rows = sorted((dict(t) for t in runner_a["trades"]), key=lambda r: r["exitTime"])

    cumulative = 0.0
    peak = 0.0
    for row in rows:
        current_dd = cumulative - peak
        size_mult = RISK_OFF_SIZE if current_dd <= RISK_OFF_DD_R else 1.0
        row["preTradeCumulativeR"] = cumulative
        row["preTradeDrawdownR"] = current_dd
        row["sizeMultiplier"] = size_mult
        row["baseR"] = row["rMultiple"]
        row["riskOffR"] = row["rMultiple"] * size_mult
        cumulative += row["riskOffR"]
        peak = max(peak, cumulative)
        row["postTradeCumulativeR"] = cumulative
        row["postTradeDrawdownR"] = cumulative - peak

    total_r = sum(r["riskOffR"] for r in rows)
    wins = sum(1 for r in rows if r["riskOffR"] > 0)
    acct_rows, final_equity = account_rows(rows)
    result = {
        "generatedAt": iso_now(),
        "systemVersion": "NXT v3.1 + Runner A + Risk-Off",
        "source": source.get("source"),
        "period": source["period"],
        "symbols": source["symbols"],
        "riskOffRule": {
            "triggerDrawdownR": RISK_OFF_DD_R,
            "sizeMultiplierWhenTriggered": RISK_OFF_SIZE,
            "description": "If the closed-trade equity curve is at or below -4R from its peak before a new trade closes, that trade is counted at 40% size. Otherwise it is counted at full size.",
        },
        "stats": {
            "trades": len(rows),
            "wins": wins,
            "losses": len(rows) - wins,
            "winRate": wins / len(rows),
            "totalR": total_r,
            "avgR": total_r / len(rows),
            "maxDrawdownR": max_drawdown(rows, "riskOffR"),
            "baseTotalR": sum(r["baseR"] for r in rows),
            "baseMaxDrawdownR": max_drawdown(rows, "baseR"),
            "riskOffTrades": sum(1 for r in rows if r["sizeMultiplier"] < 1),
            "finalEquity20K": final_equity,
            "compoundReturn": final_equity / STARTING_EQUITY - 1,
        },
        "summary": summarize(rows),
        "trades": rows,
        "datasets": source["datasets"],
        "template": str(TEMPLATE),
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")

    template_wb = load_workbook(TEMPLATE)
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        sheet_clear(ws)

    summary = wb["Summary"]
    summary["A1"] = "Latest - NXT v3.1 Runner A + Risk-Off"
    summary["A2"] = "BTC/SOL/SUI 6Y | 20K account | risk-off: DD <= -4R uses 40% size"
    for r, vals in enumerate(
        [
            ["Metric", "Value"],
            ["System", result["systemVersion"]],
            ["Total trades", result["stats"]["trades"]],
            ["Win rate", result["stats"]["winRate"]],
            ["Total R", result["stats"]["totalR"]],
            ["Average R / trade", result["stats"]["avgR"]],
            ["Max DD R", result["stats"]["maxDrawdownR"]],
            ["Risk-off trades", result["stats"]["riskOffTrades"]],
            ["Final equity 20K", result["stats"]["finalEquity20K"]],
            ["Compound return", result["stats"]["compoundReturn"]],
            ["Base Runner A Total R", result["stats"]["baseTotalR"]],
            ["Base Runner A Max DD R", result["stats"]["baseMaxDrawdownR"]],
        ],
        4,
    ):
        set_row(summary, r, vals)
    set_row(summary, 4, ["Metric", "Value", "", "Symbol", "Trades", "Wins", "Losses", "Win Rate", "Total R", "Avg R", "Best R", "Worst R"])
    for i, row in enumerate(result["summary"], 5):
        set_row(summary, i, [None, None, None, row["symbol"].replace("USDT", ""), row["trades"], row["wins"], row["losses"], row["winRate"], row["totalR"], row["avgR"], row["bestR"], row["worstR"]])

    headers = [
        "Symbol", "No", "Side", "Signal Time", "Entry Time", "Entry Price", "Initial Stop", "Final Stop", "Risk / Unit",
        "TP", "TP Time", "TP Ref", "Exit Time", "Exit Price", "Exit Reason", "Base R", "Size Mult", "Risk-Off R",
        "Pre DD R", "Post DD R", "TP Hit", "EMA20", "EMA50", "ATR14", "SSL Signal", "Net Volume", "Distance EMA50 ATR", "Notes",
    ]

    def trade_values(t):
        return [
            t["symbol"].replace("USDT", ""), t["tradeNo"], t["side"], t["signalTime"].replace("T", " ").replace("Z", ""),
            t["entryTime"].replace("T", " ").replace("Z", ""), t["entryPrice"], t["initialStop"], t["finalStop"], t["riskPerUnit"],
            t["tp1"], t["tp1Time"].replace("T", " ").replace("Z", "") if t["tp1Time"] else "", t["tp2"],
            t["exitTime"].replace("T", " ").replace("Z", ""), t["exitPrice"], t["exitReason"], t["baseR"], t["sizeMultiplier"],
            t["riskOffR"], t["preTradeDrawdownR"], t["postTradeDrawdownR"], t["tp1Hit"], t["ema20"], t["ema50"], t["atr14"],
            t["sslAtSignal"], t["netVolume"], t["distanceToEma50Atr"], t.get("notes", ""),
        ]

    for sheet_name, sheet_rows, title in [
        ("Trades", rows, "Detailed Trades - Latest Risk-Off"),
        ("BTC", [r for r in rows if r["symbol"] == "BTCUSDT"], "BTC Trades - Latest Risk-Off"),
        ("SOL", [r for r in rows if r["symbol"] == "SOLUSDT"], "SOL Trades - Latest Risk-Off"),
        ("SUI", [r for r in rows if r["symbol"] == "SUIUSDT"], "SUI Trades - Latest Risk-Off"),
    ]:
        ws = wb[sheet_name]
        ws["A1"] = title
        ws["A2"] = "One completed trade per row; Base R is Runner A, Risk-Off R applies size multiplier."
        set_row(ws, 4, headers)
        for i, trade in enumerate(sheet_rows, 5):
            set_row(ws, i, trade_values(trade))

    account = wb["20K Account"]
    account["A1"] = "20K Account - Latest Risk-Off"
    account["A2"] = "Compounded sequence at 2.0% risk per trade using Risk-Off R."
    set_row(account, 4, ["Trade", "Exit Time", "Symbol", "Side", "Risk-Off R", "Equity Before", "Risk USD", "P/L USD", "Equity After", "Drawdown"])
    for i, vals in enumerate(acct_rows, 5):
        vals = list(vals)
        vals[1] = vals[1].replace("T", " ").replace("Z", "")
        set_row(account, i, vals)

    assumptions = wb["Assumptions"]
    assumptions["A1"] = "Backtest Assumptions"
    assumptions["A2"] = "Latest selected system."
    assumption_rows = [
        ["#", "Assumption"],
        [1, "System is NXT v3.1 with RSI 50/50 and Runner A."],
        [2, "Runner A closes 50% at 2.5 ATR and keeps 50% runner until opposite SSL flip or breakeven stop."],
        [3, "Risk-off overlay is applied on the closed-trade equity curve."],
        [4, "When pre-trade equity drawdown is <= -4R, trade size is reduced to 40%."],
        [5, "When drawdown is above -4R, trade size returns to 100%."],
        [6, "The overlay changes sizing only; it does not change entry or exit signals."],
    ]
    for i, vals in enumerate(assumption_rows, 4):
        set_row(assumptions, i, vals)

    quality = wb["Data Quality"]
    quality["A1"] = "Data Quality"
    quality["A2"] = "Binance spot daily candles from project cache."
    set_row(quality, 4, ["Symbol", "Daily Candles", "Weekly Candles", "First Daily", "Last Daily", "Source"])
    for i, (symbol, q) in enumerate(result["datasets"].items(), 5):
        set_row(quality, i, [symbol.replace("USDT", ""), q["dailyCount"], q["weeklyCount"], q["firstDaily"][:10], q["lastDaily"][:10], q["source"]])

    for sheet_name in ["Summary", "Trades", "BTC", "SOL", "SUI", "Equity Curve", "20K Account", "Assumptions", "Data Quality"]:
        if sheet_name in template_wb.sheetnames and sheet_name in wb.sheetnames:
            apply_template_layout(template_wb[sheet_name], wb[sheet_name])

    wb.save(OUT_XLSX)

    OUT_MD.write_text(
        "\n".join(
            [
                "# Latest NXT System",
                "",
                "System: NXT v3.1 + RSI 50/50 + Runner A + Risk-Off",
                "",
                f"Total R: {result['stats']['totalR']:.2f}R",
                f"Max DD R: {result['stats']['maxDrawdownR']:.2f}R",
                f"Trades: {result['stats']['trades']}",
                f"Win rate: {result['stats']['winRate']:.2%}",
                f"Final 20K equity: ${result['stats']['finalEquity20K']:,.0f}",
                "",
                "Risk-off rule: if the closed-trade equity curve is at or below -4R from peak, new trade size is reduced to 40%; otherwise use full size.",
                "",
                f"Workbook: {OUT_XLSX.name}",
                f"JSON: {OUT_JSON.name}",
            ]
        ),
        encoding="utf-8",
    )

    shutil.copy2(TEMPLATE, LATEST / TEMPLATE.name)


if __name__ == "__main__":
    build()
