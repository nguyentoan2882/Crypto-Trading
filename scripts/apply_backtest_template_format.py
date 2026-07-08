from __future__ import annotations

import copy
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\Workspace\Codex\Crypto trading")
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"


def copy_sheet_layout(template_ws, target_ws) -> None:
    target_ws.sheet_format.defaultColWidth = template_ws.sheet_format.defaultColWidth
    target_ws.sheet_format.defaultRowHeight = template_ws.sheet_format.defaultRowHeight
    target_ws.freeze_panes = template_ws.freeze_panes
    target_ws.sheet_view.showGridLines = template_ws.sheet_view.showGridLines

    max_col = max(template_ws.max_column, target_ws.max_column)
    max_row = max(template_ws.max_row, target_ws.max_row)

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        src_dim = template_ws.column_dimensions.get(letter)
        dst_dim = target_ws.column_dimensions[letter]
        if src_dim and src_dim.width:
            dst_dim.width = src_dim.width
        dst_dim.hidden = src_dim.hidden if src_dim else False

    template_data_row = 5 if template_ws.max_row >= 5 else template_ws.max_row
    for row_idx in range(1, max_row + 1):
        src_idx = row_idx if row_idx <= template_ws.max_row else template_data_row
        src_dim = template_ws.row_dimensions.get(src_idx)
        dst_dim = target_ws.row_dimensions[row_idx]
        if src_dim and src_dim.height:
            dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden if src_dim else False

    for row_idx in range(1, max_row + 1):
        src_row = row_idx if row_idx <= template_ws.max_row else template_data_row
        for col_idx in range(1, max_col + 1):
            src_col = col_idx if col_idx <= template_ws.max_column else template_ws.max_column
            src = template_ws.cell(row=src_row, column=src_col)
            dst = target_ws.cell(row=row_idx, column=col_idx)
            if src.has_style:
                dst._style = copy.copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.alignment:
                dst.alignment = copy.copy(src.alignment)
            if src.font:
                dst.font = copy.copy(src.font)
            if src.fill:
                dst.fill = copy.copy(src.fill)
            if src.border:
                dst.border = copy.copy(src.border)


def apply_template(target_path: Path, output_path: Path | None = None) -> None:
    template_wb = load_workbook(TEMPLATE)
    target_wb = load_workbook(target_path)

    for sheet_name in [
        "Summary",
        "Compounding Summary",
        "Compounding Yearly",
        "Compounding Trades",
        "Trades",
        "BTC",
        "SOL",
        "SUI",
        "Equity Curve",
        "20K Account",
        "Assumptions",
        "Data Quality",
    ]:
        if sheet_name in template_wb.sheetnames and sheet_name in target_wb.sheetnames:
            copy_sheet_layout(template_wb[sheet_name], target_wb[sheet_name])

    # Excel rejects a worksheet view that keeps selection.pane="bottomLeft"
    # after freeze_panes has been removed. openpyxl can save that inconsistent
    # state without warning, which makes Excel show a recovery dialog.
    for ws in target_wb.worksheets:
        if ws.freeze_panes is None:
            for selection in ws.sheet_view.selection:
                selection.pane = None

    target_wb.save(output_path or target_path)


if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        raise SystemExit("Usage: apply_backtest_template_format.py <xlsx_path> [output_xlsx_path]")
    apply_template(Path(sys.argv[1]), Path(sys.argv[2]) if len(sys.argv) == 3 else None)
