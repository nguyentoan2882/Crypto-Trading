from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"

SHEETS = {
    "Compounding Summary": [
        "Scenario",
        "Risk Model",
        "Ending Equity",
        "Net Profit",
        "Max DD $",
        "Max DD % Peak",
        "Trades",
    ],
    "Compounding Yearly": [
        "Year",
        "Fixed Latest P&L",
        "Fixed Latest End Equity",
        "Compound 2% P&L",
        "Compound 2% End Equity",
        "Compound 5% P&L",
        "Compound 5% End Equity",
    ],
    "Compounding Trades": [
        "Seq",
        "Symbol",
        "Trade #",
        "Signal Type",
        "Side",
        "Entry Date",
        "Exit Date",
        "Net R",
        "Fixed Risk",
        "Fixed P&L",
        "Fixed Equity",
        "2% Risk",
        "2% P&L",
        "2% Equity",
        "5% Risk",
        "5% P&L",
        "5% Equity",
    ],
}


def style_header(ws, row: int = 4) -> None:
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def setup_sheet(ws, title: str, subtitle: str, headers: list[str]) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    ws["A1"] = title
    ws["A2"] = subtitle
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"].font = Font(color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True)
    for col, header in enumerate(headers, 1):
        ws.cell(4, col).value = header
    style_header(ws)
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A4:{ws.cell(4, len(headers)).coordinate}"


def main() -> None:
    wb = load_workbook(TEMPLATE)
    for name in SHEETS:
        if name in wb.sheetnames:
            del wb[name]

    summary = wb.create_sheet("Compounding Summary", 1)
    setup_sheet(
        summary,
        "NXT Compounding Comparison",
        "Funding-adjusted NXT results. Compounding risk is fixed from realized equity at entry and P&L is recognized at exit.",
        SHEETS["Compounding Summary"],
    )
    summary_widths = [27, 28, 18, 18, 18, 17, 10]
    for col, width in enumerate(summary_widths, 1):
        summary.column_dimensions[summary.cell(1, col).column_letter].width = width
    for row in range(5, 8):
        for col in range(3, 6):
            summary.cell(row, col).number_format = '$#,##0'
        summary.cell(row, 6).number_format = "0.0%"

    yearly = wb.create_sheet("Compounding Yearly", 2)
    setup_sheet(
        yearly,
        "NXT Compounding by Year",
        "Annual realized P&L and ending equity for fixed latest, compound 2%, and compound 5%.",
        SHEETS["Compounding Yearly"],
    )
    yearly.column_dimensions["A"].width = 10
    for col in range(2, 8):
        yearly.column_dimensions[yearly.cell(1, col).column_letter].width = 22
        for row in range(5, 20):
            yearly.cell(row, col).number_format = '$#,##0'
    chart = LineChart()
    chart.title = "Ending Equity by Year"
    chart.y_axis.title = "Equity"
    chart.x_axis.title = "Year"
    chart.y_axis.numFmt = '$#,##0'
    chart.add_data(Reference(yearly, min_col=3, max_col=7, min_row=4, max_row=12), titles_from_data=True)
    chart.set_categories(Reference(yearly, min_col=1, min_row=5, max_row=12))
    chart.height = 9
    chart.width = 18
    yearly.add_chart(chart, "I4")

    trades = wb.create_sheet("Compounding Trades", 3)
    setup_sheet(
        trades,
        "NXT Compounding Trade Detail",
        "One completed trade per row. Risk amount is locked at entry; P&L and equity update at exit.",
        SHEETS["Compounding Trades"],
    )
    widths = [8, 12, 10, 15, 10, 13, 13, 10, 14, 14, 16, 14, 14, 16, 14, 14, 16]
    for col, width in enumerate(widths, 1):
        trades.column_dimensions[trades.cell(1, col).column_letter].width = width
    for row in range(5, 505):
        trades.cell(row, 8).number_format = "0.0000"
        for col in range(9, 18):
            trades.cell(row, col).number_format = '$#,##0'
    trades.auto_filter.ref = "A4:Q504"

    wb.save(TEMPLATE)
    print(TEMPLATE)


if __name__ == "__main__":
    main()
