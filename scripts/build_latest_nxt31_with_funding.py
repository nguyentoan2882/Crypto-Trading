from __future__ import annotations

import json
import time
import urllib.parse
import urllib.request
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\Workspace\Codex\Crypto trading")
LATEST = ROOT / "latest"
SOURCE_JSON = LATEST / "NXT_Latest_NXT31_RunnerA_RiskOff_6Y_BTC_SOL_SUI_20K.json"
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
CACHE = ROOT / "data_cache" / "binance_futures_funding"
OUT_JSON = LATEST / "NXT_Latest_NXT31_RunnerA_RiskOff_Funding_6Y_BTC_SOL_SUI_20K.json"
OUT_XLSX = LATEST / "NXT_Latest_NXT31_RunnerA_RiskOff_Funding_6Y_BTC_SOL_SUI_20K.xlsx"


def parse_ms(value: str) -> int:
    return int(datetime.fromisoformat(value.replace("Z", "+00:00")).timestamp() * 1000)


def iso_ms(ms: int) -> str:
    return datetime.utcfromtimestamp(ms / 1000).replace(microsecond=0).isoformat() + "Z"


def fetch_funding(symbol: str, start_ms: int, end_ms: int):
    CACHE.mkdir(parents=True, exist_ok=True)
    cache_path = CACHE / f"{symbol}_{start_ms}_{end_ms}.json"
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))

    rows = []
    cursor = start_ms
    while cursor <= end_ms:
        params = urllib.parse.urlencode(
            {
                "symbol": symbol,
                "startTime": cursor,
                "endTime": end_ms,
                "limit": 1000,
            }
        )
        url = f"https://fapi.binance.com/fapi/v1/fundingRate?{params}"
        with urllib.request.urlopen(url, timeout=30) as response:
            batch = json.loads(response.read().decode("utf-8"))
        if not batch:
            break
        rows.extend(batch)
        next_cursor = int(batch[-1]["fundingTime"]) + 1
        if next_cursor <= cursor:
            break
        cursor = next_cursor
        time.sleep(0.08)

    normalized = [
        {
            "symbol": r["symbol"],
            "fundingTime": int(r["fundingTime"]),
            "fundingRate": float(r["fundingRate"]),
            "markPrice": float(r["markPrice"]) if r.get("markPrice") not in ("", None) else None,
        }
        for r in rows
    ]
    cache_path.write_text(json.dumps(normalized, indent=2), encoding="utf-8")
    return normalized


def apply_template_layout(template_ws, target_ws):
    target_ws.sheet_format.defaultColWidth = template_ws.sheet_format.defaultColWidth
    target_ws.sheet_format.defaultRowHeight = template_ws.sheet_format.defaultRowHeight
    target_ws.freeze_panes = template_ws.freeze_panes
    target_ws.sheet_view.showGridLines = template_ws.sheet_view.showGridLines
    max_col = max(template_ws.max_column, target_ws.max_column)
    max_row = max(template_ws.max_row, target_ws.max_row)
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        src_dim = template_ws.column_dimensions.get(letter)
        dst_dim = target_ws.column_dimensions[letter]
        if src_dim and src_dim.width:
            dst_dim.width = src_dim.width
        dst_dim.hidden = src_dim.hidden if src_dim else False
    template_data_row = 5 if template_ws.max_row >= 5 else template_ws.max_row
    for row_idx in range(1, max_row + 1):
        src_idx = row_idx if row_idx <= template_ws.max_row else template_data_row
        src_dim = template_ws.row_dimensions.get(src_idx)
        dst_dim = target_ws.row_dimensions[row_idx]
        if src_dim and src_dim.height:
            dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden if src_dim else False
    for row_idx in range(1, max_row + 1):
        src_row = row_idx if row_idx <= template_ws.max_row else template_data_row
        for col_idx in range(1, max_col + 1):
            src_col = col_idx if col_idx <= template_ws.max_column else template_ws.max_column
            src = template_ws.cell(row=src_row, column=src_col)
            dst = target_ws.cell(row=row_idx, column=col_idx)
            if src.has_style:
                dst._style = copy(src._style)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)


def clear(ws, rows=320, cols=34):
    for row in ws.iter_rows(min_row=1, max_row=max(rows, ws.max_row), min_col=1, max_col=max(cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def rowset(ws, row_idx, values):
    for col_idx, value in enumerate(values, 1):
        ws.cell(row=row_idx, column=col_idx).value = value


def max_dd(rows, key):
    cum = peak = dd = 0.0
    for row in rows:
        cum += row[key]
        peak = max(peak, cum)
        dd = min(dd, cum - peak)
    return dd


def summarize(rows, key):
    out = []
    for symbol in ["BTCUSDT", "SOLUSDT", "SUIUSDT"]:
        subset = [r for r in rows if r["symbol"] == symbol]
        total = sum(r[key] for r in subset)
        wins = sum(1 for r in subset if r[key] > 0)
        out.append(
            {
                "symbol": symbol,
                "trades": len(subset),
                "wins": wins,
                "losses": len(subset) - wins,
                "winRate": wins / len(subset) if subset else 0,
                "totalR": total,
                "avgR": total / len(subset) if subset else 0,
                "fundingR": sum(r["fundingR"] for r in subset),
            }
        )
    return out


def account(rows):
    equity = 20000.0
    peak = equity
    acct = []
    for idx, row in enumerate(rows, 1):
        before = equity
        risk = before * 0.02
        pnl = risk * row["netRAfterFunding"]
        equity += pnl
        peak = max(peak, equity)
        acct.append([idx, row["exitTime"], row["symbol"].replace("USDT", ""), row["side"], row["netRAfterFunding"], before, risk, pnl, equity, equity / peak - 1])
    return acct, equity


def build():
    data = json.loads(SOURCE_JSON.read_text(encoding="utf-8"))
    trades = sorted((dict(t) for t in data["trades"]), key=lambda t: t["exitTime"])
    start = min(parse_ms(t["entryTime"]) for t in trades) - 8 * 60 * 60 * 1000
    end = max(parse_ms(t["exitTime"]) for t in trades) + 8 * 60 * 60 * 1000
    funding = {symbol: fetch_funding(symbol, start, end) for symbol in data["symbols"]}

    for trade in trades:
        entry = parse_ms(trade["entryTime"])
        exit_ = parse_ms(trade["exitTime"])
        rows = [r for r in funding[trade["symbol"]] if entry < r["fundingTime"] <= exit_]
        sum_rate = sum(r["fundingRate"] for r in rows)
        notional_to_risk = trade["entryPrice"] / trade["riskPerUnit"]
        side_sign = 1 if trade["side"] == "LONG" else -1
        funding_r_full_size = side_sign * sum_rate * notional_to_risk
        trade["fundingEvents"] = len(rows)
        trade["fundingRateSum"] = sum_rate
        trade["notionalToRisk"] = notional_to_risk
        trade["fundingR"] = funding_r_full_size * trade["sizeMultiplier"]
        trade["netRAfterFunding"] = trade["riskOffR"] - trade["fundingR"]

    total_funding = sum(t["fundingR"] for t in trades)
    total_after = sum(t["netRAfterFunding"] for t in trades)
    acct, final_equity = account(trades)
    result = {
        **data,
        "generatedAt": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "systemVersion": data["systemVersion"] + " + Actual Binance Futures Funding",
        "fundingSource": "Binance USD-M Futures fundingRate API",
        "statsAfterFunding": {
            "trades": len(trades),
            "totalRBeforeFunding": data["stats"]["totalR"],
            "fundingR": total_funding,
            "totalRAfterFunding": total_after,
            "maxDrawdownRAfterFunding": max_dd(trades, "netRAfterFunding"),
            "finalEquity20KAfterFunding": final_equity,
            "compoundReturnAfterFunding": final_equity / 20000 - 1,
        },
        "summaryAfterFunding": summarize(trades, "netRAfterFunding"),
        "trades": trades,
        "fundingDataQuality": {
            symbol: {
                "events": len(rows),
                "first": iso_ms(rows[0]["fundingTime"]) if rows else "",
                "last": iso_ms(rows[-1]["fundingTime"]) if rows else "",
                "sumFundingRate": sum(r["fundingRate"] for r in rows),
            }
            for symbol, rows in funding.items()
        },
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")

    template_wb = load_workbook(TEMPLATE)
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        clear(ws)

    summary = wb["Summary"]
    summary["A1"] = "Latest - NXT v3.1 Risk-Off With Funding"
    summary["A2"] = "Actual Binance USD-M Futures funding applied per held trade."
    rows = [
        ["Metric", "Value"],
        ["Total R before funding", data["stats"]["totalR"]],
        ["Funding R impact", total_funding],
        ["Total R after funding", total_after],
        ["Max DD R before funding", data["stats"]["maxDrawdownR"]],
        ["Max DD R after funding", result["statsAfterFunding"]["maxDrawdownRAfterFunding"]],
        ["Final 20K before funding", data["stats"]["finalEquity20K"]],
        ["Final 20K after funding", final_equity],
    ]
    for i, r in enumerate(rows, 4):
        rowset(summary, i, r)
    rowset(summary, 4, ["Metric", "Value", "", "Symbol", "Trades", "Win Rate", "Funding R", "Total R After Funding"])
    for i, r in enumerate(result["summaryAfterFunding"], 5):
        rowset(summary, i, [None, None, None, r["symbol"].replace("USDT", ""), r["trades"], r["winRate"], r["fundingR"], r["totalR"]])

    headers = [
        "Symbol", "No", "Side", "Signal Time", "Entry Time", "Entry Price", "Initial Stop", "Final Stop", "Risk / Unit",
        "TP", "TP Time", "TP Ref", "Exit Time", "Exit Price", "Exit Reason", "Risk-Off R", "Funding Events",
        "Funding Rate Sum", "Notional/Risk", "Funding R", "Net R After Funding", "Size Mult", "Pre DD R", "Post DD R",
        "TP Hit", "ATR14", "Distance EMA50 ATR", "Notes",
    ]

    def trade_row(t):
        return [
            t["symbol"].replace("USDT", ""), t["tradeNo"], t["side"], t["signalTime"].replace("T", " ").replace("Z", ""),
            t["entryTime"].replace("T", " ").replace("Z", ""), t["entryPrice"], t["initialStop"], t["finalStop"], t["riskPerUnit"],
            t["tp1"], t["tp1Time"].replace("T", " ").replace("Z", "") if t["tp1Time"] else "", t["tp2"],
            t["exitTime"].replace("T", " ").replace("Z", ""), t["exitPrice"], t["exitReason"], t["riskOffR"], t["fundingEvents"],
            t["fundingRateSum"], t["notionalToRisk"], t["fundingR"], t["netRAfterFunding"], t["sizeMultiplier"],
            t["preTradeDrawdownR"], t["postTradeDrawdownR"], t["tp1Hit"], t["atr14"], t["distanceToEma50Atr"], t.get("notes", ""),
        ]

    for sheet_name, subset in [
        ("Trades", trades),
        ("BTC", [t for t in trades if t["symbol"] == "BTCUSDT"]),
        ("SOL", [t for t in trades if t["symbol"] == "SOLUSDT"]),
        ("SUI", [t for t in trades if t["symbol"] == "SUIUSDT"]),
    ]:
        ws = wb[sheet_name]
        ws["A1"] = f"{sheet_name} Trades - With Funding" if sheet_name != "Trades" else "Detailed Trades - With Funding"
        ws["A2"] = "Funding R: positive value is a cost; negative value is a credit."
        rowset(ws, 4, headers)
        for i, t in enumerate(subset, 5):
            rowset(ws, i, trade_row(t))

    acct_ws = wb["20K Account"]
    acct_ws["A1"] = "20K Account - After Funding"
    acct_ws["A2"] = "Compounded sequence at 2.0% risk per trade using net R after funding."
    rowset(acct_ws, 4, ["Trade", "Exit Time", "Symbol", "Side", "Net R After Funding", "Equity Before", "Risk USD", "P/L USD", "Equity After", "Drawdown"])
    for i, r in enumerate(acct, 5):
        rr = list(r)
        rr[1] = rr[1].replace("T", " ").replace("Z", "")
        rowset(acct_ws, i, rr)

    assumptions = wb["Assumptions"]
    assumptions["A1"] = "Funding Assumptions"
    assumptions["A2"] = "Actual Binance USD-M Futures funding history."
    for i, r in enumerate(
        [
            ["#", "Assumption"],
            [1, "Funding rate is applied for events with entryTime < fundingTime <= exitTime."],
            [2, "LONG pays positive funding and receives negative funding."],
            [3, "SHORT receives positive funding and pays negative funding."],
            [4, "Funding R = fundingRateSum x entryNotional/riskPerUnit x risk-off size multiplier."],
            [5, "Spot backtest prices are retained; funding is overlaid as a futures carrying cost/credit."],
        ],
        4,
    ):
        rowset(assumptions, i, r)

    quality = wb["Data Quality"]
    quality["A1"] = "Data Quality"
    quality["A2"] = "Candle source plus Binance funding data coverage."
    rowset(quality, 4, ["Symbol", "Funding Events", "First Funding", "Last Funding", "Sum Funding Rate", "Source"])
    for i, (symbol, q) in enumerate(result["fundingDataQuality"].items(), 5):
        rowset(quality, i, [symbol.replace("USDT", ""), q["events"], q["first"][:10], q["last"][:10], q["sumFundingRate"], result["fundingSource"]])

    for sheet in ["Summary", "Trades", "BTC", "SOL", "SUI", "Equity Curve", "20K Account", "Assumptions", "Data Quality"]:
        if sheet in wb.sheetnames and sheet in template_wb.sheetnames:
            apply_template_layout(template_wb[sheet], wb[sheet])

    wb.save(OUT_XLSX)


if __name__ == "__main__":
    build()
