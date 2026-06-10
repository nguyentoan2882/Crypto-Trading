from __future__ import annotations

import json
import sys
from datetime import date, datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import backtest_nxt31_utc7_latest as base
import backtest_nxt32_native_1d_latest as native
import rebuild_nxt32_native_1d_tv_atr_latest as tv_atr
from openpyxl import load_workbook


ROOT = Path.cwd()
OUT_DIR = ROOT / "outputs" / "nxt32_rsi_51_49"
OUT_JSON = OUT_DIR / "nxt32_rsi_51_49_results.json"
OUT_XLSX = OUT_DIR / "NXT32_RSI_51_49_6Y_BTC_SOL_SUI_20K.xlsx"
BASELINE_JSON = ROOT / "latest" / "NXT_Latest_NXT32_Native1D_RunnerA_NoContinuation_NoRiskOff_6Y_BTC_SOL_SUI_20K.json"

LONG_RSI_MIN = 51.0
SHORT_RSI_MAX = 49.0
WHIPSAW_WINDOWS = [
    ("2020-05-29", "2020-07-07"),
    ("2022-05-31", "2022-10-05"),
    ("2023-08-02", "2023-10-12"),
]


def trade_in_window(trade: dict, start: str, end: str) -> bool:
    entry = date.fromisoformat(trade["entryTime"])
    return date.fromisoformat(start) <= entry <= date.fromisoformat(end)


def window_breakdown(trades: list[dict]) -> list[dict]:
    rows = []
    for start, end in WHIPSAW_WINDOWS:
        subset = [t for t in trades if trade_in_window(t, start, end)]
        rows.append(
            {
                "start": start,
                "end": end,
                "trades": len(subset),
                "wins": sum(1 for t in subset if t["rMultiple"] > 0),
                "totalR": sum(t["rMultiple"] for t in subset),
            }
        )
    rows.append(
        {
            "start": "combined",
            "end": "combined",
            "trades": sum(r["trades"] for r in rows),
            "wins": sum(r["wins"] for r in rows),
            "totalR": sum(r["totalR"] for r in rows),
        }
    )
    return rows


def fix_workbook_labels(path: Path) -> None:
    wb = load_workbook(path)
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        ws["A1"] = "NXT v3.2 Latest Test - RSI 51/49"
        ws["A2"] = "BTC/SOL/SUI 6Y | Runner A | RSI LONG > 51, SHORT < 49 | 20% reversal-skip | no continuation | no risk-off."
    if "Trades" in wb.sheetnames:
        wb["Trades"]["A1"] = "Detailed Trades - NXT v3.2 RSI 51/49 Test"
    for sheet in ["BTC", "SOL", "SUI"]:
        if sheet in wb.sheetnames:
            wb[sheet]["A1"] = f"{sheet} - NXT v3.2 RSI 51/49 Test"
    wb.save(path)


def backtest_symbol(symbol: str, candles: list[dict]) -> tuple[list[dict], list[dict]]:
    trades, skipped, pos, n = [], [], None, 1
    skip_reversal = None
    for i in range(55, len(candles) - 1):
        c, prev, nxt = candles[i], candles[i - 1], candles[i + 1]
        next_date = base.date.fromisoformat(nxt["localDate"])
        if next_date < base.START_DATE or next_date >= base.END_DATE:
            continue

        skip_reversal = None
        if pos:
            side = pos["side"]
            if side == "LONG":
                pos["favorable20"] = pos["favorable20"] or c["high"] >= pos["entry"] * 1.20
            else:
                pos["favorable20"] = pos["favorable20"] or c["low"] <= pos["entry"] * 0.80

            ssl_flip = (side == "LONG" and prev["ssl"] == 1 and c["ssl"] == -1) or (
                side == "SHORT" and prev["ssl"] == -1 and c["ssl"] == 1
            )
            exit_price = reason = None
            if side == "LONG":
                if c["low"] <= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["high"] >= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["tp"] - pos["entry"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bearish flip"
                        if pos["favorable20"]:
                            skip_reversal = "SHORT"
            else:
                if c["high"] >= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["low"] <= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["entry"] - pos["tp"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bullish flip"
                        if pos["favorable20"]:
                            skip_reversal = "LONG"

            if exit_price is not None:
                rem = 0.5 if pos["triggered"] else 1.0
                rem_r = (exit_price - pos["entry"]) / pos["risk"] if side == "LONG" else (pos["entry"] - exit_price) / pos["risk"]
                gross = pos["realizedR"] + rem * rem_r
                cost = base.cost_r(pos["entry"], pos["risk"])
                net = gross - cost
                trades.append(
                    {
                        "symbol": symbol,
                        "tradeNo": n,
                        "side": side,
                        "signalTime": pos["signalDate"],
                        "entryTime": pos["entryDate"],
                        "entryPrice": pos["entry"],
                        "initialStop": pos["initialStop"],
                        "finalStop": pos["stop"],
                        "riskPerUnit": pos["risk"],
                        "tp1": pos["tp"],
                        "tp1Time": pos["tp1Time"],
                        "exitTime": c["localDate"],
                        "exitPrice": exit_price,
                        "exitReason": reason,
                        "grossRMultiple": gross,
                        "costR": cost,
                        "rMultiple": net,
                        "atr14": pos["atr14"],
                        "rsi14": pos["rsi14"],
                        "distanceToEma50Atr": pos["distance"],
                        "favorable20PctReached": pos["favorable20"],
                        "notes": "Binance native 1D candles; TradingView ATR RMA; RSI entry thresholds LONG > 51 and SHORT < 49; 20pct reversal-skip retained",
                    }
                )
                n += 1
                pos = None
            if pos:
                continue

        if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]) or prev["ssl"] is None:
            continue
        dist = abs(c["close"] - c["ema50"]) / c["atr14"]
        long_ok = (
            prev["ssl"] == -1
            and c["ssl"] == 1
            and base.recent_cross(candles, i, "LONG")
            and dist <= 2
            and c["rsi14"] > LONG_RSI_MIN
        )
        short_ok = (
            prev["ssl"] == 1
            and c["ssl"] == -1
            and base.recent_cross(candles, i, "SHORT")
            and dist <= 2
            and c["rsi14"] < SHORT_RSI_MAX
        )

        if skip_reversal == "LONG" and long_ok:
            skipped.append({"symbol": symbol, "date": c["localDate"], "side": "LONG", "reason": "Previous SHORT moved at least 20pct in favor before bullish SSL flip"})
            long_ok = False
        if skip_reversal == "SHORT" and short_ok:
            skipped.append({"symbol": symbol, "date": c["localDate"], "side": "SHORT", "reason": "Previous LONG moved at least 20pct in favor before bearish SSL flip"})
            short_ok = False

        if not (long_ok or short_ok):
            continue
        side = "LONG" if long_ok else "SHORT"
        risk = c["atr14"] * 1.5
        entry = nxt["open"]
        pos = {
            "side": side,
            "signalDate": c["localDate"],
            "entryDate": nxt["localDate"],
            "entry": entry,
            "initialStop": entry - risk if side == "LONG" else entry + risk,
            "stop": entry - risk if side == "LONG" else entry + risk,
            "risk": risk,
            "tp": entry + c["atr14"] * 2.5 if side == "LONG" else entry - c["atr14"] * 2.5,
            "triggered": False,
            "tp1Time": "",
            "realizedR": 0.0,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": dist,
            "favorable20": False,
        }
    return trades, skipped


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    baseline = json.loads(BASELINE_JSON.read_text(encoding="utf-8"))
    all_trades = []
    all_skipped = []
    datasets = {}
    for symbol in base.SYMBOLS:
        candles = tv_atr.enrich_tv_atr(native.fetch_native_1d(symbol))
        datasets[symbol] = {
            "dailyRows": len(candles),
            "firstDay": candles[0]["localDate"],
            "lastDay": candles[-1]["localDate"],
            "source": "Binance spot native 1D klines",
        }
        trades, skipped = backtest_symbol(symbol, candles)
        all_trades.extend(trades)
        all_skipped.extend(skipped)

    all_trades.sort(key=lambda x: x["exitTime"])
    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "systemVersion": "NXT v3.2 Latest Test + RSI 51/49",
        "period": {
            "start": base.START_DATE.isoformat(),
            "end": (base.END_DATE - base.timedelta(days=1)).isoformat(),
            "timezone": "Binance native daily candles",
        },
        "symbols": base.SYMBOLS,
        "stats": base.stats(all_trades),
        "trades": all_trades,
        "skippedSignals": all_skipped,
        "datasets": datasets,
        "baseline": {
            "stats": baseline["stats"],
            "whipsawWindows": window_breakdown(baseline["trades"]),
        },
        "whipsawWindows": window_breakdown(all_trades),
        "assumptions": [
            "Baseline is current latest NXT v3.2 with Binance native 1D candles, TradingView ATR RMA, Runner A, 20% reversal-skip, no continuation, and no risk-off.",
            "Only changed entry RSI thresholds: LONG requires RSI14 > 51 and SHORT requires RSI14 < 49.",
            "This is a test output only and is not promoted to latest.",
        ],
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    native.OUT_XLSX = OUT_XLSX
    native.build_workbook(result)
    fix_workbook_labels(OUT_XLSX)
    print(
        json.dumps(
            {
                "outJson": str(OUT_JSON),
                "outXlsx": str(OUT_XLSX),
                "baselineStats": baseline["stats"],
                "stats": result["stats"],
                "baselineWhipsawCombinedR": result["baseline"]["whipsawWindows"][-1]["totalR"],
                "whipsawCombinedR": result["whipsawWindows"][-1]["totalR"],
                "skippedSignals": len(all_skipped),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
