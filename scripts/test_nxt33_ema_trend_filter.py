from __future__ import annotations

import json
import sys
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

import backtest_nxt31_utc7_latest as base
import backtest_nxt32_native_1d_latest as native


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "outputs" / "nxt33_ema_trend_filter_test"
OUT_JSON = OUT_DIR / "nxt33_ema_trend_filter_test_results.json"
OUT_XLSX = OUT_DIR / "NXT33_EMA20_EMA50_Trend_Filter_Test.xlsx"
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"

VARIANTS = [
    {"name": "Baseline", "mode": "none", "gapAtr": None},
    {"name": "Strict align only", "mode": "strict_align", "gapAtr": 0.0},
    {"name": "Strict align + gap >= 0.25 ATR", "mode": "strict_align", "gapAtr": 0.25},
    {"name": "Strict align + gap >= 0.50 ATR", "mode": "strict_align", "gapAtr": 0.50},
    {"name": "Block countertrend gap >= 0.10 ATR", "mode": "block_countertrend", "gapAtr": 0.10},
    {"name": "Block countertrend gap >= 0.25 ATR", "mode": "block_countertrend", "gapAtr": 0.25},
    {"name": "Block countertrend gap >= 0.50 ATR", "mode": "block_countertrend", "gapAtr": 0.50},
    {"name": "Block countertrend gap >= 0.75 ATR", "mode": "block_countertrend", "gapAtr": 0.75},
    {"name": "Block countertrend gap >= 1.00 ATR", "mode": "block_countertrend", "gapAtr": 1.00},
]


def copy_layout(src, dst):
    dst.sheet_format.defaultColWidth = src.sheet_format.defaultColWidth
    dst.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    for c in range(1, max(src.max_column, dst.max_column) + 1):
        letter = get_column_letter(c)
        if src.column_dimensions.get(letter) and src.column_dimensions[letter].width:
            dst.column_dimensions[letter].width = src.column_dimensions[letter].width
    for r in range(1, max(src.max_row, dst.max_row) + 1):
        sr = r if r <= src.max_row else 5
        if src.row_dimensions.get(sr) and src.row_dimensions[sr].height:
            dst.row_dimensions[r].height = src.row_dimensions[sr].height
        for c in range(1, max(src.max_column, dst.max_column) + 1):
            a = src.cell(sr, min(c, src.max_column))
            b = dst.cell(r, c)
            if a.has_style:
                b._style = copy(a._style)
            b.number_format = a.number_format
            b.alignment = copy(a.alignment)
            b.font = copy(a.font)
            b.fill = copy(a.fill)
            b.border = copy(a.border)


def clear(ws, rows=900, cols=36):
    for row in ws.iter_rows(min_row=1, max_row=max(rows, ws.max_row), min_col=1, max_col=max(cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def rowset(ws, r, values):
    for c, v in enumerate(values, 1):
        ws.cell(r, c).value = v


def pass_ema_filter(c: dict, side: str, variant: dict) -> bool:
    if variant["mode"] == "none":
        return True
    if c["ema20"] is None or c["ema50"] is None or c["atr14"] in (None, 0):
        return False
    gap_atr = abs(c["ema20"] - c["ema50"]) / c["atr14"]
    is_countertrend = (side == "LONG" and c["ema20"] < c["ema50"]) or (side == "SHORT" and c["ema20"] > c["ema50"])
    if variant["mode"] == "strict_align":
        if is_countertrend or c["ema20"] == c["ema50"]:
            return False
        return gap_atr >= variant["gapAtr"]
    if variant["mode"] == "block_countertrend":
        return not (is_countertrend and gap_atr >= variant["gapAtr"])
    raise ValueError(f"Unknown EMA filter mode: {variant['mode']}")


def backtest_symbol(symbol: str, candles: list[dict], variant: dict) -> list[dict]:
    trades, pos, n = [], None, 1
    last_profitable_runner_exit = None
    for i in range(55, len(candles) - 1):
        c, prev, nxt = candles[i], candles[i - 1], candles[i + 1]
        next_date = base.date.fromisoformat(nxt["localDate"])
        if next_date < native.START_DATE or next_date >= native.END_DATE:
            continue
        if pos:
            side = pos["side"]
            ssl_flip = (side == "LONG" and prev["ssl"] == 1 and c["ssl"] == -1) or (side == "SHORT" and prev["ssl"] == -1 and c["ssl"] == 1)
            exit_price = reason = None
            if side == "LONG":
                if c["low"] <= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["high"] >= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["tp"] - pos["entry"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bearish flip"
            else:
                if c["high"] >= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["low"] <= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["entry"] - pos["tp"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bullish flip"
            if exit_price is not None:
                rem = 0.5 if pos["triggered"] else 1.0
                rem_r = (exit_price - pos["entry"]) / pos["risk"] if side == "LONG" else (pos["entry"] - exit_price) / pos["risk"]
                gross = pos["realizedR"] + rem * rem_r
                net = gross - base.cost_r(pos["entry"], pos["risk"])
                trades.append({
                    "variant": variant["name"],
                    "symbol": symbol,
                    "tradeNo": n,
                    "side": side,
                    "signalTime": pos["signalDate"],
                    "entryTime": pos["entryDate"],
                    "entryPrice": pos["entry"],
                    "initialStop": pos["initialStop"],
                    "finalStop": pos["stop"],
                    "riskPerUnit": pos["risk"],
                    "tp1": pos["tp"],
                    "tp1Time": pos["tp1Time"],
                    "exitTime": c["localDate"],
                    "exitPrice": exit_price,
                    "exitReason": reason,
                    "grossRMultiple": gross,
                    "costR": base.cost_r(pos["entry"], pos["risk"]),
                    "rMultiple": net,
                    "atr14": pos["atr14"],
                    "rsi14": pos["rsi14"],
                    "distanceToEma50Atr": pos["distance"],
                    "ema20": pos["ema20"],
                    "ema50": pos["ema50"],
                    "emaGapAtr": pos["emaGapAtr"],
                    "notes": f"NXT v3.3 Runner A; {variant['name']}",
                })
                if net > 0 and reason.startswith("Runner exit"):
                    last_profitable_runner_exit = {"index": i, "side": side}
                n += 1
                pos = None
            if pos:
                continue

        if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]) or prev["ssl"] is None:
            continue
        dist = abs(c["close"] - c["ema50"]) / c["atr14"]
        long_ok = prev["ssl"] == -1 and c["ssl"] == 1 and base.recent_cross(candles, i, "LONG") and dist <= 2 and c["rsi14"] > 50
        short_ok = prev["ssl"] == 1 and c["ssl"] == -1 and base.recent_cross(candles, i, "SHORT") and dist <= 2 and c["rsi14"] < 50
        if long_ok and not pass_ema_filter(c, "LONG", variant):
            long_ok = False
        if short_ok and not pass_ema_filter(c, "SHORT", variant):
            short_ok = False
        if last_profitable_runner_exit and i - last_profitable_runner_exit["index"] <= 1:
            if long_ok and last_profitable_runner_exit["side"] == "SHORT":
                long_ok = False
            if short_ok and last_profitable_runner_exit["side"] == "LONG":
                short_ok = False
        if not (long_ok or short_ok):
            continue
        side = "LONG" if long_ok else "SHORT"
        risk = c["atr14"] * 1.5
        entry = nxt["open"]
        pos = {
            "side": side,
            "signalDate": c["localDate"],
            "entryDate": nxt["localDate"],
            "entry": entry,
            "initialStop": entry - risk if side == "LONG" else entry + risk,
            "stop": entry - risk if side == "LONG" else entry + risk,
            "risk": risk,
            "tp": entry + c["atr14"] * 2.5 if side == "LONG" else entry - c["atr14"] * 2.5,
            "triggered": False,
            "tp1Time": "",
            "realizedR": 0.0,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": dist,
            "ema20": c["ema20"],
            "ema50": c["ema50"],
            "emaGapAtr": abs(c["ema20"] - c["ema50"]) / c["atr14"],
        }
    return trades


def side_stats(rows: list[dict]) -> dict:
    out = {}
    for side in ["LONG", "SHORT"]:
        subset = [t for t in rows if t["side"] == side]
        out[side] = base.stats(subset)
    return out


def summarize_variant(variant: dict, rows: list[dict], baseline_stats: dict | None = None) -> dict:
    st = base.stats(rows)
    gross_profit = sum(t["rMultiple"] for t in rows if t["rMultiple"] > 0)
    gross_loss = -sum(t["rMultiple"] for t in rows if t["rMultiple"] < 0)
    st["profitFactor"] = gross_profit / gross_loss if gross_loss else None
    st["ending20k"] = 20000 + st["totalR"] * 1000
    st["gapAtr"] = variant["gapAtr"]
    st["mode"] = variant["mode"]
    if baseline_stats:
        st["deltaTrades"] = st["trades"] - baseline_stats["trades"]
        st["deltaTotalR"] = st["totalR"] - baseline_stats["totalR"]
        st["deltaMaxDrawdownR"] = st["maxDrawdownR"] - baseline_stats["maxDrawdownR"]
    return st


def build_workbook(result: dict) -> None:
    tpl = load_workbook(TEMPLATE)
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        clear(ws)

    summary = wb["Summary"]
    summary["A1"] = "NXT v3.3 EMA20/EMA50 Trend Filter Test"
    summary["A2"] = "Baseline versus EMA alignment + EMA gap ATR filters. Latest files are not overwritten."
    headers = [
        "Variant", "Trades", "Win Rate", "Total R", "Delta R", "Avg R", "Max DD R",
        "Delta DD R", "Best R", "Worst R", "Profit Factor", "20K Ending", "Rule",
    ]
    rowset(summary, 4, headers)
    for i, row in enumerate(result["summary"], 5):
        if row["mode"] == "none":
            rule = "Baseline"
        elif row["mode"] == "strict_align":
            rule = f"Require EMA20/EMA50 alignment; gap >= {row['gapAtr']:.2f} ATR"
        else:
            rule = f"Block countertrend only when EMA gap >= {row['gapAtr']:.2f} ATR"
        rowset(summary, i, [
            row["variant"], row["trades"], row["winRate"], row["totalR"], row.get("deltaTotalR", 0),
            row["avgR"], row["maxDrawdownR"], row.get("deltaMaxDrawdownR", 0), row["bestR"],
            row["worstR"], row["profitFactor"], row["ending20k"], rule,
        ])

    trades = wb["Trades"]
    trades["A1"] = "Detailed Trades - Best EMA Filter Variant"
    best_name = result["bestVariant"]
    trades["A2"] = f"Variant: {best_name}"
    trade_headers = [
        "Symbol", "No", "Side", "Signal Date", "Entry Date", "Entry Price", "Initial Stop",
        "Final Stop", "Risk / Unit", "TP1", "TP1 Date", "Exit Date", "Exit Price",
        "Exit Reason", "R", "ATR14", "RSI14", "Distance EMA50 ATR", "EMA20", "EMA50",
        "EMA Gap ATR", "Notes",
    ]
    rowset(trades, 4, trade_headers)
    for i, t in enumerate(result["variants"][best_name]["trades"], 5):
        rowset(trades, i, [
            t["symbol"].replace("USDT", ""), t["tradeNo"], t["side"], t["signalTime"], t["entryTime"],
            t["entryPrice"], t["initialStop"], t["finalStop"], t["riskPerUnit"], t["tp1"],
            t["tp1Time"], t["exitTime"], t["exitPrice"], t["exitReason"], t["rMultiple"],
            t["atr14"], t["rsi14"], t["distanceToEma50Atr"], t["ema20"], t["ema50"],
            t["emaGapAtr"], t["notes"],
        ])

    ass = wb["Assumptions"]
    ass["A1"] = "Assumptions"
    rowset(ass, 4, ["#", "Assumption"])
    for i, line in enumerate(result["assumptions"], 5):
        rowset(ass, i, [i - 4, line])

    quality = wb["Data Quality"]
    quality["A1"] = "Data Quality"
    rowset(quality, 4, ["Symbol", "Daily Rows", "First Day", "Last Day", "Source"])
    for i, (sym, q) in enumerate(result["datasets"].items(), 5):
        rowset(quality, i, [sym.replace("USDT", ""), q["dailyRows"], q["firstDay"], q["lastDay"], q["source"]])

    by_side = wb["BTC"]
    by_side.title = "By Side"
    by_side["A1"] = "By Side Summary"
    rowset(by_side, 4, ["Variant", "Side", "Trades", "Win Rate", "Total R", "Avg R", "Max DD R", "Best R", "Worst R"])
    r = 5
    for name, payload in result["variants"].items():
        for side, st in payload["sideStats"].items():
            rowset(by_side, r, [name, side, st["trades"], st["winRate"], st["totalR"], st["avgR"], st["maxDrawdownR"], st["bestR"], st["worstR"]])
            r += 1

    removed = wb["SOL"]
    removed.title = "Baseline Removed"
    removed["A1"] = "Baseline Trades Removed By Best Variant"
    rowset(removed, 4, trade_headers)
    for i, t in enumerate(result["removedByBest"], 5):
        rowset(removed, i, [
            t["symbol"].replace("USDT", ""), t["tradeNo"], t["side"], t["signalTime"], t["entryTime"],
            t["entryPrice"], t["initialStop"], t["finalStop"], t["riskPerUnit"], t["tp1"],
            t["tp1Time"], t["exitTime"], t["exitPrice"], t["exitReason"], t["rMultiple"],
            t["atr14"], t["rsi14"], t["distanceToEma50Atr"], t["ema20"], t["ema50"],
            t["emaGapAtr"], t["notes"],
        ])

    for ws in [summary, trades, by_side, removed, ass, quality]:
        ws.freeze_panes = "A5"
        for cell in ws[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = min(max(ws.column_dimensions[get_column_letter(col)].width or 12, 12), 24)
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["M"].width = 42
    trades.column_dimensions["V"].width = 40
    removed.column_dimensions["V"].width = 40
    wb.save(OUT_XLSX)


def trade_key(t: dict) -> tuple:
    return (t["symbol"], t["side"], t["signalTime"], t["entryTime"])


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    datasets = {}
    candles_by_symbol = {}
    for symbol in native.SYMBOLS:
        candles = base.enrich(native.fetch_native_1d(symbol))
        candles_by_symbol[symbol] = candles
        datasets[symbol] = {
            "dailyRows": len(candles),
            "firstDay": candles[0]["localDate"],
            "lastDay": candles[-1]["localDate"],
            "source": "Binance spot native 1D klines",
        }

    variants = {}
    baseline_stats = None
    summary = []
    for variant in VARIANTS:
        rows = []
        for symbol, candles in candles_by_symbol.items():
            rows.extend(backtest_symbol(symbol, candles, variant))
        rows.sort(key=lambda x: x["exitTime"])
        st = summarize_variant(variant, rows, baseline_stats)
        st["variant"] = variant["name"]
        if baseline_stats is None:
            baseline_stats = st
        summary.append(st)
        variants[variant["name"]] = {"stats": st, "sideStats": side_stats(rows), "trades": rows}

    candidates = [x for x in summary if x["variant"] != "Baseline"]
    best = sorted(candidates, key=lambda x: (x["totalR"], x["maxDrawdownR"]), reverse=True)[0]
    baseline_keys = {trade_key(t): t for t in variants["Baseline"]["trades"]}
    best_keys = {trade_key(t) for t in variants[best["variant"]]["trades"]}
    removed = [t for k, t in baseline_keys.items() if k not in best_keys]
    removed.sort(key=lambda x: x["signalTime"])

    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "systemVersion": "NXT v3.3 EMA20/EMA50 trend filter test",
        "period": {"start": native.START_DATE.isoformat(), "end": (native.END_DATE - base.timedelta(days=1)).isoformat()},
        "summary": summary,
        "bestVariant": best["variant"],
        "variants": variants,
        "removedByBest": removed,
        "datasets": datasets,
        "assumptions": [
            "Baseline is NXT v3.3 Binance native 1D, Runner A, anti-immediate-reversal, no continuation, no risk-off.",
            "EMA trend filter: LONG requires EMA20 > EMA50; SHORT requires EMA20 < EMA50.",
            "EMA gap is normalized as ABS(EMA20 - EMA50) / ATR14.",
            "Strict tests require every LONG/SHORT to align with EMA20/EMA50 and optionally require a minimum EMA gap.",
            "Countertrend-block tests only block LONG below EMA trend or SHORT above EMA trend when EMA gap is already large.",
            "Entry remains next daily open after signal close; exit/risk/cost rules are unchanged.",
            "Latest production files were not overwritten.",
        ],
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    build_workbook(result)
    print(json.dumps({
        "outJson": str(OUT_JSON),
        "outXlsx": str(OUT_XLSX),
        "bestVariant": result["bestVariant"],
        "summary": summary,
    }, indent=2))


if __name__ == "__main__":
    main()
