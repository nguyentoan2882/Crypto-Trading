from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

import backtest_nxt31_utc7_latest as base
import backtest_nxt32_native_1d_latest as native


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "outputs" / "nxt33_market_20pct_trend_skip"
OUT_JSON = OUT_DIR / "nxt33_market_20pct_trend_skip_results.json"
OUT_XLSX = OUT_DIR / "NXT33_Market_20Pct_Trend_Skip_Test.xlsx"

VARIANTS = [
    {"name": "Baseline", "lookback": None, "requirePriceSideEma50": False},
    {"name": "Market 20pct skip LB20", "lookback": 20, "requirePriceSideEma50": False},
    {"name": "Market 20pct skip LB30", "lookback": 30, "requirePriceSideEma50": False},
    {"name": "Market 20pct skip LB45", "lookback": 45, "requirePriceSideEma50": False},
    {"name": "Market 20pct skip LB60", "lookback": 60, "requirePriceSideEma50": False},
    {"name": "20pct + close above/below EMA50 LB20", "lookback": 20, "requirePriceSideEma50": True},
    {"name": "20pct + close above/below EMA50 LB30", "lookback": 30, "requirePriceSideEma50": True},
    {"name": "20pct + close above/below EMA50 LB45", "lookback": 45, "requirePriceSideEma50": True},
]


def skip_by_market_20pct(candles: list[dict], i: int, side: str, variant: dict) -> tuple[bool, str]:
    lookback = variant["lookback"]
    if lookback is None:
        return False, ""
    c = candles[i]
    if c["ema20"] is None or c["ema50"] is None:
        return False, ""
    window = candles[max(0, i - lookback): i + 1]
    lowest_low = min(x["low"] for x in window)
    highest_high = max(x["high"] for x in window)
    if side == "SHORT":
        rally = highest_high / lowest_low - 1
        price_ok = (not variant["requirePriceSideEma50"]) or c["close"] > c["ema50"]
        if rally >= 0.20 and c["ema20"] > c["ema50"] and price_ok:
            return True, f"skip SHORT: prior {lookback}D rally {rally:.1%}, EMA20 > EMA50"
    else:
        drop = 1 - lowest_low / highest_high
        price_ok = (not variant["requirePriceSideEma50"]) or c["close"] < c["ema50"]
        if drop >= 0.20 and c["ema20"] < c["ema50"] and price_ok:
            return True, f"skip LONG: prior {lookback}D drop {drop:.1%}, EMA20 < EMA50"
    return False, ""


def backtest_symbol(symbol: str, candles: list[dict], variant: dict) -> tuple[list[dict], list[dict]]:
    trades, skipped, pos, n = [], [], None, 1
    last_profitable_runner_exit = None
    for i in range(55, len(candles) - 1):
        c, prev, nxt = candles[i], candles[i - 1], candles[i + 1]
        next_date = base.date.fromisoformat(nxt["localDate"])
        if next_date < native.START_DATE or next_date >= native.END_DATE:
            continue
        if pos:
            side = pos["side"]
            ssl_flip = (side == "LONG" and prev["ssl"] == 1 and c["ssl"] == -1) or (side == "SHORT" and prev["ssl"] == -1 and c["ssl"] == 1)
            exit_price = reason = None
            if side == "LONG":
                if c["low"] <= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["high"] >= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["tp"] - pos["entry"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bearish flip"
            else:
                if c["high"] >= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["low"] <= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["entry"] - pos["tp"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bullish flip"
            if exit_price is not None:
                rem = 0.5 if pos["triggered"] else 1.0
                rem_r = (exit_price - pos["entry"]) / pos["risk"] if side == "LONG" else (pos["entry"] - exit_price) / pos["risk"]
                gross = pos["realizedR"] + rem * rem_r
                net = gross - base.cost_r(pos["entry"], pos["risk"])
                trades.append({
                    "variant": variant["name"],
                    "symbol": symbol,
                    "tradeNo": n,
                    "side": side,
                    "signalTime": pos["signalDate"],
                    "entryTime": pos["entryDate"],
                    "entryPrice": pos["entry"],
                    "initialStop": pos["initialStop"],
                    "finalStop": pos["stop"],
                    "riskPerUnit": pos["risk"],
                    "tp1": pos["tp"],
                    "tp1Time": pos["tp1Time"],
                    "exitTime": c["localDate"],
                    "exitPrice": exit_price,
                    "exitReason": reason,
                    "grossRMultiple": gross,
                    "costR": base.cost_r(pos["entry"], pos["risk"]),
                    "rMultiple": net,
                    "atr14": pos["atr14"],
                    "rsi14": pos["rsi14"],
                    "distanceToEma50Atr": pos["distance"],
                    "notes": f"NXT v3.3 Runner A; {variant['name']}",
                })
                if net > 0 and reason.startswith("Runner exit"):
                    last_profitable_runner_exit = {"index": i, "side": side}
                n += 1
                pos = None
            if pos:
                continue

        if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]) or prev["ssl"] is None:
            continue
        dist = abs(c["close"] - c["ema50"]) / c["atr14"]
        long_ok = prev["ssl"] == -1 and c["ssl"] == 1 and base.recent_cross(candles, i, "LONG") and dist <= 2 and c["rsi14"] > 50
        short_ok = prev["ssl"] == 1 and c["ssl"] == -1 and base.recent_cross(candles, i, "SHORT") and dist <= 2 and c["rsi14"] < 50

        if last_profitable_runner_exit and i - last_profitable_runner_exit["index"] <= 1:
            if long_ok and last_profitable_runner_exit["side"] == "SHORT":
                long_ok = False
            if short_ok and last_profitable_runner_exit["side"] == "LONG":
                short_ok = False

        if long_ok:
            skip, reason = skip_by_market_20pct(candles, i, "LONG", variant)
            if skip:
                skipped.append({"symbol": symbol, "date": c["localDate"], "side": "LONG", "reason": reason})
                long_ok = False
        if short_ok:
            skip, reason = skip_by_market_20pct(candles, i, "SHORT", variant)
            if skip:
                skipped.append({"symbol": symbol, "date": c["localDate"], "side": "SHORT", "reason": reason})
                short_ok = False

        if not (long_ok or short_ok):
            continue
        side = "LONG" if long_ok else "SHORT"
        risk = c["atr14"] * 1.5
        entry = nxt["open"]
        pos = {
            "side": side,
            "signalDate": c["localDate"],
            "entryDate": nxt["localDate"],
            "entry": entry,
            "initialStop": entry - risk if side == "LONG" else entry + risk,
            "stop": entry - risk if side == "LONG" else entry + risk,
            "risk": risk,
            "tp": entry + c["atr14"] * 2.5 if side == "LONG" else entry - c["atr14"] * 2.5,
            "triggered": False,
            "tp1Time": "",
            "realizedR": 0.0,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": dist,
        }
    return trades, skipped


def profit_factor(rows: list[dict]) -> float | None:
    gp = sum(t["rMultiple"] for t in rows if t["rMultiple"] > 0)
    gl = -sum(t["rMultiple"] for t in rows if t["rMultiple"] < 0)
    return gp / gl if gl else None


def make_summary(variant: dict, rows: list[dict], skipped: list[dict], baseline: dict | None) -> dict:
    st = base.stats(rows)
    st["variant"] = variant["name"]
    st["lookback"] = variant["lookback"]
    st["profitFactor"] = profit_factor(rows)
    st["ending20k"] = 20000 + st["totalR"] * 1000
    st["skipped"] = len(skipped)
    if baseline:
        st["deltaTotalR"] = st["totalR"] - baseline["totalR"]
        st["deltaMaxDrawdownR"] = st["maxDrawdownR"] - baseline["maxDrawdownR"]
        st["deltaTrades"] = st["trades"] - baseline["trades"]
    else:
        st["deltaTotalR"] = 0
        st["deltaMaxDrawdownR"] = 0
        st["deltaTrades"] = 0
    return st


def build_workbook(result: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "NXT v3.3 Market 20% Trend-Skip Test"
    ws["A2"] = "Rule test: skip countertrend SSL flip after a recent 20% market move while EMA trend still points the other way."
    headers = ["Variant", "Lookback", "Trades", "Skipped", "Win Rate", "Total R", "Delta R", "Avg R", "Max DD R", "Delta DD R", "PF", "20K Ending"]
    ws.append([])
    ws.append(headers)
    for row in result["summary"]:
        ws.append([row["variant"], row["lookback"], row["trades"], row["skipped"], row["winRate"], row["totalR"], row["deltaTotalR"], row["avgR"], row["maxDrawdownR"], row["deltaMaxDrawdownR"], row["profitFactor"], row["ending20k"]])

    skipped_ws = wb.create_sheet("Skipped Signals")
    skipped_ws.append(["Variant", "Symbol", "Date", "Side", "Reason"])
    for variant, payload in result["variants"].items():
        for s in payload["skippedSignals"]:
            skipped_ws.append([variant, s["symbol"].replace("USDT", ""), s["date"], s["side"], s["reason"]])

    trade_ws = wb.create_sheet("Best Trades")
    trade_ws.append(["Symbol", "No", "Side", "Signal", "Entry", "Exit", "Exit Reason", "R", "Notes"])
    best = result["bestVariant"]
    for t in result["variants"][best]["trades"]:
        trade_ws.append([t["symbol"].replace("USDT", ""), t["tradeNo"], t["side"], t["signalTime"], t["entryTime"], t["exitTime"], t["exitReason"], t["rMultiple"], t["notes"]])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A4" if sheet.title == "Summary" else "A2"
        header_row = 4 if sheet.title == "Summary" else 1
        for cell in sheet[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 16
    skipped_ws.column_dimensions["E"].width = 60
    trade_ws.column_dimensions["I"].width = 42
    ws.column_dimensions["A"].width = 28
    wb.save(OUT_XLSX)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    candles_by_symbol = {symbol: base.enrich(native.fetch_native_1d(symbol)) for symbol in native.SYMBOLS}
    variants, summary = {}, []
    baseline = None
    for variant in VARIANTS:
        rows, skipped = [], []
        for symbol, candles in candles_by_symbol.items():
            t, s = backtest_symbol(symbol, candles, variant)
            rows.extend(t)
            skipped.extend(s)
        rows.sort(key=lambda x: x["exitTime"])
        st = make_summary(variant, rows, skipped, baseline)
        if baseline is None:
            baseline = st
        summary.append(st)
        variants[variant["name"]] = {"stats": st, "trades": rows, "skippedSignals": skipped}

    candidates = [s for s in summary if s["variant"] != "Baseline"]
    best = sorted(candidates, key=lambda x: (x["totalR"], x["maxDrawdownR"]), reverse=True)[0]["variant"]
    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "systemVersion": "NXT v3.3 market 20pct trend-skip test",
        "summary": summary,
        "bestVariant": best,
        "variants": variants,
        "assumptions": [
            "Baseline is NXT v3.3 native 1D, Runner A, anti-immediate-reversal, no continuation, no risk-off.",
            "For SHORT: if the recent lookback window rallied at least 20% from lowest low to highest high and EMA20 > EMA50, skip bearish SSL flip.",
            "For LONG: if the recent lookback window dropped at least 20% from highest high to lowest low and EMA20 < EMA50, skip bullish SSL flip.",
            "This test is designed to catch cases like BTC SHORT 2021-01-21 after the prior bull impulse.",
        ],
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    build_workbook(result)
    jan_case = {
        name: [s for s in payload["skippedSignals"] if s["symbol"] == "BTCUSDT" and s["date"] == "2021-01-21"]
        for name, payload in variants.items()
    }
    print(json.dumps({"outJson": str(OUT_JSON), "outXlsx": str(OUT_XLSX), "bestVariant": best, "summary": summary, "btc20210121Skipped": jan_case}, indent=2))


if __name__ == "__main__":
    main()
