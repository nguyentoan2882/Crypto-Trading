from __future__ import annotations

import json
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\Workspace\Codex\Crypto trading")
DATA = ROOT / "data_cache" / "binance_spot_1d"
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
OUT_DIR = ROOT / "outputs" / "nxt31_8020_continuation_6y"
OUT_JSON = OUT_DIR / "nxt31_8020_continuation_6y_results.json"
OUT_XLSX = OUT_DIR / "NXT31_8020_Continuation_6Y_BTC_SOL_SUI_20K.xlsx"

SYMBOLS = ["BTCUSDT", "SOLUSDT", "SUIUSDT"]
START = int(datetime(2020, 5, 17, tzinfo=timezone.utc).timestamp() * 1000)
END = int(datetime(2026, 5, 17, tzinfo=timezone.utc).timestamp() * 1000)
FEE = 0.0006
SLIPPAGE = 0.0005
ROUND_TRIP = 2 * (FEE + SLIPPAGE)


def iso(ms: int) -> str:
    return datetime.fromtimestamp(ms / 1000, timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def sma(values, period):
    out = [None] * len(values)
    total = 0.0
    for i, value in enumerate(values):
        total += value
        if i >= period:
            total -= values[i - period]
        if i >= period - 1:
            out[i] = total / period
    return out


def ema(values, period):
    out = [None] * len(values)
    k = 2 / (period + 1)
    total = 0.0
    for i, value in enumerate(values):
        if i < period:
            total += value
        if i == period - 1:
            out[i] = total / period
        elif i >= period:
            out[i] = value * k + out[i - 1] * (1 - k)
    return out


def rma(values, period):
    out = [None] * len(values)
    total = 0.0
    for i, value in enumerate(values):
        if i < period:
            total += value
        if i == period - 1:
            out[i] = total / period
        elif i >= period:
            out[i] = (out[i - 1] * (period - 1) + value) / period
    return out


def rsi(values, period=14):
    out = [None] * len(values)
    avg_gain = avg_loss = 0.0
    for i in range(1, len(values)):
        change = values[i] - values[i - 1]
        gain = max(change, 0)
        loss = max(-change, 0)
        if i <= period:
            avg_gain += gain
            avg_loss += loss
            if i == period:
                avg_gain /= period
                avg_loss /= period
                out[i] = 100 if avg_loss == 0 else 100 - 100 / (1 + avg_gain / avg_loss)
        else:
            avg_gain = (avg_gain * (period - 1) + gain) / period
            avg_loss = (avg_loss * (period - 1) + loss) / period
            out[i] = 100 if avg_loss == 0 else 100 - 100 / (1 + avg_gain / avg_loss)
    return out


def enrich(candles):
    closes = [c["close"] for c in candles]
    highs = [c["high"] for c in candles]
    lows = [c["low"] for c in candles]
    ema20 = ema(closes, 20)
    ema50 = ema(closes, 50)
    atr_tr = []
    for i, c in enumerate(candles):
        if i == 0:
            atr_tr.append(c["high"] - c["low"])
        else:
            pc = candles[i - 1]["close"]
            atr_tr.append(max(c["high"] - c["low"], abs(c["high"] - pc), abs(c["low"] - pc)))
    atr14 = rma(atr_tr, 14)
    rsi14 = rsi(closes, 14)
    high_sma = sma(highs, 10)
    low_sma = sma(lows, 10)
    ssl = []
    state = 0
    for i, c in enumerate(candles):
        if high_sma[i] is None or low_sma[i] is None:
            ssl.append(None)
            continue
        if c["close"] > high_sma[i]:
            state = 1
        elif c["close"] < low_sma[i]:
            state = -1
        ssl.append(state)
    for i, c in enumerate(candles):
        slope5 = ema50[i] - ema50[i - 5] if i >= 5 and ema50[i] is not None and ema50[i - 5] is not None else None
        c.update({"ema20": ema20[i], "ema50": ema50[i], "ema50Slope5": slope5, "atr14": atr14[i], "rsi14": rsi14[i], "ssl": ssl[i]})
    return candles


def crossed_up(cs, i, key="ema20"):
    return cs[i - 1]["close"] <= cs[i - 1][key] and cs[i]["close"] > cs[i][key]


def crossed_down(cs, i, key="ema20"):
    return cs[i - 1]["close"] >= cs[i - 1][key] and cs[i]["close"] < cs[i][key]


def recent_cross(cs, i, side, lookback=3):
    start = max(1, i - lookback + 1)
    if side == "LONG":
        return any(crossed_up(cs, j) for j in range(start, i + 1))
    return any(crossed_down(cs, j) for j in range(start, i + 1))


def recent_touch_reclaim(cs, i, side, lookback=5):
    start = max(1, i - lookback + 1)
    if side == "LONG":
        touched = any(cs[j]["low"] <= cs[j]["ema20"] for j in range(start, i + 1) if cs[j]["ema20"] is not None)
        return touched and cs[i]["close"] > cs[i]["ema20"] and cs[i - 1]["close"] <= cs[i]["close"]
    touched = any(cs[j]["high"] >= cs[j]["ema20"] for j in range(start, i + 1) if cs[j]["ema20"] is not None)
    return touched and cs[i]["close"] < cs[i]["ema20"] and cs[i - 1]["close"] >= cs[i]["close"]


def continuation_ok(cs, i, side, variant):
    c = cs[i]
    long_rsi = variant.get("contLongRsi", 55)
    short_rsi = variant.get("contShortRsi", 45)
    slope_ok_long = not variant.get("contSlope") or (c["ema50Slope5"] is not None and c["ema50Slope5"] > 0)
    slope_ok_short = not variant.get("contSlope") or (c["ema50Slope5"] is not None and c["ema50Slope5"] < 0)
    if side == "LONG":
        return (
            not variant.get("contShortOnly", False)
            and
            c["ssl"] == 1
            and c["close"] > c["ema20"] > c["ema50"]
            and c["rsi14"] > long_rsi
            and slope_ok_long
            and recent_touch_reclaim(cs, i, side, 5)
        )
    return (
        not variant.get("contLongOnly", False)
        and
        c["ssl"] == -1
        and c["close"] < c["ema20"] < c["ema50"]
        and c["rsi14"] < short_rsi
        and slope_ok_short
        and recent_touch_reclaim(cs, i, side, 5)
    )


VARIANTS = [
    {"key": "base_full_25", "name": "Base full close 2.5 ATR", "lock": 1.0, "runner": 0.0, "trail": "none", "continuation": False},
    {"key": "runner_a_50_ssl", "name": "Runner A 50/50 SSL", "lock": 0.5, "runner": 0.5, "trail": "ssl", "continuation": False},
    {"key": "tp80_runner20_ssl", "name": "TP1 80% + 20% SSL trailing", "lock": 0.8, "runner": 0.2, "trail": "ssl", "continuation": False},
    {"key": "tp80_runner20_chandelier", "name": "TP1 80% + 20% Chandelier 3ATR", "lock": 0.8, "runner": 0.2, "trail": "chandelier", "continuation": False},
    {"key": "tp80_runner20_ssl_cont", "name": "80/20 SSL + continuation", "lock": 0.8, "runner": 0.2, "trail": "ssl", "continuation": True},
    {"key": "runner_a_ssl_cont", "name": "Runner A + continuation", "lock": 0.5, "runner": 0.5, "trail": "ssl", "continuation": True},
    {"key": "runner_a_cont_hq", "name": "Runner A + HQ continuation", "lock": 0.5, "runner": 0.5, "trail": "ssl", "continuation": True, "contLongRsi": 60, "contShortRsi": 40, "contMaxDist": 2.3, "contSlope": True},
    {"key": "tp80_cont_hq", "name": "80/20 SSL + HQ continuation", "lock": 0.8, "runner": 0.2, "trail": "ssl", "continuation": True, "contLongRsi": 60, "contShortRsi": 40, "contMaxDist": 2.3, "contSlope": True},
    {"key": "runner_a_cont_long_only", "name": "Runner A + LONG continuation only", "lock": 0.5, "runner": 0.5, "trail": "ssl", "continuation": True, "contLongRsi": 60, "contShortRsi": 0, "contMaxDist": 2.5, "contSlope": True, "contLongOnly": True},
    {"key": "tp80_cont_long_only", "name": "80/20 SSL + LONG continuation only", "lock": 0.8, "runner": 0.2, "trail": "ssl", "continuation": True, "contLongRsi": 60, "contShortRsi": 0, "contMaxDist": 2.5, "contSlope": True, "contLongOnly": True},
]


def cost_r(entry, risk):
    return entry * ROUND_TRIP / risk


def backtest_symbol(symbol, candles, variant):
    trades = []
    position = None
    trade_no = 1
    for i in range(55, len(candles) - 1):
        prev = candles[i - 1]
        c = candles[i]
        nxt = candles[i + 1]
        if nxt["time"] < START or nxt["time"] >= END:
            continue

        if position:
            side = position["side"]
            ssl_flip = (side == "LONG" and prev["ssl"] == 1 and c["ssl"] == -1) or (side == "SHORT" and prev["ssl"] == -1 and c["ssl"] == 1)
            exit_price = None
            reason = None

            if side == "LONG":
                if c["low"] <= position["stop"]:
                    exit_price = position["stop"]
                    reason = "Breakeven/trailing stop" if position["triggered"] else "Stop loss"
                else:
                    if not position["triggered"] and c["high"] >= position["tp"]:
                        position["triggered"] = True
                        position["triggerTime"] = c["time"]
                        position["realizedR"] += position["lock"] * ((position["tp"] - position["entry"]) / position["risk"])
                        position["stop"] = position["entry"]
                    if position["triggered"] and position["trail"] == "chandelier":
                        position["hh"] = max(position["hh"], c["high"])
                        position["stop"] = max(position["stop"], position["hh"] - 3 * c["atr14"])
                    if variant["runner"] == 0 and c["high"] >= position["tp"]:
                        exit_price, reason = position["tp"], "TP 2.5 ATR full close"
                    elif position["triggered"] and ssl_flip:
                        exit_price, reason = c["close"], "Runner exit: SSL bearish flip"
            else:
                if c["high"] >= position["stop"]:
                    exit_price = position["stop"]
                    reason = "Breakeven/trailing stop" if position["triggered"] else "Stop loss"
                else:
                    if not position["triggered"] and c["low"] <= position["tp"]:
                        position["triggered"] = True
                        position["triggerTime"] = c["time"]
                        position["realizedR"] += position["lock"] * ((position["entry"] - position["tp"]) / position["risk"])
                        position["stop"] = position["entry"]
                    if position["triggered"] and position["trail"] == "chandelier":
                        position["ll"] = min(position["ll"], c["low"])
                        position["stop"] = min(position["stop"], position["ll"] + 3 * c["atr14"])
                    if variant["runner"] == 0 and c["low"] <= position["tp"]:
                        exit_price, reason = position["tp"], "TP 2.5 ATR full close"
                    elif position["triggered"] and ssl_flip:
                        exit_price, reason = c["close"], "Runner exit: SSL bullish flip"

            if exit_price is not None:
                remaining = position["runner"] if position["triggered"] else 1.0
                rem_r = (exit_price - position["entry"]) / position["risk"] if side == "LONG" else (position["entry"] - exit_price) / position["risk"]
                gross = position["realizedR"] + remaining * rem_r
                net = gross - cost_r(position["entry"], position["risk"])
                trades.append({
                    "symbol": symbol,
                    "tradeNo": trade_no,
                    "side": side,
                    "signalType": position["signalType"],
                    "signalTime": iso(position["signalTime"]),
                    "entryTime": iso(position["entryTime"]),
                    "entryPrice": position["entry"],
                    "initialStop": position["initialStop"],
                    "finalStop": position["stop"],
                    "riskPerUnit": position["risk"],
                    "tp1": position["tp"],
                    "tp1Time": iso(position["triggerTime"]) if position["triggerTime"] else "",
                    "exitTime": iso(c["time"]),
                    "exitPrice": exit_price,
                    "exitReason": reason,
                    "grossR": gross,
                    "costR": cost_r(position["entry"], position["risk"]),
                    "rMultiple": net,
                    "atr14": position["atr14"],
                    "rsi14": position["rsi14"],
                    "distanceToEma50Atr": position["distance"],
                    "notes": position["notes"],
                })
                trade_no += 1
                position = None
            if position:
                continue

        if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]) or prev["ssl"] is None:
            continue
        dist = abs(c["close"] - c["ema50"]) / c["atr14"]
        long_flip = prev["ssl"] == -1 and c["ssl"] == 1 and recent_cross(candles, i, "LONG") and dist <= 2 and c["rsi14"] > 50
        short_flip = prev["ssl"] == 1 and c["ssl"] == -1 and recent_cross(candles, i, "SHORT") and dist <= 2 and c["rsi14"] < 50
        cont_max_dist = variant.get("contMaxDist", 2.8)
        long_cont = variant["continuation"] and continuation_ok(candles, i, "LONG", variant) and dist <= cont_max_dist
        short_cont = variant["continuation"] and continuation_ok(candles, i, "SHORT", variant) and dist <= cont_max_dist
        if not (long_flip or short_flip or long_cont or short_cont):
            continue
        side = "LONG" if long_flip or long_cont else "SHORT"
        signal_type = "Continuation" if long_cont or short_cont else "Primary flip"
        entry = nxt["open"]
        risk = c["atr14"] * 1.5
        stop = entry - risk if side == "LONG" else entry + risk
        tp = entry + c["atr14"] * 2.5 if side == "LONG" else entry - c["atr14"] * 2.5
        position = {
            "side": side,
            "signalType": signal_type,
            "signalTime": c["time"],
            "entryTime": nxt["time"],
            "entry": entry,
            "initialStop": stop,
            "stop": stop,
            "risk": risk,
            "tp": tp,
            "lock": variant["lock"],
            "runner": variant["runner"],
            "trail": variant["trail"],
            "triggered": False,
            "triggerTime": None,
            "realizedR": 0.0,
            "hh": entry,
            "ll": entry,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": dist,
            "notes": f"{signal_type}; {variant['name']}; ATR uses TradingView-style RMA.",
        }

    return trades


def max_dd(trades):
    cum = peak = dd = 0.0
    for t in sorted(trades, key=lambda x: x["exitTime"]):
        cum += t["rMultiple"]
        peak = max(peak, cum)
        dd = min(dd, cum - peak)
    return dd


def stats(trades):
    total = sum(t["rMultiple"] for t in trades)
    wins = sum(1 for t in trades if t["rMultiple"] > 0)
    cont = sum(1 for t in trades if t["signalType"] == "Continuation")
    return {
        "trades": len(trades),
        "wins": wins,
        "losses": len(trades) - wins,
        "winRate": wins / len(trades) if trades else 0,
        "totalR": total,
        "avgR": total / len(trades) if trades else 0,
        "maxDrawdownR": max_dd(trades),
        "bestR": max((t["rMultiple"] for t in trades), default=0),
        "worstR": min((t["rMultiple"] for t in trades), default=0),
        "continuationTrades": cont,
        "continuationR": sum(t["rMultiple"] for t in trades if t["signalType"] == "Continuation"),
    }


def template_layout(template_ws, target_ws):
    target_ws.sheet_format.defaultColWidth = template_ws.sheet_format.defaultColWidth
    target_ws.sheet_format.defaultRowHeight = template_ws.sheet_format.defaultRowHeight
    target_ws.sheet_view.showGridLines = template_ws.sheet_view.showGridLines
    for col_idx in range(1, max(template_ws.max_column, target_ws.max_column) + 1):
        letter = get_column_letter(col_idx)
        if template_ws.column_dimensions.get(letter) and template_ws.column_dimensions[letter].width:
            target_ws.column_dimensions[letter].width = template_ws.column_dimensions[letter].width
    for row_idx in range(1, max(template_ws.max_row, target_ws.max_row) + 1):
        src_idx = row_idx if row_idx <= template_ws.max_row else 5
        if template_ws.row_dimensions.get(src_idx) and template_ws.row_dimensions[src_idx].height:
            target_ws.row_dimensions[row_idx].height = template_ws.row_dimensions[src_idx].height
        for col_idx in range(1, max(template_ws.max_column, target_ws.max_column) + 1):
            src = template_ws.cell(src_idx, min(col_idx, template_ws.max_column))
            dst = target_ws.cell(row_idx, col_idx)
            if src.has_style:
                dst._style = copy(src._style)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)


def clear(ws, rows=360, cols=30):
    for row in ws.iter_rows(min_row=1, max_row=max(rows, ws.max_row), min_col=1, max_col=max(cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def write_row(ws, row, values):
    for col, value in enumerate(values, 1):
        ws.cell(row, col).value = value


def build_workbook(result):
    template = load_workbook(TEMPLATE)
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        clear(ws)
    summary = wb["Summary"]
    summary["A1"] = "NXT v3.1 - 80/20 TP1 and Trend Continuation Tests"
    summary["A2"] = "ATR uses TradingView-style Wilder/RMA. Outputs use project template."
    write_row(summary, 4, ["Variant", "Trades", "Win Rate", "Total R", "Avg R", "Max DD R", "Best R", "Worst R", "Continuation Trades", "Continuation R"])
    for i, row in enumerate(result["ranking"], 5):
        write_row(summary, i, [row["name"], row["trades"], row["winRate"], row["totalR"], row["avgR"], row["maxDrawdownR"], row["bestR"], row["worstR"], row["continuationTrades"], row["continuationR"]])

    best = result["variants"][result["ranking"][0]["key"]]
    headers = ["Symbol", "No", "Side", "Signal Type", "Signal Time", "Entry Time", "Entry Price", "Initial Stop", "Final Stop", "Risk / Unit", "TP1", "TP1 Time", "Exit Time", "Exit Price", "Exit Reason", "Gross R", "Cost R", "Net R", "ATR14", "RSI14", "Distance EMA50 ATR", "Notes"]
    for sheet_name, rows in [("Trades", best["trades"]), *[(s.replace("USDT", ""), [t for t in best["trades"] if t["symbol"] == s]) for s in SYMBOLS]]:
        ws = wb[sheet_name]
        ws["A1"] = f"{sheet_name} - {best['name']}"
        ws["A2"] = "Best-ranked variant trade list."
        write_row(ws, 4, headers)
        for r, t in enumerate(rows, 5):
            write_row(ws, r, [t["symbol"].replace("USDT", ""), t["tradeNo"], t["side"], t["signalType"], t["signalTime"].replace("T", " ").replace("Z", ""), t["entryTime"].replace("T", " ").replace("Z", ""), t["entryPrice"], t["initialStop"], t["finalStop"], t["riskPerUnit"], t["tp1"], t["tp1Time"].replace("T", " ").replace("Z", "") if t["tp1Time"] else "", t["exitTime"].replace("T", " ").replace("Z", ""), t["exitPrice"], t["exitReason"], t["grossR"], t["costR"], t["rMultiple"], t["atr14"], t["rsi14"], t["distanceToEma50Atr"], t["notes"]])

    ass = wb["Assumptions"]
    ass["A1"] = "Assumptions"
    lines = [
        "NXT v3.1 entry base: SSL flip, EMA20 cross within 3 candles, RSI 50/50, distance to EMA50 <= 2 ATR.",
        "TP1 is 2.5 ATR. 80/20 variants close 80% at TP1 and let 20% trail.",
        "SSL trailing exits runner on opposite SSL flip. Chandelier trailing uses 3 ATR from highest high/lowest low after TP1.",
        "Continuation rule: trend already aligned with SSL and EMA20/EMA50; RSI >55 for long or <45 for short; recent EMA20 touch/reclaim; distance <= 2.8 ATR.",
        "ATR uses TradingView-style Wilder RMA, not SMA.",
    ]
    write_row(ass, 4, ["#", "Assumption"])
    for i, line in enumerate(lines, 5):
        write_row(ass, i, [i - 4, line])
    for name in ["Summary", "Trades", "BTC", "SOL", "SUI", "Assumptions", "Data Quality", "20K Account", "Equity Curve"]:
        if name in wb.sheetnames and name in template.sheetnames:
            template_layout(template[name], wb[name])
    wb.save(OUT_XLSX)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    data = {s: enrich(json.loads((DATA / f"{s}.json").read_text(encoding="utf-8"))) for s in SYMBOLS}
    variants = {}
    for variant in VARIANTS:
        trades = []
        for symbol in SYMBOLS:
            trades.extend(backtest_symbol(symbol, data[symbol], variant))
        trades.sort(key=lambda x: x["exitTime"])
        variants[variant["key"]] = {"name": variant["name"], "config": variant, "stats": stats(trades), "trades": trades}
    ranking = sorted(({ "key": k, "name": v["name"], **v["stats"] } for k, v in variants.items()), key=lambda x: x["totalR"], reverse=True)
    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "period": {"start": iso(START), "end": iso(END - 1)},
        "symbols": SYMBOLS,
        "source": "Binance spot daily klines cache",
        "atrMethod": "Wilder RMA / TradingView default",
        "ranking": ranking,
        "variants": variants,
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    build_workbook(result)
    print(json.dumps({"outJson": str(OUT_JSON), "outXlsx": str(OUT_XLSX), "ranking": ranking}, indent=2))


if __name__ == "__main__":
    main()
