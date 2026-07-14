from __future__ import annotations

import json
import sys
import time
import urllib.parse
import urllib.request
from copy import copy
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

import backtest_nxt31_utc7_latest as base
import backtest_nxt32_native_1d_latest as native
import test_nxt33_long_only_pullback_continuation as cont
from test_nxt33_ssl14 import enrich_with_ssl_period


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
CACHE = ROOT / "data_cache" / "yahoo_us_stocks_1d"
OUT_DIR = ROOT / "outputs" / "nxt35_us_top3_stocks_3y"
OUT_JSON = OUT_DIR / "NXT35_US_Top3_Stocks_3Y.json"
OUT_XLSX = OUT_DIR / "NXT35_US_Top3_Stocks_3Y_20K.xlsx"

SYMBOLS = ["NVDA", "AAPL", "GOOGL"]
MARKET_LABEL = "US Stocks"
MARKET_CAP_SOURCE = "CompaniesMarketCap snapshot checked 2026-07-09: NVDA, AAPL, Alphabet/GOOG. GOOGL used as the actively traded Alphabet class."
START_DATE = date(2023, 7, 9)
END_DATE = date(2026, 7, 10)  # exclusive for signal entries; allows the latest available Yahoo close.
WARMUP_DATE = date(2023, 1, 1)
STARTING_EQUITY = 20_000
ONE_R_DOLLARS = 1_000


def run_label() -> str:
    years = round((END_DATE - START_DATE).days / 365.2425)
    return f"{years}Y"


def unix_day(day: date) -> int:
    return int(datetime(day.year, day.month, day.day, tzinfo=timezone.utc).timestamp())


def http_json(url: str) -> object:
    req = urllib.request.Request(url, headers={"User-Agent": "nxt-us-stock-backtest/1.0"})
    with urllib.request.urlopen(req, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def fetch_yahoo_1d(symbol: str) -> list[dict]:
    CACHE.mkdir(parents=True, exist_ok=True)
    cache_path = CACHE / f"{symbol}_{WARMUP_DATE}_{END_DATE}.json"
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))

    params = urllib.parse.urlencode(
        {
            "period1": unix_day(WARMUP_DATE),
            "period2": unix_day(END_DATE + timedelta(days=1)),
            "interval": "1d",
            "events": "history",
            "includeAdjustedClose": "true",
        }
    )
    data = http_json(f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?{params}")
    result = data["chart"]["result"][0]
    timestamps = result.get("timestamp") or []
    quote = result["indicators"]["quote"][0]
    adjclose = (result["indicators"].get("adjclose") or [{}])[0].get("adjclose") or quote["close"]

    rows = []
    for i, ts in enumerate(timestamps):
        raw = {k: quote[k][i] for k in ["open", "high", "low", "close", "volume"]}
        if any(raw[k] is None for k in ["open", "high", "low", "close"]):
            continue
        close = float(raw["close"])
        adj = float(adjclose[i]) if adjclose[i] is not None else close
        factor = adj / close if close else 1.0
        day = datetime.fromtimestamp(ts, timezone.utc).date()
        rows.append(
            {
                "time": int(ts * 1000),
                "localDate": day.isoformat(),
                "open": float(raw["open"]) * factor,
                "high": float(raw["high"]) * factor,
                "low": float(raw["low"]) * factor,
                "close": adj,
                "volume": float(raw["volume"] or 0),
            }
        )
    rows = [r for r in rows if WARMUP_DATE <= date.fromisoformat(r["localDate"]) <= END_DATE]
    cache_path.write_text(json.dumps(rows, indent=2), encoding="utf-8")
    time.sleep(0.2)
    return rows


def profit_factor(rows: list[dict]) -> float | None:
    gross_profit = sum(t["rMultiple"] for t in rows if t["rMultiple"] > 0)
    gross_loss = -sum(t["rMultiple"] for t in rows if t["rMultiple"] < 0)
    return gross_profit / gross_loss if gross_loss else None


def stats(rows: list[dict]) -> dict:
    out = base.stats(rows)
    out["profitFactor"] = profit_factor(rows)
    out["ending20k"] = STARTING_EQUITY + out["totalR"] * ONE_R_DOLLARS
    return out


def group_stats(rows: list[dict], key_name: str) -> list[dict]:
    groups: dict[str, list[dict]] = {}
    for row in rows:
        groups.setdefault(row[key_name], []).append(row)
    result = []
    for key, subset in sorted(groups.items()):
        item = stats(subset)
        item[key_name] = key
        result.append(item)
    return result


def by_year_stats(rows: list[dict]) -> list[dict]:
    groups: dict[str, list[dict]] = {}
    for row in rows:
        groups.setdefault(row["exitTime"][:4], []).append(row)
    result = []
    for year, subset in sorted(groups.items()):
        item = stats(subset)
        item["year"] = year
        result.append(item)
    return result


def equity_curve(rows: list[dict]) -> list[dict]:
    equity = STARTING_EQUITY
    peak = equity
    curve = []
    for i, trade in enumerate(rows, 1):
        pnl = trade["rMultiple"] * ONE_R_DOLLARS
        equity += pnl
        peak = max(peak, equity)
        curve.append(
            {
                "no": i,
                "exitTime": trade["exitTime"],
                "symbol": trade["symbol"],
                "side": trade["side"],
                "signalType": trade["signalType"],
                "rMultiple": trade["rMultiple"],
                "pnl": pnl,
                "equity": equity,
                "drawdown": equity - peak,
            }
        )
    return curve


def clear(ws, rows=1200, cols=40):
    for row in ws.iter_rows(min_row=1, max_row=max(rows, ws.max_row), min_col=1, max_col=max(cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def rowset(ws, r, values):
    for c, v in enumerate(values, 1):
        ws.cell(r, c).value = v


def style_table(ws, header_row=4):
    ws.freeze_panes = f"A{header_row + 1}"
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(max(ws.column_dimensions[get_column_letter(col)].width or 12, 12), 24)


def copy_layout(src, dst):
    dst.sheet_format.defaultColWidth = src.sheet_format.defaultColWidth
    dst.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    for c in range(1, max(src.max_column, dst.max_column) + 1):
        letter = get_column_letter(c)
        if src.column_dimensions.get(letter) and src.column_dimensions[letter].width:
            dst.column_dimensions[letter].width = src.column_dimensions[letter].width
    for r in range(1, max(src.max_row, dst.max_row) + 1):
        sr = r if r <= src.max_row else 5
        if src.row_dimensions.get(sr) and src.row_dimensions[sr].height:
            dst.row_dimensions[r].height = src.row_dimensions[sr].height
        for c in range(1, max(src.max_column, dst.max_column) + 1):
            a = src.cell(sr, min(c, src.max_column))
            b = dst.cell(r, c)
            if a.has_style:
                b._style = copy(a._style)
            b.number_format = a.number_format
            b.alignment = copy(a.alignment)
            b.font = copy(a.font)
            b.fill = copy(a.fill)
            b.border = copy(a.border)


def write_workbook(result: dict) -> None:
    tpl = load_workbook(TEMPLATE)
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        clear(ws)

    summary = wb["Summary"]
    summary["A1"] = f"NXT v3.5 - Top 3 {MARKET_LABEL} {run_label()}"
    summary["A2"] = f"{START_DATE.isoformat()} to {(END_DATE - timedelta(days=1)).isoformat()} | 1R = ${ONE_R_DOLLARS:,} | Data: Yahoo adjusted daily OHLCV"
    rowset(summary, 4, ["Metric", "Value"])
    for r, (label, key) in enumerate(
        [
            ("Trades", "trades"),
            ("Win Rate", "winRate"),
            ("Total R", "totalR"),
            ("Average R", "avgR"),
            ("Max DD R", "maxDrawdownR"),
            ("Best R", "bestR"),
            ("Worst R", "worstR"),
            ("Profit Factor", "profitFactor"),
            ("20K Ending", "ending20k"),
        ],
        5,
    ):
        rowset(summary, r, [label, result["stats"][key]])
    style_table(summary)

    headers = [
        "Symbol", "No", "Signal Type", "Side", "Signal Date", "Entry Date", "Entry Price",
        "Initial Stop", "Final Stop", "Risk / Unit", "TP1", "TP1 Date", "Early BE", "Early BE Date",
        "Exit Date", "Exit Price", "Exit Reason", "Gross R", "Cost R", "Net R", "ATR14", "RSI14",
        "Distance EMA50 ATR", "EMA20", "EMA50", "Notes",
    ]
    trades = wb["Trades"]
    trades["A1"] = "Detailed Trades - NXT v3.5 US Stocks"
    trades["A2"] = "One completed trade per row."
    rowset(trades, 4, headers)
    for i, t in enumerate(result["trades"], 5):
        rowset(
            trades,
            i,
            [
                t["symbol"], t["tradeNo"], t["signalType"], t["side"], t["signalTime"], t["entryTime"],
                t["entryPrice"], t["initialStop"], t["finalStop"], t["riskPerUnit"], t["tp1"], t["tp1Time"],
                t["earlyBeTriggered"], t["earlyBeTime"], t["exitTime"], t["exitPrice"], t["exitReason"],
                t["grossRMultiple"], t["costR"], t["rMultiple"], t["atr14"], t["rsi14"],
                t["distanceToEma50Atr"], t["ema20"], t["ema50"], t["notes"],
            ],
        )
    style_table(trades)
    trades.column_dimensions["Z"].width = 44

    for sheet_name, title, rows, key in [
        ("BTC", "Stats By Symbol", result["bySymbol"], "symbol"),
        ("SOL", "Stats By Exit Year", result["byYear"], "year"),
    ]:
        ws = wb[sheet_name]
        ws.title = title
        ws["A1"] = title
        rowset(ws, 4, [key.title(), "Trades", "Win Rate", "Total R", "Avg R", "Max DD R", "Best R", "Worst R", "Profit Factor", "20K Ending"])
        for i, row in enumerate(rows, 5):
            rowset(ws, i, [row[key], row["trades"], row["winRate"], row["totalR"], row["avgR"], row["maxDrawdownR"], row["bestR"], row["worstR"], row["profitFactor"], row["ending20k"]])
        style_table(ws)

    account = wb["20K Account"]
    account["A1"] = "20K Account"
    rowset(account, 4, ["Starting Equity", "1R", "Total R", "Ending Equity", "Max DD R"])
    rowset(account, 5, [STARTING_EQUITY, ONE_R_DOLLARS, result["stats"]["totalR"], result["stats"]["ending20k"], result["stats"]["maxDrawdownR"]])
    style_table(account)

    curve = wb["Equity Curve"]
    curve["A1"] = "Equity Curve"
    rowset(curve, 4, ["No", "Exit Date", "Symbol", "Side", "Signal Type", "Net R", "PnL", "Equity", "Drawdown"])
    for i, row in enumerate(result["equityCurve"], 5):
        rowset(curve, i, [row["no"], row["exitTime"], row["symbol"], row["side"], row["signalType"], row["rMultiple"], row["pnl"], row["equity"], row["drawdown"]])
    style_table(curve)

    assumptions = wb["Assumptions"]
    assumptions["A1"] = "Assumptions"
    rowset(assumptions, 4, ["#", "Assumption"])
    for i, line in enumerate(result["assumptions"], 5):
        rowset(assumptions, i, [i - 4, line])
    style_table(assumptions)
    assumptions.column_dimensions["B"].width = 88

    dq = wb["Data Quality"]
    dq["A1"] = "Data Quality"
    rowset(dq, 4, ["Symbol", "Daily Rows", "First Day", "Last Day", "Source"])
    for i, (symbol, q) in enumerate(result["datasets"].items(), 5):
        rowset(dq, i, [symbol, q["dailyRows"], q["firstDay"], q["lastDay"], q["source"]])
    style_table(dq)

    if "SUI" in wb.sheetnames:
        del wb["SUI"]
    for sheet in ["Summary", "Trades", "Assumptions", "Data Quality", "20K Account", "Equity Curve"]:
        if sheet in wb.sheetnames and sheet in tpl.sheetnames:
            copy_layout(tpl[sheet], wb[sheet])
    wb.save(OUT_XLSX)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    old_start, old_end, old_warmup = native.START_DATE, native.END_DATE, native.WARMUP_DATE
    try:
        native.START_DATE = START_DATE
        native.END_DATE = END_DATE
        native.WARMUP_DATE = WARMUP_DATE
        all_trades = []
        datasets = {}
        for symbol in SYMBOLS:
            candles = fetch_yahoo_1d(symbol)
            enriched = enrich_with_ssl_period(candles, 14)
            trades = cont.backtest_symbol(symbol, enriched)
            all_trades.extend(trades)
            datasets[symbol] = {
                "dailyRows": len(candles),
                "firstDay": candles[0]["localDate"] if candles else None,
                "lastDay": candles[-1]["localDate"] if candles else None,
                "source": "Yahoo Finance chart API daily OHLCV adjusted by adjclose/close",
            }
        all_trades.sort(key=lambda t: (t["exitTime"], t["symbol"], t["tradeNo"]))
    finally:
        native.START_DATE, native.END_DATE, native.WARMUP_DATE = old_start, old_end, old_warmup

    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "systemVersion": f"NXT v3.5 adapted to {MARKET_LABEL} {run_label()}: native daily OHLCV, SSL14, Runner A, Early-BE 7%, anti-immediate-reversal, LONG-only pullback continuation",
        "marketCapSelection": MARKET_CAP_SOURCE,
        "symbols": SYMBOLS,
        "period": {"start": START_DATE.isoformat(), "end": (END_DATE - timedelta(days=1)).isoformat(), "timezone": "US exchange daily bars from Yahoo"},
        "stats": stats(all_trades),
        "bySymbol": group_stats(all_trades, "symbol"),
        "byYear": by_year_stats(all_trades),
        "equityCurve": equity_curve(all_trades),
        "datasets": datasets,
        "assumptions": [
            "Top 3 selection was checked live on 2026-07-09: NVIDIA, Apple, Alphabet. GOOGL is used for Alphabet trading data.",
            "This is a rule-fit smoke test on US equities, not a promoted NXT latest version.",
            "Daily OHLCV comes from Yahoo Finance chart API; open/high/low/close are adjusted by adjclose/close to keep split-adjusted continuity.",
            "No crypto funding overlay is applied to stocks.",
            "The existing NXT cost model is retained: 0.06% fee plus 0.05% slippage per side.",
            "Entry is the next available daily open after the signal candle closes; missing weekends and holidays are skipped naturally.",
            "Short signals are modeled as shortable stock trades without borrow fees or hard-to-borrow constraints.",
            "1R is shown as $1,000 on a $20,000 account for comparability with the existing NXT workbooks.",
        ],
        "trades": all_trades,
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    write_workbook(result)
    print(json.dumps({"outJson": str(OUT_JSON), "outXlsx": str(OUT_XLSX), "stats": result["stats"], "bySymbol": result["bySymbol"], "datasets": result["datasets"]}, indent=2))


if __name__ == "__main__":
    main()
