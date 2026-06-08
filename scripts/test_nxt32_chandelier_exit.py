from __future__ import annotations

import json
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import backtest_nxt31_utc7_latest as base


ROOT = Path(r"D:\Workspace\Codex\Crypto trading")
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
OUT_DIR = ROOT / "outputs" / "nxt32_chandelier_exit_grid_6y"
OUT_JSON = OUT_DIR / "nxt32_chandelier_exit_grid_6y_results.json"
OUT_XLSX = OUT_DIR / "NXT32_Chandelier_Exit_Grid_6Y_BTC_SOL_SUI_20K.xlsx"

MULTIPLIERS = [2.5, 3.0, 3.5, 4.0]


def backtest_symbol(symbol: str, candles: list[dict], chandelier_mult: float) -> list[dict]:
    trades, pos, n = [], None, 1
    for i in range(55, len(candles) - 1):
        c, prev, nxt = candles[i], candles[i - 1], candles[i + 1]
        next_date = base.date.fromisoformat(nxt["localDate"])
        if next_date < base.START_DATE or next_date >= base.END_DATE:
            continue

        if pos:
            side = pos["side"]
            exit_price = reason = None
            if side == "LONG":
                if c["low"] <= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Chandelier stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["high"] >= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["tp"] - pos["entry"]) / pos["risk"])
                    if pos["triggered"]:
                        pos["highestHigh"] = max(pos["highestHigh"], c["high"])
                        pos["stop"] = max(pos["stop"], pos["highestHigh"] - chandelier_mult * c["atr14"])
            else:
                if c["high"] >= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Chandelier stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["low"] <= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["entry"] - pos["tp"]) / pos["risk"])
                    if pos["triggered"]:
                        pos["lowestLow"] = min(pos["lowestLow"], c["low"])
                        pos["stop"] = min(pos["stop"], pos["lowestLow"] + chandelier_mult * c["atr14"])

            if exit_price is not None:
                rem = 0.5 if pos["triggered"] else 1.0
                rem_r = (exit_price - pos["entry"]) / pos["risk"] if side == "LONG" else (pos["entry"] - exit_price) / pos["risk"]
                gross = pos["realizedR"] + rem * rem_r
                net = gross - base.cost_r(pos["entry"], pos["risk"])
                trades.append(
                    {
                        "symbol": symbol,
                        "tradeNo": n,
                        "side": side,
                        "signalTime": pos["signalDate"],
                        "entryTime": pos["entryDate"],
                        "entryPrice": pos["entry"],
                        "initialStop": pos["initialStop"],
                        "finalStop": pos["stop"],
                        "riskPerUnit": pos["risk"],
                        "tp1": pos["tp"],
                        "tp1Time": pos["tp1Time"],
                        "exitTime": c["localDate"],
                        "exitPrice": exit_price,
                        "exitReason": reason,
                        "grossRMultiple": gross,
                        "costR": base.cost_r(pos["entry"], pos["risk"]),
                        "rMultiple": net,
                        "atr14": pos["atr14"],
                        "rsi14": pos["rsi14"],
                        "distanceToEma50Atr": pos["distance"],
                        "notes": f"UTC+7 daily candles; TP1 2.5 ATR 50%; runner exits by Chandelier ATR14 x {chandelier_mult}",
                    }
                )
                n += 1
                pos = None
            if pos:
                continue

        if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]) or prev["ssl"] is None:
            continue
        dist = abs(c["close"] - c["ema50"]) / c["atr14"]
        long_ok = prev["ssl"] == -1 and c["ssl"] == 1 and base.recent_cross(candles, i, "LONG") and dist <= 2 and c["rsi14"] > 50
        short_ok = prev["ssl"] == 1 and c["ssl"] == -1 and base.recent_cross(candles, i, "SHORT") and dist <= 2 and c["rsi14"] < 50
        if not (long_ok or short_ok):
            continue
        side = "LONG" if long_ok else "SHORT"
        risk = c["atr14"] * 1.5
        entry = nxt["open"]
        initial_stop = entry - risk if side == "LONG" else entry + risk
        pos = {
            "side": side,
            "signalDate": c["localDate"],
            "entryDate": nxt["localDate"],
            "entry": entry,
            "initialStop": initial_stop,
            "stop": initial_stop,
            "risk": risk,
            "tp": entry + c["atr14"] * 2.5 if side == "LONG" else entry - c["atr14"] * 2.5,
            "triggered": False,
            "tp1Time": "",
            "realizedR": 0.0,
            "highestHigh": entry,
            "lowestLow": entry,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": dist,
        }
    return trades


def account_stats(trades: list[dict]) -> dict:
    by_symbol = {}
    for symbol in base.SYMBOLS:
        rows = [t for t in trades if t["symbol"] == symbol]
        by_symbol[symbol] = base.stats(rows)
    end_equity = 20000 + sum(t["rMultiple"] for t in trades) * 1000
    out = base.stats(trades)
    out["ending20kAccount"] = end_equity
    out["bySymbol"] = by_symbol
    out["tp1Hits"] = sum(1 for t in trades if t["tp1Time"])
    out["tp1HitRate"] = out["tp1Hits"] / len(trades) if trades else 0
    return out


def clear(ws, rows=700, cols=32):
    for row in ws.iter_rows(min_row=1, max_row=max(rows, ws.max_row), min_col=1, max_col=max(cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def rowset(ws, r, values):
    for c, v in enumerate(values, 1):
        ws.cell(r, c).value = v


def copy_layout(src, dst):
    dst.sheet_format.defaultColWidth = src.sheet_format.defaultColWidth
    dst.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    for c in range(1, max(src.max_column, dst.max_column) + 1):
        letter = get_column_letter(c)
        if src.column_dimensions.get(letter) and src.column_dimensions[letter].width:
            dst.column_dimensions[letter].width = src.column_dimensions[letter].width
    for r in range(1, max(src.max_row, dst.max_row) + 1):
        sr = r if r <= src.max_row else 5
        if src.row_dimensions.get(sr) and src.row_dimensions[sr].height:
            dst.row_dimensions[r].height = src.row_dimensions[sr].height
        for c in range(1, max(src.max_column, dst.max_column) + 1):
            a = src.cell(sr, min(c, src.max_column))
            b = dst.cell(r, c)
            if a.has_style:
                b._style = copy(a._style)
            b.number_format = a.number_format
            b.alignment = copy(a.alignment)
            b.font = copy(a.font)
            b.fill = copy(a.fill)
            b.border = copy(a.border)


def build_workbook(result: dict):
    tpl = load_workbook(TEMPLATE)
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        clear(ws)

    ws = wb["Summary"]
    ws["A1"] = "NXT v3.2 Chandelier Exit Grid"
    ws["A2"] = "BTC/SOL/SUI 6Y | UTC+7 | no continuation | no risk-off | only runner exit changed."
    headers = ["Variant", "Trades", "Win Rate", "TP1 Hit Rate", "Total R", "Avg R", "Max DD R", "Best R", "Worst R", "20K Ending"]
    rowset(ws, 4, headers)
    for r, item in enumerate(result["variants"], 5):
        s = item["stats"]
        rowset(ws, r, [item["name"], s["trades"], s["winRate"], s["tp1HitRate"], s["totalR"], s["avgR"], s["maxDrawdownR"], s["bestR"], s["worstR"], s["ending20kAccount"]])

    best = max(result["variants"], key=lambda x: x["stats"]["totalR"])
    headers = ["Symbol", "No", "Side", "Signal Date UTC+7", "Entry Date UTC+7", "Entry Price", "Initial Stop", "Final Stop", "Risk / Unit", "TP1", "TP1 Date", "Exit Date UTC+7", "Exit Price", "Exit Reason", "R", "ATR14", "RSI14", "Distance EMA50 ATR", "Notes"]
    for sheet, subset in [
        ("Trades", best["trades"]),
        ("BTC", [t for t in best["trades"] if t["symbol"] == "BTCUSDT"]),
        ("SOL", [t for t in best["trades"] if t["symbol"] == "SOLUSDT"]),
        ("SUI", [t for t in best["trades"] if t["symbol"] == "SUIUSDT"]),
    ]:
        ws = wb[sheet]
        ws["A1"] = f"{sheet} - {best['name']}" if sheet != "Trades" else f"Detailed Trades - {best['name']}"
        ws["A2"] = "One completed trade per row. Detail sheets show best Total R variant from this grid."
        rowset(ws, 4, headers)
        for i, t in enumerate(subset, 5):
            rowset(ws, i, [t["symbol"].replace("USDT", ""), t["tradeNo"], t["side"], t["signalTime"], t["entryTime"], t["entryPrice"], t["initialStop"], t["finalStop"], t["riskPerUnit"], t["tp1"], t["tp1Time"], t["exitTime"], t["exitPrice"], t["exitReason"], t["rMultiple"], t["atr14"], t["rsi14"], t["distanceToEma50Atr"], t["notes"]])

    ws = wb["20K Account"]
    ws["A1"] = "20K Account Projection"
    ws["A2"] = "Assumes 1R = $1,000 on a $20,000 starting account."
    rowset(ws, 4, ["Variant", "Starting Equity", "Total R", "Ending Equity"])
    for r, item in enumerate(result["variants"], 5):
        rowset(ws, r, [item["name"], 20000, item["stats"]["totalR"], item["stats"]["ending20kAccount"]])

    ws = wb["Assumptions"]
    ws["A1"] = "Assumptions"
    rowset(ws, 4, ["#", "Assumption"])
    for i, line in enumerate(result["assumptions"], 5):
        rowset(ws, i, [i - 4, line])

    ws = wb["Data Quality"]
    ws["A1"] = "Data Quality"
    rowset(ws, 4, ["Symbol", "1H Rows", "UTC+7 Daily Candles", "First UTC+7 Day", "Last UTC+7 Day", "Source"])
    for i, (sym, q) in enumerate(result["datasets"].items(), 5):
        rowset(ws, i, [sym.replace("USDT", ""), q["hourlyRows"], q["dailyRows"], q["firstDay"], q["lastDay"], q["source"]])

    for sheet in ["Summary", "Trades", "BTC", "SOL", "SUI", "Assumptions", "Data Quality", "20K Account", "Equity Curve"]:
        if sheet in wb.sheetnames and sheet in tpl.sheetnames:
            copy_layout(tpl[sheet], wb[sheet])
    wb.save(OUT_XLSX)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    datasets = {}
    candles_by_symbol = {}
    for symbol in base.SYMBOLS:
        h = base.fetch_1h(symbol)
        d = base.enrich(base.resample_utc7(h))
        candles_by_symbol[symbol] = d
        datasets[symbol] = {
            "hourlyRows": len(h),
            "dailyRows": len(d),
            "firstDay": d[0]["localDate"],
            "lastDay": d[-1]["localDate"],
            "source": "Binance spot 1H klines resampled to Asia/Saigon UTC+7 day",
        }

    variants = []
    for mult in MULTIPLIERS:
        trades = []
        for symbol, candles in candles_by_symbol.items():
            trades.extend(backtest_symbol(symbol, candles, mult))
        trades.sort(key=lambda x: x["exitTime"])
        variants.append(
            {
                "key": f"chandelier_{mult:g}",
                "name": f"Chandelier ATR14 x {mult:g}",
                "chandelierMultiplier": mult,
                "stats": account_stats(trades),
                "trades": trades,
            }
        )

    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "systemVersion": "NXT v3.2 Simple UTC+7 Runner A, runner exit replaced by Chandelier Exit grid",
        "period": {"start": base.START_DATE.isoformat(), "end": (base.END_DATE - base.timedelta(days=1)).isoformat(), "timezone": "Asia/Saigon UTC+7"},
        "symbols": base.SYMBOLS,
        "variants": variants,
        "datasets": datasets,
        "assumptions": [
            "Entry rules are unchanged from latest NXT v3.2 Simple: SSL flip, EMA20 cross within 3 candles, RSI 50/50, EMA50 distance <= 2 ATR.",
            "No continuation module and no risk-off overlay are used in this test.",
            "TP1 remains 2.5 ATR and closes 50% of the position; remaining 50% uses Chandelier Exit ATR14 multiplier.",
            "Before TP1, initial stop remains 1.5 ATR from entry.",
            "After TP1, stop is moved to breakeven and then trailed by Chandelier: LONG highest high since entry minus ATR14 x multiplier; SHORT lowest low since entry plus ATR14 x multiplier.",
            "Cost model remains 0.06% fee and 0.05% slippage per side; funding is not included.",
        ],
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    build_workbook(result)
    compact = [{k: v for k, v in item.items() if k not in {"trades"}} for item in variants]
    print(json.dumps({"outJson": str(OUT_JSON), "outXlsx": str(OUT_XLSX), "variants": compact}, indent=2))


if __name__ == "__main__":
    main()
