from __future__ import annotations

import csv
import io
import json
import time
import urllib.error
import urllib.request
import zipfile
from copy import copy
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\Workspace\Codex\Crypto trading")
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
OUT_DIR = ROOT / "outputs" / "nxt31_utc7_latest_6y"
OUT_JSON = OUT_DIR / "nxt31_utc7_latest_6y_results.json"
OUT_XLSX = OUT_DIR / "NXT31_UTC7_Latest_RunnerA_RiskOff_6Y_BTC_SOL_SUI_20K.xlsx"
CACHE = ROOT / "data_cache" / "binance_spot_1h"

SYMBOLS = ["BTCUSDT", "SOLUSDT", "SUIUSDT"]
START_DATE = date(2020, 5, 17)
END_DATE = date(2026, 5, 17)
WARMUP_DATE = date(2019, 11, 1)
FEE = 0.0006
SLIPPAGE = 0.0005
ROUND_TRIP = 2 * (FEE + SLIPPAGE)


def month_iter(start: date, end: date):
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        yield y, m
        m += 1
        if m == 13:
            y += 1
            m = 1


def normalize_ms(v: str) -> int:
    n = int(float(v))
    return n // 1000 if n > 10_000_000_000_000 else n


def read_zip_csv(url: str):
    try:
        with urllib.request.urlopen(url, timeout=30) as response:
            payload = response.read()
    except urllib.error.HTTPError as exc:
        if exc.code == 404:
            return []
        raise
    rows = []
    with zipfile.ZipFile(io.BytesIO(payload)) as zf:
        names = [n for n in zf.namelist() if n.endswith(".csv")]
        if not names:
            return []
        with zf.open(names[0]) as f:
            for row in csv.reader(io.TextIOWrapper(f, encoding="utf-8")):
                if not row or row[0] == "open_time":
                    continue
                rows.append(
                    {
                        "time": normalize_ms(row[0]),
                        "open": float(row[1]),
                        "high": float(row[2]),
                        "low": float(row[3]),
                        "close": float(row[4]),
                        "volume": float(row[5]),
                        "closeTime": normalize_ms(row[6]),
                        "takerBuyBaseVolume": float(row[9]),
                    }
                )
    return rows


def fetch_1h(symbol: str):
    CACHE.mkdir(parents=True, exist_ok=True)
    cache_path = CACHE / f"{symbol}_1h_2019-11_2026-05.json"
    if cache_path.exists():
        return json.loads(cache_path.read_text(encoding="utf-8"))
    rows = []
    for y, m in month_iter(WARMUP_DATE, END_DATE):
        url = f"https://data.binance.vision/data/spot/monthly/klines/{symbol}/1h/{symbol}-1h-{y}-{m:02d}.zip"
        batch = read_zip_csv(url)
        if not batch and (y, m) == (END_DATE.year, END_DATE.month):
            day = date(y, m, 1)
            while day <= END_DATE:
                daily_url = f"https://data.binance.vision/data/spot/daily/klines/{symbol}/1h/{symbol}-1h-{day:%Y-%m-%d}.zip"
                batch.extend(read_zip_csv(daily_url))
                day += timedelta(days=1)
        rows.extend(batch)
        if batch:
            print(symbol, y, f"{m:02d}", len(batch))
        time.sleep(0.03)
    rows = sorted({r["time"]: r for r in rows}.values(), key=lambda r: r["time"])
    cache_path.write_text(json.dumps(rows), encoding="utf-8")
    return rows


def resample_utc7(rows):
    bars = []
    current = None
    for r in rows:
        local_date = (datetime.fromtimestamp(r["time"] / 1000, timezone.utc) + timedelta(hours=7)).date()
        if current is None or current["localDate"] != local_date.isoformat():
            current = {
                "localDate": local_date.isoformat(),
                "time": int((datetime.combine(local_date, datetime.min.time(), timezone.utc) - timedelta(hours=7)).timestamp() * 1000),
                "open": r["open"],
                "high": r["high"],
                "low": r["low"],
                "close": r["close"],
                "volume": r["volume"],
                "takerBuyBaseVolume": r["takerBuyBaseVolume"],
            }
            bars.append(current)
        else:
            current["high"] = max(current["high"], r["high"])
            current["low"] = min(current["low"], r["low"])
            current["close"] = r["close"]
            current["volume"] += r["volume"]
            current["takerBuyBaseVolume"] += r["takerBuyBaseVolume"]
    return [b for b in bars if WARMUP_DATE <= date.fromisoformat(b["localDate"]) <= END_DATE]


def sma(values, period):
    out, s = [None] * len(values), 0.0
    for i, v in enumerate(values):
        s += v
        if i >= period:
            s -= values[i - period]
        if i >= period - 1:
            out[i] = s / period
    return out


def ema(values, period):
    out, s = [None] * len(values), 0.0
    k = 2 / (period + 1)
    for i, v in enumerate(values):
        if i < period:
            s += v
        if i == period - 1:
            out[i] = s / period
        elif i >= period:
            out[i] = v * k + out[i - 1] * (1 - k)
    return out


def rsi(values, period=14):
    out = [None] * len(values)
    gain = loss = 0.0
    for i in range(1, len(values)):
        ch = values[i] - values[i - 1]
        g, l = max(ch, 0), max(-ch, 0)
        if i <= period:
            gain += g
            loss += l
            if i == period:
                gain /= period
                loss /= period
                out[i] = 100 if loss == 0 else 100 - 100 / (1 + gain / loss)
        else:
            gain = (gain * (period - 1) + g) / period
            loss = (loss * (period - 1) + l) / period
            out[i] = 100 if loss == 0 else 100 - 100 / (1 + gain / loss)
    return out


def atr_sma(candles, period=14):
    tr = []
    for i, c in enumerate(candles):
        if i == 0:
            tr.append(c["high"] - c["low"])
        else:
            pc = candles[i - 1]["close"]
            tr.append(max(c["high"] - c["low"], abs(c["high"] - pc), abs(c["low"] - pc)))
    return sma(tr, period)


def enrich(candles):
    closes = [c["close"] for c in candles]
    highs = [c["high"] for c in candles]
    lows = [c["low"] for c in candles]
    ema20 = ema(closes, 20)
    ema50 = ema(closes, 50)
    atr14 = atr_sma(candles, 14)
    rsi14 = rsi(closes, 14)
    hs = sma(highs, 10)
    ls = sma(lows, 10)
    ssl, state = [], 0
    for i, c in enumerate(candles):
        if hs[i] is None or ls[i] is None:
            ssl.append(None)
            continue
        if c["close"] > hs[i]:
            state = 1
        elif c["close"] < ls[i]:
            state = -1
        ssl.append(state)
    for i, c in enumerate(candles):
        c.update({"ema20": ema20[i], "ema50": ema50[i], "atr14": atr14[i], "rsi14": rsi14[i], "ssl": ssl[i]})
    return candles


def crossed_up(candles, i):
    return candles[i - 1]["close"] <= candles[i - 1]["ema20"] and candles[i]["close"] > candles[i]["ema20"]


def crossed_down(candles, i):
    return candles[i - 1]["close"] >= candles[i - 1]["ema20"] and candles[i]["close"] < candles[i]["ema20"]


def recent_cross(candles, i, side, lookback=3):
    start = max(1, i - lookback + 1)
    if side == "LONG":
        return any(crossed_up(candles, j) for j in range(start, i + 1))
    return any(crossed_down(candles, j) for j in range(start, i + 1))


def cost_r(entry, risk):
    return entry * ROUND_TRIP / risk


def backtest_symbol(symbol, candles):
    trades, pos, n = [], None, 1
    for i in range(55, len(candles) - 1):
        c, prev, nxt = candles[i], candles[i - 1], candles[i + 1]
        next_date = date.fromisoformat(nxt["localDate"])
        if next_date < START_DATE or next_date >= END_DATE:
            continue
        if pos:
            side = pos["side"]
            ssl_flip = (side == "LONG" and prev["ssl"] == 1 and c["ssl"] == -1) or (side == "SHORT" and prev["ssl"] == -1 and c["ssl"] == 1)
            exit_price = reason = None
            if side == "LONG":
                if c["low"] <= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["high"] >= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["tp"] - pos["entry"]) / pos["risk"])
                    if pos["triggered"] and ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bearish flip"
            else:
                if c["high"] >= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["low"] <= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["entry"] - pos["tp"]) / pos["risk"])
                    if pos["triggered"] and ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bullish flip"
            if exit_price is not None:
                rem = 0.5 if pos["triggered"] else 1.0
                rem_r = (exit_price - pos["entry"]) / pos["risk"] if side == "LONG" else (pos["entry"] - exit_price) / pos["risk"]
                gross = pos["realizedR"] + rem * rem_r
                net = gross - cost_r(pos["entry"], pos["risk"])
                trades.append({
                    "symbol": symbol,
                    "tradeNo": n,
                    "side": side,
                    "signalTime": pos["signalDate"],
                    "entryTime": pos["entryDate"],
                    "entryPrice": pos["entry"],
                    "initialStop": pos["initialStop"],
                    "finalStop": pos["stop"],
                    "riskPerUnit": pos["risk"],
                    "tp1": pos["tp"],
                    "tp1Time": pos["tp1Time"],
                    "exitTime": c["localDate"],
                    "exitPrice": exit_price,
                    "exitReason": reason,
                    "grossRMultiple": gross,
                    "costR": cost_r(pos["entry"], pos["risk"]),
                    "rMultiple": net,
                    "atr14": pos["atr14"],
                    "rsi14": pos["rsi14"],
                    "distanceToEma50Atr": pos["distance"],
                    "notes": "UTC+7 daily candles; NXT v3.1 Runner A",
                })
                n += 1
                pos = None
            if pos:
                continue

        if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]) or prev["ssl"] is None:
            continue
        dist = abs(c["close"] - c["ema50"]) / c["atr14"]
        long_ok = prev["ssl"] == -1 and c["ssl"] == 1 and recent_cross(candles, i, "LONG") and dist <= 2 and c["rsi14"] > 50
        short_ok = prev["ssl"] == 1 and c["ssl"] == -1 and recent_cross(candles, i, "SHORT") and dist <= 2 and c["rsi14"] < 50
        if not (long_ok or short_ok):
            continue
        side = "LONG" if long_ok else "SHORT"
        risk = c["atr14"] * 1.5
        entry = nxt["open"]
        pos = {
            "side": side,
            "signalDate": c["localDate"],
            "entryDate": nxt["localDate"],
            "entry": entry,
            "initialStop": entry - risk if side == "LONG" else entry + risk,
            "stop": entry - risk if side == "LONG" else entry + risk,
            "risk": risk,
            "tp": entry + c["atr14"] * 2.5 if side == "LONG" else entry - c["atr14"] * 2.5,
            "triggered": False,
            "tp1Time": "",
            "realizedR": 0.0,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": dist,
        }
    return trades


def max_dd(rows, key="rMultiple"):
    cum = peak = dd = 0.0
    for t in sorted(rows, key=lambda x: x["exitTime"]):
        cum += t[key]
        peak = max(peak, cum)
        dd = min(dd, cum - peak)
    return dd


def stats(rows, key="rMultiple"):
    total = sum(t[key] for t in rows)
    wins = sum(1 for t in rows if t[key] > 0)
    return {
        "trades": len(rows),
        "wins": wins,
        "losses": len(rows) - wins,
        "winRate": wins / len(rows) if rows else 0,
        "totalR": total,
        "avgR": total / len(rows) if rows else 0,
        "maxDrawdownR": max_dd(rows, key),
        "bestR": max((t[key] for t in rows), default=0),
        "worstR": min((t[key] for t in rows), default=0),
    }


def riskoff(rows, threshold=-4.0, scale=0.4):
    out, cum, peak = [], 0.0, 0.0
    for t in sorted(rows, key=lambda x: x["exitTime"]):
        r = dict(t)
        pre = cum - peak
        mult = scale if pre <= threshold else 1.0
        r["preTradeDrawdownR"] = pre
        r["sizeMultiplier"] = mult
        r["riskOffR"] = r["rMultiple"] * mult
        cum += r["riskOffR"]
        peak = max(peak, cum)
        r["postTradeDrawdownR"] = cum - peak
        out.append(r)
    return out


def clear(ws, rows=360, cols=30):
    for row in ws.iter_rows(min_row=1, max_row=max(rows, ws.max_row), min_col=1, max_col=max(cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def rowset(ws, r, values):
    for c, v in enumerate(values, 1):
        ws.cell(r, c).value = v


def copy_layout(src, dst):
    dst.sheet_format.defaultColWidth = src.sheet_format.defaultColWidth
    dst.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    for c in range(1, max(src.max_column, dst.max_column) + 1):
        letter = get_column_letter(c)
        if src.column_dimensions.get(letter) and src.column_dimensions[letter].width:
            dst.column_dimensions[letter].width = src.column_dimensions[letter].width
    for r in range(1, max(src.max_row, dst.max_row) + 1):
        sr = r if r <= src.max_row else 5
        if src.row_dimensions.get(sr) and src.row_dimensions[sr].height:
            dst.row_dimensions[r].height = src.row_dimensions[sr].height
        for c in range(1, max(src.max_column, dst.max_column) + 1):
            a = src.cell(sr, min(c, src.max_column))
            b = dst.cell(r, c)
            if a.has_style:
                b._style = copy(a._style)
            b.number_format = a.number_format
            b.alignment = copy(a.alignment)
            b.font = copy(a.font)
            b.fill = copy(a.fill)
            b.border = copy(a.border)


def build_workbook(result):
    tpl = load_workbook(TEMPLATE)
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        clear(ws)
    ws = wb["Summary"]
    ws["A1"] = "NXT v3.1 Latest Rerun - UTC+7 Daily Candles"
    ws["A2"] = "BTC/SOL/SUI 6Y | Runner A + risk-off | Daily candles aligned to Asia/Saigon / UTC+7."
    rows = [
        ["Metric", "Before Risk-Off", "After Risk-Off"],
        ["Trades", result["stats"]["trades"], result["riskOffStats"]["trades"]],
        ["Win Rate", result["stats"]["winRate"], result["riskOffStats"]["winRate"]],
        ["Total R", result["stats"]["totalR"], result["riskOffStats"]["totalR"]],
        ["Average R", result["stats"]["avgR"], result["riskOffStats"]["avgR"]],
        ["Max DD R", result["stats"]["maxDrawdownR"], result["riskOffStats"]["maxDrawdownR"]],
        ["Best R", result["stats"]["bestR"], result["riskOffStats"]["bestR"]],
        ["Worst R", result["stats"]["worstR"], result["riskOffStats"]["worstR"]],
    ]
    for i, row in enumerate(rows, 4):
        rowset(ws, i, row)

    headers = ["Symbol", "No", "Side", "Signal Date UTC+7", "Entry Date UTC+7", "Entry Price", "Initial Stop", "Final Stop", "Risk / Unit", "TP1", "TP1 Date", "Exit Date UTC+7", "Exit Price", "Exit Reason", "Base R", "Risk-Off R", "Size Mult", "Pre DD R", "Post DD R", "ATR14", "RSI14", "Distance EMA50 ATR", "Notes"]
    for sheet, subset in [
        ("Trades", result["riskOffTrades"]),
        ("BTC", [t for t in result["riskOffTrades"] if t["symbol"] == "BTCUSDT"]),
        ("SOL", [t for t in result["riskOffTrades"] if t["symbol"] == "SOLUSDT"]),
        ("SUI", [t for t in result["riskOffTrades"] if t["symbol"] == "SUIUSDT"]),
    ]:
        ws = wb[sheet]
        ws["A1"] = f"{sheet} - UTC+7 Trades" if sheet != "Trades" else "Detailed Trades - UTC+7"
        ws["A2"] = "One completed trade per row."
        rowset(ws, 4, headers)
        for i, t in enumerate(subset, 5):
            rowset(ws, i, [t["symbol"].replace("USDT", ""), t["tradeNo"], t["side"], t["signalTime"], t["entryTime"], t["entryPrice"], t["initialStop"], t["finalStop"], t["riskPerUnit"], t["tp1"], t["tp1Time"], t["exitTime"], t["exitPrice"], t["exitReason"], t["rMultiple"], t["riskOffR"], t["sizeMultiplier"], t["preTradeDrawdownR"], t["postTradeDrawdownR"], t["atr14"], t["rsi14"], t["distanceToEma50Atr"], t["notes"]])

    ass = wb["Assumptions"]
    ass["A1"] = "Assumptions"
    rowset(ass, 4, ["#", "Assumption"])
    for i, line in enumerate(result["assumptions"], 5):
        rowset(ass, i, [i - 4, line])
    quality = wb["Data Quality"]
    quality["A1"] = "Data Quality"
    rowset(quality, 4, ["Symbol", "1H Rows", "UTC+7 Daily Candles", "First UTC+7 Day", "Last UTC+7 Day", "Source"])
    for i, (sym, q) in enumerate(result["datasets"].items(), 5):
        rowset(quality, i, [sym.replace("USDT", ""), q["hourlyRows"], q["dailyRows"], q["firstDay"], q["lastDay"], q["source"]])
    for sheet in ["Summary", "Trades", "BTC", "SOL", "SUI", "Assumptions", "Data Quality", "20K Account", "Equity Curve"]:
        if sheet in wb.sheetnames and sheet in tpl.sheetnames:
            copy_layout(tpl[sheet], wb[sheet])
    wb.save(OUT_XLSX)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    datasets = {}
    all_trades = []
    for symbol in SYMBOLS:
        h = fetch_1h(symbol)
        d = enrich(resample_utc7(h))
        datasets[symbol] = {"hourlyRows": len(h), "dailyRows": len(d), "firstDay": d[0]["localDate"], "lastDay": d[-1]["localDate"], "source": "Binance spot 1H klines resampled to UTC+7 day"}
        all_trades.extend(backtest_symbol(symbol, d))
    all_trades.sort(key=lambda x: x["exitTime"])
    ro = riskoff(all_trades)
    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "systemVersion": "NXT v3.1 Runner A + Risk-Off, UTC+7 daily candles",
        "period": {"start": START_DATE.isoformat(), "end": (END_DATE - timedelta(days=1)).isoformat(), "timezone": "Asia/Saigon UTC+7"},
        "symbols": SYMBOLS,
        "stats": stats(all_trades),
        "riskOffStats": stats(ro, "riskOffR"),
        "trades": all_trades,
        "riskOffTrades": ro,
        "datasets": datasets,
        "assumptions": [
            "Daily candles are resampled from Binance spot 1H klines using Asia/Saigon UTC+7 calendar days.",
            "Primary NXT v3.1 rules are unchanged: SSL flip, EMA20 cross within 3 candles, RSI 50/50, EMA50 distance <= 2 ATR.",
            "Runner A exit is unchanged: close 50% at 2.5 ATR, remaining 50% exits on opposite SSL flip or breakeven stop.",
            "Risk-off overlay is unchanged: if closed-trade equity curve drawdown <= -4R, trade is counted at 40% size.",
            "Cost model remains 0.06% fee and 0.05% slippage per side; funding is not included in this rerun.",
        ],
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    build_workbook(result)
    case = [t for t in ro if t["symbol"] == "BTCUSDT" and ("2022-08" in t["signalTime"] or "2022-08" in t["entryTime"]) and t["side"] == "SHORT"]
    print(json.dumps({"outJson": str(OUT_JSON), "outXlsx": str(OUT_XLSX), "stats": result["stats"], "riskOffStats": result["riskOffStats"], "btcAug2022Shorts": case}, indent=2))


if __name__ == "__main__":
    main()
