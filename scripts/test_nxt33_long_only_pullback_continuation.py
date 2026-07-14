from __future__ import annotations

import json
import sys
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

import backtest_nxt31_utc7_latest as base
import backtest_nxt32_native_1d_latest as native
from test_nxt33_ssl14 import enrich_with_ssl_period


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
OUT_DIR = ROOT / "outputs" / "nxt33_long_only_pullback_continuation"
OUT_JSON = OUT_DIR / "nxt33_long_only_pullback_continuation_results.json"
OUT_XLSX = OUT_DIR / "NXT33_Long_Only_Pullback_Continuation_6Y_BTC_SOL_SUI_20K.xlsx"
BASELINE_JSON = ROOT / "latest" / "NXT_Latest_NXT33_Native1D_AntiReversal_RunnerA_NoContinuation_NoRiskOff_6Y_BTC_SOL_SUI_20K.json"


RULE = {
    "key": "long_only_pullback_touch_ema20",
    "name": "LONG-only pullback/touch EMA20 continuation",
    "touchLookback": 5,
}
TP1_ATR = 2.5
EARLY_BE_PROFIT_PCT = 0.07
ANTI_REVERSAL_MIN_RUNNER_R = 0.50
CONTINUATION_REQUIRE_SSL_FLIP = True
ENABLE_SHORT_CONTINUATION = False


def clear(ws, rows=1200, cols=40):
    for row in ws.iter_rows(min_row=1, max_row=max(rows, ws.max_row), min_col=1, max_col=max(cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def rowset(ws, r, values):
    for c, v in enumerate(values, 1):
        ws.cell(r, c).value = v


def style_table(ws, header_row=4):
    ws.freeze_panes = f"A{header_row + 1}"
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = min(max(ws.column_dimensions[letter].width or 12, 12), 24)


def touch_reclaim_long(candles: list[dict], i: int, lookback: int) -> bool:
    start = max(1, i - lookback + 1)
    touched = any(candles[j]["low"] <= candles[j]["ema20"] for j in range(start, i + 1) if candles[j]["ema20"] is not None)
    return touched and candles[i]["close"] > candles[i]["ema20"] and candles[i]["close"] > candles[i - 1]["close"]


def touch_reject_short(candles: list[dict], i: int, lookback: int) -> bool:
    start = max(1, i - lookback + 1)
    touched = any(candles[j]["high"] >= candles[j]["ema20"] for j in range(start, i + 1) if candles[j]["ema20"] is not None)
    return touched and candles[i]["close"] < candles[i]["ema20"] and candles[i]["close"] < candles[i - 1]["close"]


def profit_factor(rows: list[dict]) -> float | None:
    gross_profit = sum(t["rMultiple"] for t in rows if t["rMultiple"] > 0)
    gross_loss = -sum(t["rMultiple"] for t in rows if t["rMultiple"] < 0)
    return gross_profit / gross_loss if gross_loss else None


def enriched_stats(rows: list[dict]) -> dict:
    st = base.stats(rows)
    st["profitFactor"] = profit_factor(rows)
    st["ending20k"] = 20000 + st["totalR"] * 1000
    return st


def backtest_symbol(symbol: str, candles: list[dict]) -> list[dict]:
    trades, pos, n = [], None, 1
    last_profitable_runner_exit = None
    for i in range(55, len(candles) - 1):
        c, prev, nxt = candles[i], candles[i - 1], candles[i + 1]
        next_date = base.date.fromisoformat(nxt["localDate"])
        if next_date < native.START_DATE or next_date >= native.END_DATE:
            continue
        if pos:
            side = pos["side"]
            ssl_flip = (side == "LONG" and prev["ssl"] == 1 and c["ssl"] == -1) or (side == "SHORT" and prev["ssl"] == -1 and c["ssl"] == 1)
            can_trigger_early_be = c["localDate"] != pos["entryDate"]
            exit_price = reason = None
            if side == "LONG":
                if c["low"] <= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if (pos["triggered"] or pos["earlyBeTriggered"]) else "Stop loss"
                else:
                    if not pos["triggered"] and c["high"] >= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["tp"] - pos["entry"]) / pos["risk"])
                    if can_trigger_early_be and not pos["triggered"] and not pos["earlyBeTriggered"] and EARLY_BE_PROFIT_PCT is not None and c["high"] >= pos["entry"] * (1 + EARLY_BE_PROFIT_PCT):
                        pos["earlyBeTriggered"] = True
                        pos["earlyBeTime"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bearish flip"
            else:
                if c["high"] >= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if (pos["triggered"] or pos["earlyBeTriggered"]) else "Stop loss"
                else:
                    if not pos["triggered"] and c["low"] <= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["entry"] - pos["tp"]) / pos["risk"])
                    if can_trigger_early_be and not pos["triggered"] and not pos["earlyBeTriggered"] and EARLY_BE_PROFIT_PCT is not None and c["low"] <= pos["entry"] * (1 - EARLY_BE_PROFIT_PCT):
                        pos["earlyBeTriggered"] = True
                        pos["earlyBeTime"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bullish flip"
            if exit_price is not None:
                rem = 0.5 if pos["triggered"] else 1.0
                rem_r = (exit_price - pos["entry"]) / pos["risk"] if side == "LONG" else (pos["entry"] - exit_price) / pos["risk"]
                gross = pos["realizedR"] + rem * rem_r
                cost = base.cost_r(pos["entry"], pos["risk"])
                net = gross - cost
                trades.append({
                    "symbol": symbol,
                    "tradeNo": n,
                    "signalType": pos["signalType"],
                    "side": side,
                    "signalTime": pos["signalDate"],
                    "entryTime": pos["entryDate"],
                    "entryPrice": pos["entry"],
                    "initialStop": pos["initialStop"],
                    "finalStop": pos["stop"],
                    "riskPerUnit": pos["risk"],
                    "tp1": pos["tp"],
                    "tp1Time": pos["tp1Time"],
                    "earlyBeTriggered": pos["earlyBeTriggered"],
                    "earlyBeTime": pos["earlyBeTime"],
                    "exitTime": c["localDate"],
                    "exitPrice": exit_price,
                    "exitReason": reason,
                    "grossRMultiple": gross,
                    "costR": cost,
                    "rMultiple": net,
                    "atr14": pos["atr14"],
                    "rsi14": pos["rsi14"],
                    "distanceToEma50Atr": pos["distance"],
                    "ema20": pos["ema20"],
                    "ema50": pos["ema50"],
                    "notes": pos["notes"],
                })
                if net >= ANTI_REVERSAL_MIN_RUNNER_R and reason.startswith("Runner exit"):
                    last_profitable_runner_exit = {"index": i, "side": side, "netR": net}
                n += 1
                pos = None
            if pos:
                continue

        if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]) or prev["ssl"] is None:
            continue
        dist = abs(c["close"] - c["ema50"]) / c["atr14"]
        long_primary = prev["ssl"] == -1 and c["ssl"] == 1 and base.recent_cross(candles, i, "LONG") and dist <= 2 and c["rsi14"] > 50
        short_primary = prev["ssl"] == 1 and c["ssl"] == -1 and base.recent_cross(candles, i, "SHORT") and dist <= 2 and c["rsi14"] < 50
        continuation_ssl_ok = (prev["ssl"] == -1 and c["ssl"] == 1) if CONTINUATION_REQUIRE_SSL_FLIP else c["ssl"] == 1
        long_cont = continuation_ssl_ok and c["close"] > c["ema20"] > c["ema50"] and touch_reclaim_long(candles, i, RULE["touchLookback"])
        short_cont_ssl_ok = (prev["ssl"] == 1 and c["ssl"] == -1) if CONTINUATION_REQUIRE_SSL_FLIP else c["ssl"] == -1
        short_cont = ENABLE_SHORT_CONTINUATION and short_cont_ssl_ok and c["close"] < c["ema20"] < c["ema50"] and touch_reject_short(candles, i, RULE["touchLookback"])
        if last_profitable_runner_exit and i - last_profitable_runner_exit["index"] <= 1:
            if (long_primary or long_cont) and last_profitable_runner_exit["side"] == "SHORT":
                long_primary = long_cont = False
            if (short_primary or short_cont) and last_profitable_runner_exit["side"] == "LONG":
                short_primary = short_cont = False
        if not (long_primary or short_primary or long_cont or short_cont):
            continue
        side = "LONG" if (long_primary or long_cont) else "SHORT"
        signal_type = "Continuation" if ((long_cont and not long_primary) or (short_cont and not short_primary)) else "Primary"
        risk = c["atr14"] * 1.5
        entry = nxt["open"]
        pos = {
            "side": side,
            "signalType": signal_type,
            "signalDate": c["localDate"],
            "entryDate": nxt["localDate"],
            "entry": entry,
            "initialStop": entry - risk if side == "LONG" else entry + risk,
            "stop": entry - risk if side == "LONG" else entry + risk,
            "risk": risk,
            "tp": entry + c["atr14"] * TP1_ATR if side == "LONG" else entry - c["atr14"] * TP1_ATR,
            "triggered": False,
            "earlyBeTriggered": False,
            "earlyBeTime": "",
            "tp1Time": "",
            "realizedR": 0.0,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": dist,
            "ema20": c["ema20"],
            "ema50": c["ema50"],
            "notes": "Primary NXT v3.5" if signal_type == "Primary" else (RULE["name"] if side == "LONG" else "SHORT pullback/touch EMA20 continuation"),
        }
    return trades


def group_stats(rows: list[dict], key_fn) -> list[dict]:
    groups = {}
    for t in rows:
        groups.setdefault(key_fn(t), []).append(t)
    out = []
    for key, subset in sorted(groups.items()):
        st = enriched_stats(subset)
        st["group"] = key
        out.append(st)
    return out


def trade_key(t: dict) -> tuple:
    return (t["symbol"], t["side"], t["signalTime"], t["entryTime"])


def build_workbook(result: dict) -> None:
    wb = load_workbook(TEMPLATE)
    for ws in wb.worksheets:
        clear(ws)

    summary = wb["Summary"]
    summary["A1"] = "NXT v3.3 LONG-only Pullback Continuation Test"
    summary["A2"] = "Adds only LONG continuation trades using SSL bullish trend, EMA20 pullback/touch, and EMA20 reclaim. Latest files are not overwritten."
    rowset(summary, 4, ["Metric", "Baseline", "LONG-only Continuation", "Delta"])
    metrics = [
        ("Trades", "trades"),
        ("Win Rate", "winRate"),
        ("Total R", "totalR"),
        ("Average R", "avgR"),
        ("Max DD R", "maxDrawdownR"),
        ("Best R", "bestR"),
        ("Worst R", "worstR"),
        ("Profit Factor", "profitFactor"),
        ("20K Ending", "ending20k"),
    ]
    for r, (label, key) in enumerate(metrics, 5):
        b = result["baselineStats"][key]
        v = result["variantStats"][key]
        rowset(summary, r, [label, b, v, v - b if isinstance(v, (int, float)) and isinstance(b, (int, float)) else None])
    rowset(summary, 16, ["Continuation Trades", result["continuationStats"]["trades"]])
    rowset(summary, 17, ["Continuation R", result["continuationStats"]["totalR"]])
    rowset(summary, 18, ["Continuation Win Rate", result["continuationStats"]["winRate"]])
    rowset(summary, 19, ["Rule", "SSL bullish; close > EMA20 > EMA50; low touched EMA20 in last 5 bars; close > EMA20; close > prior close. No RSI, distance, or EMA50 slope filter."])
    style_table(summary)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["D"].width = 16

    headers = [
        "Symbol", "No", "Signal Type", "Side", "Signal Date", "Entry Date", "Entry Price",
        "Initial Stop", "Final Stop", "Risk / Unit", "TP1", "TP1 Date", "Exit Date",
        "Exit Price", "Exit Reason", "R", "ATR14", "RSI14", "Distance EMA50 ATR",
        "EMA20", "EMA50", "Notes",
    ]
    trades = wb["Trades"]
    trades["A1"] = "Detailed Trades - LONG-only Pullback Continuation"
    rowset(trades, 4, headers)
    for i, t in enumerate(result["trades"], 5):
        rowset(trades, i, [
            t["symbol"].replace("USDT", ""), t["tradeNo"], t["signalType"], t["side"],
            t["signalTime"], t["entryTime"], t["entryPrice"], t["initialStop"], t["finalStop"],
            t["riskPerUnit"], t["tp1"], t["tp1Time"], t["exitTime"], t["exitPrice"],
            t["exitReason"], t["rMultiple"], t["atr14"], t["rsi14"], t["distanceToEma50Atr"],
            t["ema20"], t["ema50"], t["notes"],
        ])
    style_table(trades)
    trades.column_dimensions["V"].width = 42

    btc = wb["BTC"]
    btc.title = "Continuation Trades"
    btc["A1"] = "Continuation Trades Only"
    rowset(btc, 4, headers)
    for i, t in enumerate(result["continuationTrades"], 5):
        rowset(btc, i, [
            t["symbol"].replace("USDT", ""), t["tradeNo"], t["signalType"], t["side"],
            t["signalTime"], t["entryTime"], t["entryPrice"], t["initialStop"], t["finalStop"],
            t["riskPerUnit"], t["tp1"], t["tp1Time"], t["exitTime"], t["exitPrice"],
            t["exitReason"], t["rMultiple"], t["atr14"], t["rsi14"], t["distanceToEma50Atr"],
            t["ema20"], t["ema50"], t["notes"],
        ])
    style_table(btc)
    btc.column_dimensions["V"].width = 42

    sol = wb["SOL"]
    sol.title = "By Symbol"
    sol["A1"] = "Stats By Symbol"
    rowset(sol, 4, ["Symbol", "Trades", "Win Rate", "Total R", "Avg R", "Max DD R", "Best R", "Worst R", "Profit Factor"])
    for i, row in enumerate(result["bySymbol"], 5):
        rowset(sol, i, [row["group"].replace("USDT", ""), row["trades"], row["winRate"], row["totalR"], row["avgR"], row["maxDrawdownR"], row["bestR"], row["worstR"], row["profitFactor"]])
    style_table(sol)

    sui = wb["SUI"]
    sui.title = "By Year"
    sui["A1"] = "Stats By Exit Year"
    rowset(sui, 4, ["Year", "Trades", "Win Rate", "Total R", "Avg R", "Max DD R", "Best R", "Worst R", "Profit Factor"])
    for i, row in enumerate(result["byYear"], 5):
        rowset(sui, i, [row["group"], row["trades"], row["winRate"], row["totalR"], row["avgR"], row["maxDrawdownR"], row["bestR"], row["worstR"], row["profitFactor"]])
    style_table(sui)

    account = wb["20K Account"]
    account["A1"] = "20K Account"
    rowset(account, 4, ["Starting Equity", "Baseline Ending", "Variant Ending", "Delta"])
    rowset(account, 5, [20000, result["baselineStats"]["ending20k"], result["variantStats"]["ending20k"], result["variantStats"]["ending20k"] - result["baselineStats"]["ending20k"]])
    style_table(account)

    ass = wb["Assumptions"]
    ass["A1"] = "Assumptions"
    rowset(ass, 4, ["#", "Assumption"])
    for i, line in enumerate(result["assumptions"], 5):
        rowset(ass, i, [i - 4, line])
    style_table(ass)
    ass.column_dimensions["B"].width = 72

    dq = wb["Data Quality"]
    dq["A1"] = "Data Quality"
    rowset(dq, 4, ["Symbol", "Daily Rows", "First Day", "Last Day", "Source"])
    for i, (sym, q) in enumerate(result["datasets"].items(), 5):
        rowset(dq, i, [sym.replace("USDT", ""), q["dailyRows"], q["firstDay"], q["lastDay"], q["source"]])
    style_table(dq)

    wb.save(OUT_XLSX)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    baseline = json.loads(BASELINE_JSON.read_text(encoding="utf-8"))
    baseline_rows = baseline["trades"]
    baseline_stats = enriched_stats(baseline_rows)

    all_trades = []
    datasets = {}
    for symbol in native.SYMBOLS:
        candles = enrich_with_ssl_period(native.fetch_native_1d(symbol), 14)
        datasets[symbol] = {
            "dailyRows": len(candles),
            "firstDay": candles[0]["localDate"],
            "lastDay": candles[-1]["localDate"],
            "source": "Binance spot native 1D klines",
        }
        all_trades.extend(backtest_symbol(symbol, candles))
    all_trades.sort(key=lambda x: x["exitTime"])

    baseline_keys = {trade_key(t) for t in baseline_rows}
    continuation = [t for t in all_trades if t["signalType"] == "Continuation"]
    added = [t for t in all_trades if trade_key(t) not in baseline_keys]
    variant_stats = enriched_stats(all_trades)
    continuation_stats = enriched_stats(continuation)
    added_stats = enriched_stats(added)

    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "systemVersion": "NXT v3.3 + LONG-only pullback/touch EMA20 continuation",
        "period": baseline["period"],
        "rule": RULE,
        "baselineStats": baseline_stats,
        "variantStats": variant_stats,
        "continuationStats": continuation_stats,
        "addedStats": added_stats,
        "trades": all_trades,
        "continuationTrades": continuation,
        "addedTrades": added,
        "bySymbol": group_stats(all_trades, lambda t: t["symbol"]),
        "byContinuationSymbol": group_stats(continuation, lambda t: t["symbol"]),
        "byYear": group_stats(all_trades, lambda t: t["exitTime"][:4]),
        "byContinuationYear": group_stats(continuation, lambda t: t["exitTime"][:4]),
        "datasets": datasets,
        "assumptions": [
            "Baseline is current latest NXT v3.3 with Binance native 1D candles, SSL14, Runner A, anti-immediate-reversal, no continuation, and no risk-off.",
            "Primary LONG/SHORT rules are unchanged from latest NXT v3.3.",
            "Continuation is LONG-only.",
            "Continuation LONG requires SSL already bullish, close > EMA20 > EMA50, a low touching EMA20 within the last 5 candles, close > EMA20, and close > previous close.",
            "Continuation does not require RSI, distance-to-EMA50, or EMA50 slope filters.",
            "Only one position per symbol is open at a time; continuation is skipped while an existing trade is open.",
            "Entry remains next daily open after signal close; stop, TP1, Runner A exit, cost model, and anti-immediate-reversal rule are unchanged.",
            "Latest production files were not overwritten.",
        ],
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    build_workbook(result)
    print(json.dumps({
        "outJson": str(OUT_JSON),
        "outXlsx": str(OUT_XLSX),
        "baselineStats": baseline_stats,
        "variantStats": variant_stats,
        "continuationStats": continuation_stats,
        "addedStats": added_stats,
        "byContinuationSymbol": result["byContinuationSymbol"],
        "byContinuationYear": result["byContinuationYear"],
    }, indent=2))


if __name__ == "__main__":
    main()
