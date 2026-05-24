from __future__ import annotations

import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"D:\Workspace\Codex\Crypto trading")
LATEST = ROOT / "latest"
XLSX = LATEST / "NXT_Latest_NXT32_UTC7_RunnerA_Continuation_NoRiskOff_6Y_BTC_SOL_SUI_20K.xlsx"
JSON_PATH = LATEST / "NXT_Latest_NXT32_UTC7_RunnerA_Continuation_NoRiskOff_6Y_BTC_SOL_SUI_20K.json"
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
STARTING_EQUITY = 20_000.0
RISK_PCT = 0.02


def copy_layout(src, dst):
    dst.sheet_format.defaultColWidth = src.sheet_format.defaultColWidth
    dst.sheet_format.defaultRowHeight = src.sheet_format.defaultRowHeight
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    for c in range(1, max(src.max_column, dst.max_column) + 1):
        letter = get_column_letter(c)
        if src.column_dimensions.get(letter) and src.column_dimensions[letter].width:
            dst.column_dimensions[letter].width = src.column_dimensions[letter].width
    for r in range(1, max(src.max_row, dst.max_row) + 1):
        sr = r if r <= src.max_row else 5
        if src.row_dimensions.get(sr) and src.row_dimensions[sr].height:
            dst.row_dimensions[r].height = src.row_dimensions[sr].height
        for c in range(1, max(src.max_column, dst.max_column) + 1):
            a = src.cell(sr, min(c, src.max_column))
            b = dst.cell(r, c)
            if a.has_style:
                b._style = copy(a._style)
            b.number_format = a.number_format
            b.alignment = copy(a.alignment)
            b.font = copy(a.font)
            b.fill = copy(a.fill)
            b.border = copy(a.border)


def clear_values(ws, rows=260, cols=12):
    for row in ws.iter_rows(min_row=1, max_row=max(rows, ws.max_row), min_col=1, max_col=max(cols, ws.max_column)):
        for cell in row:
            if cell.__class__.__name__ != "MergedCell":
                cell.value = None


def rowset(ws, r, values):
    for c, v in enumerate(values, 1):
        ws.cell(r, c).value = v


def main():
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    trades = sorted(data["trades"], key=lambda t: t["exitTime"])
    wb = load_workbook(XLSX)
    tpl = load_workbook(TEMPLATE)
    ws = wb["20K Account"]
    clear_values(ws)
    ws["A1"] = "20K Account - NXT v3.2 No Risk-Off"
    ws["A2"] = "Compounded sequence at 2.0% risk per trade; leverage and funding are not modeled."
    headers = ["Trade", "Exit Date", "Symbol", "Signal Type", "Side", "Net R", "Equity Before", "Risk USD", "P/L USD", "Equity After", "Drawdown"]
    rowset(ws, 4, headers)
    equity = STARTING_EQUITY
    peak = STARTING_EQUITY
    max_dd = 0.0
    for i, t in enumerate(trades, 5):
        before = equity
        risk = before * RISK_PCT
        pnl = risk * t["rMultiple"]
        equity += pnl
        peak = max(peak, equity)
        dd = equity / peak - 1
        max_dd = min(max_dd, dd)
        rowset(ws, i, [i - 4, t["exitTime"], t["symbol"].replace("USDT", ""), t.get("signalType", ""), t["side"], t["rMultiple"], before, risk, pnl, equity, dd])

    copy_layout(tpl["20K Account"], ws)
    # Re-apply useful number formats after template style copy.
    last = len(trades) + 4
    for row in range(5, last + 1):
        ws.cell(row, 6).number_format = "0.00"
        for col in range(7, 10 + 1):
            ws.cell(row, col).number_format = "$#,##0"
        ws.cell(row, 11).number_format = "0.0%"
    wb.save(XLSX)
    print({"finalEquity": equity, "compoundReturn": equity / STARTING_EQUITY - 1, "maxDrawdownPct": max_dd, "rows": len(trades)})


if __name__ == "__main__":
    main()
