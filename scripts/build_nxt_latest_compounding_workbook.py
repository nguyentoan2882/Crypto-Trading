from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import Reference


ROOT = Path(__file__).resolve().parents[1]
LATEST_JSON = ROOT / "latest" / "NXT_Latest_NXT35_BTC_BNB_SOL_FundingAdjusted_20K.json"
TEMPLATE = ROOT / "templates" / "NXT_Backtest_Workbook_Template.xlsx"
OUT_DIR = ROOT / "outputs" / "nxt_latest_compounding"
OUT_XLSX = OUT_DIR / "NXT_Latest_Portfolio_Cap_6pct_BTC_BNB_SOL.xlsx"

STARTING_EQUITY = 20_000.0
FIXED_ONE_R = 1_000.0
SCENARIOS = [
    ("Fixed Latest 1R=$1,000", "fixed", None),
    ("Portfolio Cap 6% Equal", "portfolio_cap", {"BTCUSDT": 0.02, "BNBUSDT": 0.02, "SOLUSDT": 0.02}),
    ("Portfolio Cap 6% BTC-heavy", "portfolio_cap", {"BTCUSDT": 0.03, "BNBUSDT": 0.015, "SOLUSDT": 0.015}),
]


def max_drawdown(rows: list[dict], key: str) -> tuple[float, float]:
    peak = rows[0][key] if rows else STARTING_EQUITY
    max_dd = 0.0
    max_dd_pct = 0.0
    for row in rows:
        equity = row[key]
        peak = max(peak, equity)
        dd = equity - peak
        max_dd = min(max_dd, dd)
        max_dd_pct = min(max_dd_pct, dd / peak if peak else 0)
    return max_dd, max_dd_pct


def simulate(trades: list[dict], mode: str, allocation: dict[str, float] | None) -> list[dict]:
    events = defaultdict(lambda: {"entries": [], "exits": []})
    for index, trade in enumerate(trades):
        events[trade["entryTime"]]["entries"].append((index, trade))
        events[trade["exitTime"]]["exits"].append((index, trade))

    equity = STARTING_EQUITY
    open_risk: dict[int, dict] = {}
    rows: list[dict] = []
    sequence = 1

    def current_open_risk(symbol: str | None = None) -> float:
        return sum(
            row["riskAmount"]
            for row in open_risk.values()
            if symbol is None or row["symbol"] == symbol
        )

    for event_date in sorted(events):
        same_day_indexes = {
            index
            for index, trade in events[event_date]["entries"]
            if trade["exitTime"] == event_date
        }
        regular_exits = [
            item for item in events[event_date]["exits"]
            if item[0] not in same_day_indexes
        ]
        same_day_exits = [
            item for item in events[event_date]["exits"]
            if item[0] in same_day_indexes
        ]

        def close_trade(index: int, trade: dict) -> None:
            nonlocal equity, sequence
            risk_record = open_risk.pop(index, None)
            risk_amount = risk_record["riskAmount"] if risk_record else 0.0
            pnl = risk_amount * float(trade["netRAfterFunding"])
            equity += pnl
            rows.append(
                {
                    "seq": sequence,
                    "symbol": trade["symbol"],
                    "tradeNo": trade["tradeNo"],
                    "signalType": trade["signalType"],
                    "side": trade["side"],
                    "entryTime": trade["entryTime"],
                    "exitTime": trade["exitTime"],
                    "r": float(trade["netRAfterFunding"]),
                    "riskAmount": risk_amount,
                    "pnl": pnl,
                    "equity": equity,
                    "skipped": risk_record is None,
                }
            )
            sequence += 1

        for index, trade in sorted(regular_exits, key=lambda item: (item[1]["symbol"], item[1]["tradeNo"])):
            close_trade(index, trade)
        for index, trade in sorted(events[event_date]["entries"], key=lambda item: (item[1]["symbol"], item[1]["tradeNo"])):
            if mode == "fixed":
                risk_amount = FIXED_ONE_R
            else:
                assert allocation is not None
                total_cap = sum(allocation.values())
                symbol = trade["symbol"]
                symbol_capacity = max(0.0, equity * allocation[symbol] - current_open_risk(symbol))
                portfolio_capacity = max(0.0, equity * total_cap - current_open_risk())
                risk_amount = min(symbol_capacity, portfolio_capacity)
            if risk_amount > 0:
                open_risk[index] = {"symbol": trade["symbol"], "riskAmount": risk_amount}
        for index, trade in sorted(same_day_exits, key=lambda item: (item[1]["symbol"], item[1]["tradeNo"])):
            close_trade(index, trade)

    return rows


def yearly(rows: list[dict]) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for row in rows:
        year = row["exitTime"][:4]
        rec = out.setdefault(year, {"trades": 0, "pnl": 0.0, "end": STARTING_EQUITY})
        rec["trades"] += 1
        rec["pnl"] += row["pnl"]
        rec["end"] = row["equity"]
    return out


def add_table_rows(ws, start_row: int, rows: list[list]) -> None:
    for r, row in enumerate(rows, start_row):
        for c, value in enumerate(row, 1):
            ws.cell(r, c).value = value


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    data = json.loads(LATEST_JSON.read_text(encoding="utf-8"))
    trades = sorted(data["trades"], key=lambda t: (t["entryTime"], t["exitTime"], t["symbol"], t["tradeNo"]))
    scenario_rows = {name: simulate(trades, mode, allocation) for name, mode, allocation in SCENARIOS}

    wb = load_workbook(TEMPLATE)
    main_summary = wb["Summary"]
    main_summary["A1"] = "NXT Latest Portfolio Risk Cap Comparison"
    main_summary["A2"] = "BTC/BNB/SOL latest, funding-adjusted net R, comparing fixed 1R with portfolio-cap 6% allocation scenarios."

    ws = wb["Compounding Summary"]
    ws["A1"] = "NXT Latest Compounding Comparison"
    ws["A2"] = "BTC/BNB/SOL latest, funding-adjusted net R. Portfolio-cap scenarios size each trade at entryTime, then realize P&L at exitTime."
    summary_rows = [["Scenario", "Risk Model", "Ending Equity", "Net Profit", "Max DD $", "Max DD % Peak", "Executed Trades", "Skipped Trades"]]
    for name, mode, allocation in SCENARIOS:
        rows = scenario_rows[name]
        dd, dd_pct = max_drawdown(rows, "equity")
        executed = sum(1 for row in rows if row["riskAmount"] > 0)
        skipped = sum(1 for row in rows if row["skipped"])
        summary_rows.append(
            [
                name,
                "Fixed $1,000/R" if mode == "fixed" else ", ".join(f"{symbol.replace('USDT', '')} {pct:.1%}" for symbol, pct in allocation.items()),
                rows[-1]["equity"],
                rows[-1]["equity"] - STARTING_EQUITY,
                dd,
                dd_pct,
                executed,
                skipped,
            ]
        )
    add_table_rows(ws, 4, summary_rows)

    ys = wb["Compounding Yearly"]
    years = sorted({row["exitTime"][:4] for rows in scenario_rows.values() for row in rows})
    headers = ["Year"]
    for name, _, _ in SCENARIOS:
        headers.extend([f"{name} P&L", f"{name} End Equity"])
    add_table_rows(ys, 4, [headers])
    for r, year in enumerate(years, 5):
        values = [year]
        for name, _, _ in SCENARIOS:
            rec = yearly(scenario_rows[name]).get(year, {"pnl": 0.0, "end": None})
            values.extend([rec["pnl"], rec["end"]])
        add_table_rows(ys, r, [values])
    if ys._charts:
        chart = ys._charts[0]
        chart.set_categories(Reference(ys, min_col=1, min_row=5, max_row=4 + len(years)))

    detail = wb["Compounding Trades"]
    detail_headers = [
        "Seq",
        "Symbol",
        "Trade #",
        "Signal Type",
        "Side",
        "Entry Date",
        "Exit Date",
        "Net R",
        "Fixed Risk",
        "Fixed P&L",
        "Fixed Equity",
        "Cap Equal Risk",
        "Cap Equal P&L",
        "Cap Equal Equity",
        "Cap BTC-heavy Risk",
        "Cap BTC-heavy P&L",
        "Cap BTC-heavy Equity",
    ]
    add_table_rows(detail, 4, [detail_headers])
    fixed = scenario_rows["Fixed Latest 1R=$1,000"]
    equal = scenario_rows["Portfolio Cap 6% Equal"]
    btc_heavy = scenario_rows["Portfolio Cap 6% BTC-heavy"]
    rows = []
    for a, b, c in zip(fixed, equal, btc_heavy):
        rows.append(
            [
                a["seq"],
                a["symbol"],
                a["tradeNo"],
                a["signalType"],
                a["side"],
                a["entryTime"],
                a["exitTime"],
                a["r"],
                a["riskAmount"],
                a["pnl"],
                a["equity"],
                b["riskAmount"],
                b["pnl"],
                b["equity"],
                c["riskAmount"],
                c["pnl"],
                c["equity"],
            ]
        )
    add_table_rows(detail, 5, rows)

    assumptions = wb["Assumptions"]
    add_table_rows(
        assumptions,
        4,
        [
            ["#", "Assumption"],
            [1, "Source is latest/NXT_Latest_NXT34_BTC_BNB_SOL_FundingAdjusted_20K.json."],
            [2, "R uses netRAfterFunding, so trading cost and funding are included."],
            [3, "Fixed Latest matches the current latest workbook model: 1R = $1,000 on a $20,000 starting account."],
            [4, "Portfolio-cap scenarios calculate risk amount at entryTime from realized account equity at that moment."],
            [5, "P&L is realized at exitTime. If exits and entries occur on the same date, exits are processed before entries."],
            [6, "Open-position unrealized P&L is not marked to market between entry and exit."],
            [7, "Portfolio Cap 6% Equal allocates BTC 2%, BNB 2%, and SOL 2% of equity."],
            [8, "Portfolio Cap 6% BTC-heavy allocates BTC 3%, BNB 1.5%, and SOL 1.5% of equity."],
        ],
    )
    for sheet in wb.worksheets:
        if sheet.freeze_panes is None:
            for selection in sheet.sheet_view.selection:
                selection.pane = None
    wb.save(OUT_XLSX)
    print(json.dumps({"workbook": str(OUT_XLSX), "summary": summary_rows[1:]}, indent=2))


if __name__ == "__main__":
    main()
