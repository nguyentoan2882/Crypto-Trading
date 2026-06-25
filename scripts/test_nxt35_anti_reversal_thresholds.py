from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


sys.path.insert(0, str(Path(__file__).resolve().parent))

import audit_nxt34_btc_bnb_sol_funding_adjusted as funding
import backtest_nxt31_utc7_latest as base
import promote_nxt34_btc_bnb_sol_latest as latest
import test_nxt33_long_only_pullback_continuation as cont
from test_nxt33_ssl14 import enrich_with_ssl_period


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "outputs" / "nxt35_anti_reversal_threshold_test"
OUT_JSON = OUT_DIR / "nxt35_anti_reversal_threshold_test_results.json"
OUT_XLSX = OUT_DIR / "NXT35_Anti_Reversal_20_25_30pct_Comparison.xlsx"
THRESHOLDS = [None, 0.20, 0.25, 0.30]


def variant_name(threshold: float | None) -> str:
    return "Before 20% rule" if threshold is None else f"{threshold:.0%} threshold"


def favorable_reached(side: str, entry: float, candle: dict, threshold: float) -> bool:
    if side == "LONG":
        return candle["high"] >= entry * (1 + threshold)
    return candle["low"] <= entry * (1 - threshold)


def backtest_symbol(symbol: str, candles: list[dict], threshold: float | None) -> list[dict]:
    trades, pos, n = [], None, 1
    last_qualified_runner_exit = None
    for i in range(55, len(candles) - 1):
        c, prev, nxt = candles[i], candles[i - 1], candles[i + 1]
        next_date = base.date.fromisoformat(nxt["localDate"])
        if next_date < latest.START_DATE or next_date >= latest.END_DATE:
            continue

        if pos:
            side = pos["side"]
            if threshold is None:
                pos["qualified"] = True
            else:
                pos["qualified"] = pos["qualified"] or favorable_reached(side, pos["entry"], c, threshold)
            ssl_flip = (side == "LONG" and prev["ssl"] == 1 and c["ssl"] == -1) or (
                side == "SHORT" and prev["ssl"] == -1 and c["ssl"] == 1
            )
            exit_price = reason = None
            if side == "LONG":
                if c["low"] <= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["high"] >= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["tp"] - pos["entry"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bearish flip"
            else:
                if c["high"] >= pos["stop"]:
                    exit_price = pos["stop"]
                    reason = "Breakeven stop" if pos["triggered"] else "Stop loss"
                else:
                    if not pos["triggered"] and c["low"] <= pos["tp"]:
                        pos["triggered"] = True
                        pos["tp1Time"] = c["localDate"]
                        pos["stop"] = pos["entry"]
                        pos["realizedR"] += 0.5 * ((pos["entry"] - pos["tp"]) / pos["risk"])
                    if ssl_flip:
                        exit_price = c["close"]
                        reason = "Runner exit: SSL bullish flip"
            if exit_price is not None:
                remaining = 0.5 if pos["triggered"] else 1.0
                remaining_r = (exit_price - pos["entry"]) / pos["risk"] if side == "LONG" else (pos["entry"] - exit_price) / pos["risk"]
                gross = pos["realizedR"] + remaining * remaining_r
                cost = base.cost_r(pos["entry"], pos["risk"])
                net = gross - cost
                trades.append(
                    {
                        "symbol": symbol,
                        "tradeNo": n,
                        "signalType": pos["signalType"],
                        "side": side,
                        "signalTime": pos["signalDate"],
                        "entryTime": pos["entryDate"],
                        "entryPrice": pos["entry"],
                        "initialStop": pos["initialStop"],
                        "finalStop": pos["stop"],
                        "riskPerUnit": pos["risk"],
                        "tp1": pos["tp"],
                        "tp1Time": pos["tp1Time"],
                        "exitTime": c["localDate"],
                        "exitPrice": exit_price,
                        "exitReason": reason,
                        "grossRMultiple": gross,
                        "costR": cost,
                        "rMultiple": net,
                        "atr14": pos["atr14"],
                        "rsi14": pos["rsi14"],
                        "distanceToEma50Atr": pos["distance"],
                        "ema20": pos["ema20"],
                        "ema50": pos["ema50"],
                        "favorableThresholdReached": pos["qualified"],
                        "notes": variant_name(threshold),
                    }
                )
                qualifies_exit = pos["qualified"] and reason.startswith("Runner exit")
                if threshold is None:
                    qualifies_exit = qualifies_exit and net > 0
                if qualifies_exit:
                    last_qualified_runner_exit = {"index": i, "side": side}
                n += 1
                pos = None
            if pos:
                continue

        if any(c[k] is None for k in ["ema20", "ema50", "atr14", "rsi14", "ssl"]) or prev["ssl"] is None:
            continue
        distance = abs(c["close"] - c["ema50"]) / c["atr14"]
        long_primary = prev["ssl"] == -1 and c["ssl"] == 1 and base.recent_cross(candles, i, "LONG") and distance <= 2 and c["rsi14"] > 50
        short_primary = prev["ssl"] == 1 and c["ssl"] == -1 and base.recent_cross(candles, i, "SHORT") and distance <= 2 and c["rsi14"] < 50
        long_cont = c["ssl"] == 1 and c["close"] > c["ema20"] > c["ema50"] and cont.touch_reclaim_long(candles, i, cont.RULE["touchLookback"])
        if last_qualified_runner_exit and i - last_qualified_runner_exit["index"] <= 1:
            if (long_primary or long_cont) and last_qualified_runner_exit["side"] == "SHORT":
                long_primary = long_cont = False
            if short_primary and last_qualified_runner_exit["side"] == "LONG":
                short_primary = False
        if not (long_primary or short_primary or long_cont):
            continue
        side = "LONG" if (long_primary or long_cont) else "SHORT"
        signal_type = "Continuation" if long_cont and not long_primary else "Primary"
        risk = c["atr14"] * 1.5
        entry = nxt["open"]
        pos = {
            "side": side,
            "signalType": signal_type,
            "signalDate": c["localDate"],
            "entryDate": nxt["localDate"],
            "entry": entry,
            "initialStop": entry - risk if side == "LONG" else entry + risk,
            "stop": entry - risk if side == "LONG" else entry + risk,
            "risk": risk,
            "tp": entry + c["atr14"] * 2.5 if side == "LONG" else entry - c["atr14"] * 2.5,
            "triggered": False,
            "tp1Time": "",
            "realizedR": 0.0,
            "qualified": threshold is None,
            "atr14": c["atr14"],
            "rsi14": c["rsi14"],
            "distance": distance,
            "ema20": c["ema20"],
            "ema50": c["ema50"],
        }
    return trades


def add_funding(trades: list[dict], funding_by_symbol: dict[str, list[dict]]) -> None:
    for trade in trades:
        trade.update(funding.funding_for_trade(trade, funding_by_symbol[trade["symbol"]]))
        trade["netRAfterFunding"] = trade["rMultiple"] + trade["fundingR"]


def stats_for(trades: list[dict]) -> dict:
    stats = funding.stats_for_key(trades, "netRAfterFunding")
    stats["endingFixed"] = 20_000 + stats["totalR"] * 1_000
    c2 = funding.compounding_curve(trades, 0.02, "netRAfterFunding")
    c5 = funding.compounding_curve(trades, 0.05, "netRAfterFunding")
    stats["compound2Ending"] = c2["endingEquity"]
    stats["compound2MaxDdPct"] = c2["maxDrawdownPct"]
    stats["compound5Ending"] = c5["endingEquity"]
    stats["compound5MaxDdPct"] = c5["maxDrawdownPct"]
    return stats


def trade_identity(trade: dict) -> tuple:
    return trade["symbol"], trade["side"], trade["signalTime"], trade["entryTime"]


def style_sheet(ws, header_row: int = 4) -> None:
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False
    for cell in ws[header_row]:
        if cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(max(ws.column_dimensions[get_column_letter(col)].width or 12, 12), 24)


def build_workbook(result: dict) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "NXT v3.5 Anti-Reversal Threshold Test"
    summary["A2"] = "BTC/BNB/SOL, funding-adjusted. Compares prior anti-reversal behavior with 20%, 25%, and 30% favorable-move thresholds."
    headers = [
        "Variant", "Trades", "Win Rate", "Total R", "Delta R vs Before", "Max DD R",
        "Profit Factor", "Fixed Ending", "Compound 2% Ending", "2% Max DD",
        "Compound 5% Ending", "5% Max DD",
    ]
    summary.append([])
    summary.append(headers)
    baseline_r = result["variants"]["Before 20% rule"]["stats"]["totalR"]
    for name in [variant_name(t) for t in THRESHOLDS]:
        stats = result["variants"][name]["stats"]
        summary.append(
            [
                name, stats["trades"], stats["winRate"], stats["totalR"], stats["totalR"] - baseline_r,
                stats["maxDrawdownR"], stats["profitFactor"], stats["endingFixed"],
                stats["compound2Ending"], stats["compound2MaxDdPct"],
                stats["compound5Ending"], stats["compound5MaxDdPct"],
            ]
        )
    style_sheet(summary)

    changes = wb.create_sheet("Changed Trades")
    change_headers = ["Variant", "Change", "Symbol", "Side", "Signal", "Entry", "Exit", "Adjusted R"]
    for col, value in enumerate(change_headers, 1):
        changes.cell(4, col).value = value
    baseline = result["variants"]["Before 20% rule"]["trades"]
    baseline_map = {trade_identity(t): t for t in baseline}
    row = 5
    for name in ["20% threshold", "25% threshold", "30% threshold"]:
        rows = result["variants"][name]["trades"]
        current = {trade_identity(t): t for t in rows}
        for identity in sorted(set(current) - set(baseline_map)):
            t = current[identity]
            values = [name, "Added", t["symbol"], t["side"], t["signalTime"], t["entryTime"], t["exitTime"], t["netRAfterFunding"]]
            for col, value in enumerate(values, 1):
                changes.cell(row, col).value = value
            row += 1
        for identity in sorted(set(baseline_map) - set(current)):
            t = baseline_map[identity]
            values = [name, "Removed", t["symbol"], t["side"], t["signalTime"], t["entryTime"], t["exitTime"], t["netRAfterFunding"]]
            for col, value in enumerate(values, 1):
                changes.cell(row, col).value = value
            row += 1
    style_sheet(changes)

    assumptions = wb.create_sheet("Assumptions")
    assumptions["A1"] = "Assumptions"
    assumption_rows = [
        "The before-rule variant blocks an opposite entry after every profitable runner SSL-flip exit, matching the prior baseline.",
        "Threshold variants only qualify the anti-reversal block when the prior position reached the specified favorable percentage from entry.",
        "All other NXT v3.5 rules are unchanged: native 1D, SSL14, primary LONG/SHORT, LONG-only continuation, Runner A, no risk-off.",
        "Trading cost and Binance USD-M funding are included.",
        "Compounding risk is locked from realized equity at entry and P&L is recognized at exit.",
    ]
    assumptions.cell(4, 1).value = "#"
    assumptions.cell(4, 2).value = "Assumption"
    for row, text in enumerate(assumption_rows, 5):
        assumptions.cell(row, 1).value = row - 4
        assumptions.cell(row, 2).value = text
    assumptions.column_dimensions["B"].width = 100
    style_sheet(assumptions)
    wb.save(OUT_XLSX)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    raw_by_symbol = {
        symbol: enrich_with_ssl_period(latest.fetch_binance_native_1d(symbol), 14)
        for symbol in latest.SYMBOLS
    }
    funding_by_symbol = {
        symbol: funding.fetch_monthly_funding(symbol, latest.START_DATE, latest.END_DATE)
        for symbol in latest.SYMBOLS
    }
    variants = {}
    for threshold in THRESHOLDS:
        name = variant_name(threshold)
        trades = []
        for symbol, candles in raw_by_symbol.items():
            trades.extend(backtest_symbol(symbol, [dict(c) for c in candles], threshold))
        trades.sort(key=lambda t: (t["exitTime"], t["symbol"], t["tradeNo"]))
        add_funding(trades, funding_by_symbol)
        variants[name] = {"threshold": threshold, "stats": stats_for(trades), "trades": trades}

    result = {
        "generatedAt": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "system": "NXT v3.5 anti-reversal favorable threshold test",
        "symbols": latest.SYMBOLS,
        "variants": variants,
    }
    OUT_JSON.write_text(json.dumps(result, indent=2), encoding="utf-8")
    build_workbook(result)
    print(json.dumps({
        "json": str(OUT_JSON),
        "workbook": str(OUT_XLSX),
        "summary": {name: payload["stats"] for name, payload in variants.items()},
    }, indent=2))


if __name__ == "__main__":
    main()
